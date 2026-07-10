# -*- coding: utf-8 -*-
# python _build_oci.py               (January, default)
# PNL_MONTH=feb python _build_oci.py (February) ... may
# FVTOCI OCI公允/OCI价差 workbook. Builds PnL_OCI_2026<Month>.xlsx. Fully formula-driven.
#   universe  : opening ∪ in-month trades (new positions INCLUDED) — from _oci_<m>.json
#   cost basis: moving WAVG pool; BUY fee capitalized into cost; SELL fee nets realized (FVTOCI)
#   OCI公允_d (unrealized) = ΔHoldFV_d   where HoldFV_d = qty_d×px_d − costpool_d
#   OCI价差_d (realized)   = sell_proceeds_d − sell_fee_d − sell_qty_d × wavg_before_d
#   identity : OCI公允_d + OCI价差_d = priceP&L_d − total_fee_d  (dividend tracked separately)
#   Jan      : per-stock×day reconciliation of our OCI公允/OCI价差 vs official 2026年1月(1).xlsx
#   Feb–May  : our own daily OCI grid (no official daily source)
# Monthly recon vs 前海考核校验 逐月汇总 row8/row9 lives in _build_oci_summary.py.
import openpyxl, sys, io, json, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PNL_MONTH = os.environ.get('PNL_MONTH', 'may').lower()
_MNUM = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5}[PNL_MONTH]
IS_JAN = (_MNUM == 1)
_CN = {1:'1月',2:'2月',3:'3月',4:'4月',5:'5月'}[_MNUM]
_EN = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May'}[_MNUM]
_IN = f'_oci_{PNL_MONTH}.json'
OUTFILE = os.environ.get('OUTFILE', f'PnL_OCI_2026{_EN}.xlsx')
MONTH_LABEL = f'2026年{_CN}'

D = json.load(open(_IN, encoding='utf-8'))
start, trades = D['start'], D['trades']
DAYS = D['JAN_DAYS']
px, px_code, cls = D['px'], D['px_code'], D['cls']
official_fx = D['official_fx']
dec31_hkd = D['dec31_hkd']; fx_dec31 = D['official_fx_dec31']
off_oci_daily = D.get('off_oci_daily', {}); off_jc_daily = D.get('off_jc_daily', {})
off_oci_month = D.get('off_oci_month', {}); off_jc_month = D.get('off_jc_month', {})
codes = sorted(start.keys(), key=lambda c: (cls[c], c))
NDAY = len(DAYS)

# ---------- GICS sector map (analyst-supplied; same as original) ----------
GICS = {
    '000001':'金融 Financials','000568':'日常消费 Consumer Staples','000858':'日常消费 Consumer Staples',
    '600036':'金融 Financials','600519':'日常消费 Consumer Staples','600809':'日常消费 Consumer Staples',
    '601009':'金融 Financials','601098':'通信服务 Comm Services','601166':'金融 Financials',
    '601169':'金融 Financials','601818':'金融 Financials','601900':'通信服务 Comm Services',
    '603198':'日常消费 Consumer Staples','603369':'日常消费 Consumer Staples',
    '000333':'可选消费 Consumer Disc','600887':'日常消费 Consumer Staples','600919':'金融 Financials',
    '601825':'金融 Financials','600908':'金融 Financials','601577':'金融 Financials',
    '002807':'金融 Financials','002839':'金融 Financials','600690':'可选消费 Consumer Disc',
    '00177':'工业 Industrials','00300':'可选消费 Consumer Disc','00322':'日常消费 Consumer Staples',
    '00371':'公用事业 Utilities','00576':'工业 Industrials','00728':'通信服务 Comm Services',
    '00762':'通信服务 Comm Services','00788':'通信服务 Comm Services','00811':'通信服务 Comm Services',
    '00836':'公用事业 Utilities','00883':'能源 Energy','00939':'金融 Financials','00941':'通信服务 Comm Services',
    '00995':'工业 Industrials','00998':'金融 Financials','01071':'公用事业 Utilities','01088':'能源 Energy',
    '01368':'可选消费 Consumer Disc','01398':'金融 Financials','02016':'金融 Financials','02020':'可选消费 Consumer Disc',
    '02318':'金融 Financials','02319':'日常消费 Consumer Staples','02607':'医疗保健 Health Care',
    '03328':'金融 Financials','03933':'医疗保健 Health Care','03988':'金融 Financials','03998':'可选消费 Consumer Disc',
    '06690':'可选消费 Consumer Disc','06818':'金融 Financials','06862':'可选消费 Consumer Disc','09985':'日常消费 Consumer Staples',
}

# ---------- cash dividends (IDENTICAL to _build_workbook.py; 01368 stays 0.084 CNY) ----------
DIVIDENDS_JAN = [
    dict(code='000568', ex='2026-01-30', per_share=1.358, ccy='CNY', dtype='现金分红', record='2026-01-29', pay='2026-01-30', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/000568.phtml', note='泸州老窖 ex=2026-01-30'),
    dict(code='600036', ex='2026-01-16', per_share=1.013, ccy='CNY', dtype='现金分红', record='2026-01-15', pay='2026-01-16', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/600036.phtml', note='招商银行 ex=2026-01-16'),
]
DIVIDENDS_FEB = [
    dict(code='601166', ex='2026-02-06', per_share=0.565, ccy='CNY', dtype='现金分红', record='2026-02-05', pay='2026-02-06', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/601166.phtml', note='兴业银行 ex=2026-02-06'),
    dict(code='601818', ex='2026-02-05', per_share=0.105, ccy='CNY', dtype='现金分红', record='2026-02-04', pay='2026-02-05', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/601818.phtml', note='光大银行 ex=2026-02-05'),
]
DIVIDENDS_MAR = []
DIVIDENDS_APR = []
DIVIDENDS_MAY = [
    dict(code='01398', ex='2026-05-04', per_share=0.16890, ccy='CNY', dtype='现金分红', record='2026-05-01', pay='2026-05-04', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/01398.phtml', note='工商银行(H) RMB 0.1689'),
    dict(code='00576', ex='2026-05-08', per_share=0.39500, ccy='CNY', dtype='现金分红', record='2026-05-07', pay='2026-05-08', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00576.phtml', note='浙江沪杭甬 RMB 0.395'),
    dict(code='01368', ex='2026-05-12', per_share='=0.095*0.8677', ccy='CNY', dtype='现金分红', record='2026-05-11', pay='2026-05-12', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/01368.phtml', note='特步国际 HKD 0.095 × CSDC mid 0.8677 (除权日) = CNY 0.08243'),
    dict(code='02020', ex='2026-05-15', per_share=0.935604, ccy='CNY', dtype='现金分红', record='2026-05-14', pay='2026-05-15', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/02020.phtml', note='安踏体育 HKD 1.08 × CSDC mid 0.8663'),
    dict(code='00998', ex='2026-05-18', per_share=0.19300, ccy='CNY', dtype='现金分红', record='2026-05-15', pay='2026-05-18', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00998.phtml', note='中信银行(H) RMB 0.193'),
    dict(code='00788', ex='2026-05-20', per_share=0.32539, ccy='CNY', dtype='现金分红', record='2026-05-19', pay='2026-05-20', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00788.phtml', note='中国铁塔 RMB 0.32539'),
    dict(code='06862', ex='2026-05-21', per_share=0.333773, ccy='CNY', dtype='现金分红', record='2026-05-20', pay='2026-05-21', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/06862.phtml', note='海底捞 HKD 0.384 × CSDC mid 0.8692'),
    dict(code='00811', ex='2026-05-29', per_share=0.42000, ccy='CNY', dtype='现金分红', record='2026-05-28', pay='2026-05-29', src='新浪财经', url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00811.phtml', note='新华文轩 RMB 0.42'),
]
DIVIDENDS = {'jan':DIVIDENDS_JAN,'feb':DIVIDENDS_FEB,'mar':DIVIDENDS_MAR,'apr':DIVIDENDS_APR,'may':DIVIDENDS_MAY}[PNL_MONTH]
div_by_code = {d['code']: d for d in DIVIDENDS}

# ---------- fee rate model (same as original) ----------
A_STAMP_BPS, A_COMM_BPS, HK_STAMP_BPS, HK_OTHER_BPS = 5.0, 0.641, 10.0, 1.362

# ---------- CSDC Jan FX (same as original; for Jan FX sheet) ----------
CSDC_BUY = {'2026-01-05':0.89422,'2026-01-06':0.89577,'2026-01-07':0.89592,'2026-01-08':0.89746,'2026-01-09':0.89578,'2026-01-12':0.89569,'2026-01-13':0.89409,'2026-01-14':0.89409,'2026-01-15':0.89412,'2026-01-16':0.89348,'2026-01-19':0.89318,'2026-01-20':0.89250,'2026-01-21':0.89135,'2026-01-22':0.89279,'2026-01-23':0.89259,'2026-01-26':0.89268,'2026-01-27':0.89200,'2026-01-28':0.89186,'2026-01-29':0.88948,'2026-01-30':0.88948}
CSDC_SELL = {'2026-01-05':0.89478,'2026-01-06':0.89583,'2026-01-07':0.89628,'2026-01-08':0.89754,'2026-01-09':0.89582,'2026-01-12':0.89551,'2026-01-13':0.89411,'2026-01-14':0.89411,'2026-01-15':0.89428,'2026-01-16':0.89352,'2026-01-19':0.89322,'2026-01-20':0.89250,'2026-01-21':0.89185,'2026-01-22':0.89281,'2026-01-23':0.89261,'2026-01-26':0.89272,'2026-01-27':0.89200,'2026-01-28':0.89174,'2026-01-29':0.88952,'2026-01-30':0.88952}

# ---------- styling ----------
HDR=Font(bold=True,color='FFFFFF',size=10); HDRFILL=PatternFill('solid',fgColor='305496')
SUBHDR=Font(bold=True,size=10); BOLD=Font(bold=True)
GREY=PatternFill('solid',fgColor='D9E1F2'); YEL=PatternFill('solid',fgColor='FFF2CC')
RED=PatternFill('solid',fgColor='F8CBAD'); GRN=PatternFill('solid',fgColor='C6E0B4')
BLUE=PatternFill('solid',fgColor='BDD7EE')
thin=Side(style='thin',color='BFBFBF'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CNY='#,##0;[Red](#,##0)'; CNY2='#,##0.00;[Red](#,##0.00)'; PCT='0.0%'; PX4='0.0000'
def style_header_row(ws,row,ncol,start_col=1):
    for c in range(start_col,start_col+ncol):
        cell=ws.cell(row=row,column=c); cell.font=HDR; cell.fill=HDRFILL
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=BORD

wb = openpyxl.Workbook()

# ============================================================ README
ws = wb.active; ws.title='README'
readme = [
 (f'港股通 A+H 股票 FVTOCI 损益 / OCI公允·OCI价差 — {MONTH_LABEL} (自营权益4)',14,True),
 ('',10,False),
 (f'Period: 2026-{_MNUM:02d}  ({NDAY} trading days).  Account 自营权益4.',10,False),
 ('',10,False),
 ('METHOD — FVTOCI (IFRS-9 FVOCI equities), fully formula-driven',12,True),
 ('• 移动加权平均成本 (WAVG). 买入费用资本化进成本; 卖出费用冲减已实现 (FVTOCI).',10,False),
 ('• 每日未实现 OCI公允_d = ΔHoldFV;  HoldFV_d = 持仓数量_d × 收盘价CNY_d − 成本池_d.',10,False),
 ('• 每日已实现 OCI价差_d = 卖出净额(含费) − 卖出数量 × 期初加权成本_d.',10,False),
 ('• 恒等式: OCI公允_d + OCI价差_d = 价格盈亏_d − 交易费_d  (股息单列, 不在OCI内).',10,False),
 ('• 港股通 收盘价CNY = 收盘价HKD × 官方CSDC中间价.',10,False),
 ('',10,False),
 ('UNIVERSE — opening holdings ∪ ANY code traded in-month (new positions INCLUDED)',12,True),
 ('• 旧版工作簿仅取期初持仓, 漏掉了月中新建仓的股票 (本版已修正).',10,False),
 ('',10,False),
 ('RECONCILIATION',12,True),
 ('• Jan: 每股×每日 我方OCI公允/OCI价差 vs 官方 2026年1月(1).xlsx (本年累计, 万元×10000).',10,False),
 ('• Feb–May: 无官方每日来源, 仅构建我方每日 OCI; 月度对账见汇总簿 vs 逐月汇总 row8/row9.',10,False),
 ('',10,False),
 ('SHEETS',12,True),
 ('  Monthly_Subjects — 每股月度 OCI价差/OCI公允/股利/费用 + 合计.  ★ START HERE',10,False),
 ('  PnL_Analysis     — 分部(A股/港股通) 已实现/未实现/价格/股息/费用/净损益/期初市值.',10,False),
 ('  Daily_OCI_GY     — 每股×每日 未实现 OCI公允 (formulas).',10,False),
 ('  Daily_OCI_JC     — 每股×每日 已实现 OCI价差 (formulas).',10,False),
 ('  Recon_Daily_*    — (Jan only) 我方 − 官方 每日差异.',10,False),
 ('  CostPool/QtyEnd/PxCNY/PxNative/FX/Dividends/Fees/Trades/StartPos — 计算与原始输入.',10,False),
]
for i,(txt,sz,b) in enumerate(readme,1):
    ws.cell(row=i,column=1,value=txt).font=Font(bold=b,size=sz)
ws.column_dimensions['A'].width=108
ws['A1'].fill=HDRFILL; ws['A1'].font=Font(bold=True,size=14,color='FFFFFF')

# ============================================================ StartPos
ws=wb.create_sheet('StartPos')
ws.append(['代码 code','名称 name','类别 class','期初数量 qty','期初成本(CNY) cost',
           '期初价(CNY) px_base','期初价(原币) px_base_FC','期初FX base'])
style_header_row(ws,1,8)
sp_row={}
for i,code in enumerate(codes,start=2):
    s=start[code]
    ws.cell(row=i,column=1,value=code).number_format='@'
    ws.cell(row=i,column=2,value=s['name']); ws.cell(row=i,column=3,value=cls[code])
    ws.cell(row=i,column=4,value=s['qty'])
    ws.cell(row=i,column=5,value=s['cost']).number_format=CNY2
    ws.cell(row=i,column=6,value=s['px_cny']).number_format=PX4
    if cls[code]=='HK':
        ws.cell(row=i,column=7,value=dec31_hkd.get(px_code[code])).number_format=PX4
        ws.cell(row=i,column=8,value=fx_dec31).number_format='0.00000'
    else:
        ws.cell(row=i,column=7,value=f'=F{i}').number_format=PX4
        ws.cell(row=i,column=8,value=1).number_format='0.00000'
    sp_row[code]=i
for col,w in zip('ABCDEFGH',[10,18,7,13,16,14,16,12]): ws.column_dimensions[col].width=w
ws.freeze_panes='A2'; SP_LAST=len(codes)+1

# ============================================================ Trades
ws=wb.create_sheet('Trades')
ws.append(['日期 date','代码 code','名称 name','方向 dir','sign','数量 qty','带符号数量 signed_qty',
           '结算金额CNY 本币成交金额','现金流 cashflow','市场 market','成交金额(原币)','带符号原币现金流'])
style_header_row(ws,1,12)
for i,t in enumerate(trades,start=2):
    ws.cell(row=i,column=1,value=t['date']).number_format='@'
    ws.cell(row=i,column=2,value=t['code']).number_format='@'
    ws.cell(row=i,column=3,value=t['name']); ws.cell(row=i,column=4,value=t['dir'])
    ws.cell(row=i,column=5,value=t['sign']); ws.cell(row=i,column=6,value=t['qty'])
    ws.cell(row=i,column=7,value=t['signed_qty'])
    ws.cell(row=i,column=8,value=t['cny_amt']).number_format=CNY2
    ws.cell(row=i,column=9,value=f'=E{i}*H{i}').number_format=CNY2
    ws.cell(row=i,column=10,value=t['mkt'])
    ws.cell(row=i,column=11,value=t['trade_amt']).number_format=CNY2
    ws.cell(row=i,column=12,value=f'=E{i}*K{i}').number_format=CNY2
for col,w in zip('ABCDEFGHIJKL',[12,10,16,8,6,11,14,18,18,14,16,18]): ws.column_dimensions[col].width=w
ws.freeze_panes='A2'; TR_LAST=len(trades)+1

# ============================================================ FX
ws=wb.create_sheet('FX')
ws.append(['日期 date','买入 Buy','卖出 Sell','中间价 Mid=(B+C)/2'])
style_header_row(ws,1,4)
for i,day in enumerate(DAYS,start=2):
    ws.cell(row=i,column=1,value=day).number_format='@'
    if IS_JAN and day in CSDC_BUY:
        ws.cell(row=i,column=2,value=CSDC_BUY[day]).number_format='0.00000'
        ws.cell(row=i,column=3,value=CSDC_SELL[day]).number_format='0.00000'
        ws.cell(row=i,column=4,value=f'=(B{i}+C{i})/2').number_format='0.00000'
    else:
        ws.cell(row=i,column=4,value=official_fx[day]).number_format='0.00000'
for col,w in zip('ABCD',[13,12,12,20]): ws.column_dimensions[col].width=w
ws.freeze_panes='A2'
FX_RANGE=f'FX!$A$2:$D${NDAY+1}'; FX_COL=4

# ============================================================ matrix geometry
DAY0=4
def day_col(i): return DAY0+i
def day_letter(i): return get_column_letter(day_col(i))
def write_matrix_frame(ws):
    ws.cell(row=1,column=1,value='code').font=SUBHDR
    ws.cell(row=1,column=2,value='name').font=SUBHDR
    ws.cell(row=1,column=3,value='class').font=SUBHDR
    for i,day in enumerate(DAYS):
        c=ws.cell(row=1,column=day_col(i),value=day); c.font=HDR; c.fill=HDRFILL
        c.number_format='@'; c.alignment=Alignment(horizontal='center')
    for r,code in enumerate(codes,start=2):
        ws.cell(row=r,column=1,value=code).number_format='@'
        ws.cell(row=r,column=2,value=start[code]['name']); ws.cell(row=r,column=3,value=cls[code])
    ws.freeze_panes=get_column_letter(DAY0)+'2'
    ws.column_dimensions['A'].width=9; ws.column_dimensions['B'].width=17; ws.column_dimensions['C'].width=6
    for i in range(NDAY): ws.column_dimensions[day_letter(i)].width=11
MROW={code:r for r,code in enumerate(codes,start=2)}

# ============================================================ PxNative
ws=wb.create_sheet('PxNative'); write_matrix_frame(ws)
for code in codes:
    r=MROW[code]
    for i,day in enumerate(DAYS):
        v=px.get(px_code[code],{}).get(day)
        if v is not None: ws.cell(row=r,column=day_col(i),value=v).number_format=PX4
style_header_row(ws,1,3)

# ============================================================ PxCNY (carry-forward)
ws=wb.create_sheet('PxCNY'); write_matrix_frame(ws)
for code in codes:
    r=MROW[code]
    for i,day in enumerate(DAYS):
        L=day_letter(i); nat=f'PxNative!{L}{r}'
        fxlk=f'VLOOKUP({L}$1,{FX_RANGE},{FX_COL},FALSE)'
        conv=f'IF($C{r}="HK",{nat}*{fxlk},{nat})'
        if i==0:
            # day-0 carries the baseline mark (StartPos F) so there is no day-1 price move;
            # the baseline→day-2 move lands on the 2nd period day. Matches the official book,
            # which does not mark the first period day (港股通 first day / no fixing).
            f=f'=StartPos!$F{sp_row[code]}'
        else:
            prev=f'{day_letter(i-1)}{r}'
            f=f'=IF({nat}="",{prev},{conv})'
        ws.cell(row=r,column=day_col(i),value=f).number_format=PX4
style_header_row(ws,1,3)

# ============================================================ QtyEnd (cumulative qty)
ws=wb.create_sheet('QtyEnd'); write_matrix_frame(ws)
TR_DATE=f'Trades!$A$2:$A${TR_LAST}'; TR_CODE=f'Trades!$B$2:$B${TR_LAST}'
TR_DIR=f'Trades!$D$2:$D${TR_LAST}'; TR_SQTY=f'Trades!$G$2:$G${TR_LAST}'
TR_Q=f'Trades!$F$2:$F${TR_LAST}'; TR_CNY=f'Trades!$H$2:$H${TR_LAST}'
for code in codes:
    r=MROW[code]
    for i,day in enumerate(DAYS):
        L=day_letter(i)
        today=f'SUMIFS({TR_SQTY},{TR_CODE},$A{r},{TR_DATE},{L}$1)'
        prev=f'StartPos!$D{sp_row[code]}' if i==0 else f'{day_letter(i-1)}{r}'
        ws.cell(row=r,column=day_col(i),value=f'={prev}+{today}').number_format='#,##0'
style_header_row(ws,1,3)

# ============================================================ Dividends (position-driven)
ws=wb.create_sheet('Dividends'); write_matrix_frame(ws)
DTOT=day_col(NDAY); EXD=DTOT+1; PS=DTOT+2; SH=DTOT+3; SRC=DTOT+4; NOTE=DTOT+5
for col,txt in {DTOT:'月合计 Div_Month',EXD:'除权日 ex',PS:'每股CNY per-share',
                SH:'除权日持股 shares@ex',SRC:'数据源',NOTE:'备注 note'}.items():
    cc=ws.cell(row=1,column=col,value=txt); cc.font=HDR; cc.fill=HDRFILL
    cc.alignment=Alignment(horizontal='center',wrap_text=True)
ps_letter=get_column_letter(PS)
for code in codes:
    r=MROW[code]; dv=div_by_code.get(code)
    if dv:
        ws.cell(row=r,column=EXD,value=dv['ex']).number_format='@'
        ws.cell(row=r,column=PS,value=dv['per_share']).number_format='0.0000'
        exL=day_letter(DAYS.index(dv['ex'])) if dv['ex'] in DAYS else None
        ws.cell(row=r,column=SH,value=(f'=QtyEnd!{exL}{r}' if exL else 0)).number_format='#,##0'
        ws.cell(row=r,column=SRC,value=dv['src']); ws.cell(row=r,column=NOTE,value=dv['note'])
    for i,day in enumerate(DAYS):
        if dv and day==dv['ex']:
            ws.cell(row=r,column=day_col(i),value=f'=QtyEnd!{day_letter(i)}{r}*${ps_letter}{r}').number_format=CNY
        else:
            ws.cell(row=r,column=day_col(i),value=0).number_format=CNY
    ws.cell(row=r,column=DTOT,value=f'=SUM({day_letter(0)}{r}:{day_letter(NDAY-1)}{r})').number_format=CNY
for col,w in [(DTOT,14),(EXD,12),(PS,16),(SH,16),(SRC,11),(NOTE,40)]:
    ws.column_dimensions[get_column_letter(col)].width=w
style_header_row(ws,1,3); DIV_TOTCOL=get_column_letter(DTOT)

# ============================================================ Fees (per trade)
ws=wb.create_sheet('Fees')
ws.append(['日期 date','代码 code','名称 name','类别 class','方向 dir','成交金额CNY notional',
           '印花税率bps','佣金率bps','印花税CNY','佣金CNY','费用合计CNY fee_total'])
style_header_row(ws,1,11)
ws.cell(row=1,column=13,value='费率参数 RATE INPUTS (bps)').font=BOLD
for k,(lab,val) in enumerate([('A股 印花税(卖)',A_STAMP_BPS),('A股 佣金',A_COMM_BPS),
                              ('港股通 印花税',HK_STAMP_BPS),('港股通 其他',HK_OTHER_BPS)]):
    ws.cell(row=2+k,column=13,value=lab).font=Font(size=9)
    rc=ws.cell(row=2+k,column=14,value=val); rc.number_format='0.000'; rc.fill=YEL; rc.border=BORD
rA_stamp,rA_comm,rHK_stamp,rHK_other='$N$2','$N$3','$N$4','$N$5'
ws.column_dimensions['M'].width=22; ws.column_dimensions['N'].width=9
for i,t in enumerate(trades,start=2):
    ws.cell(row=i,column=1,value=t['date']).number_format='@'
    ws.cell(row=i,column=2,value=t['code']).number_format='@'
    ws.cell(row=i,column=3,value=t['name']); ws.cell(row=i,column=4,value=t['cls'])
    ws.cell(row=i,column=5,value=t['dir'])
    ws.cell(row=i,column=6,value=t['cny_amt']).number_format=CNY2
    ws.cell(row=i,column=7,value=f'=IF(D{i}="A",IF(E{i}="卖出",{rA_stamp},0),{rHK_stamp})').number_format='0.000'
    ws.cell(row=i,column=8,value=f'=IF(D{i}="A",{rA_comm},{rHK_other})').number_format='0.000'
    ws.cell(row=i,column=9,value=f'=F{i}*G{i}/10000').number_format=CNY2
    ws.cell(row=i,column=10,value=f'=F{i}*H{i}/10000').number_format=CNY2
    ws.cell(row=i,column=11,value=f'=I{i}+J{i}').number_format=CNY2
FEE_LAST=len(trades)+1
fbuy=FEE_LAST+1; fsell=FEE_LAST+2; ftot=FEE_LAST+3
def _fee_subtot(row,label,crit):
    ws.cell(row=row,column=3,value=label).font=BOLD
    for col in ('F','I','J','K'):
        if crit is None: f=f'=SUM({col}2:{col}{FEE_LAST})'
        else: f=f'=SUMIFS({col}2:{col}{FEE_LAST},$E$2:$E${FEE_LAST},"{crit}")'
        cc=ws.cell(row=row,column=openpyxl.utils.column_index_from_string(col),value=f)
        cc.number_format=CNY2; cc.font=BOLD; cc.fill=GREY
    for c in (1,2,3): ws.cell(row=row,column=c).fill=GREY
_fee_subtot(fbuy,'买入小计 BUY','买入'); _fee_subtot(fsell,'卖出小计 SELL','卖出'); _fee_subtot(ftot,'合计 TOTAL',None)
for c in range(1,12): ws.cell(row=ftot,column=c).fill=BLUE
for col,w in zip('ABCDEFGHIJK',[12,9,16,7,8,16,11,10,12,12,15]): ws.column_dimensions[col].width=w
ws.freeze_panes='A2'
FEE_DATE=f'Fees!$A$2:$A${FEE_LAST}'; FEE_CODE=f'Fees!$B$2:$B${FEE_LAST}'
FEE_DIR=f'Fees!$E$2:$E${FEE_LAST}'; FEE_TOT=f'Fees!$K$2:$K${FEE_LAST}'

# ============================================================ CostPool (moving WAVG, recursive)
# Same-day buys settle BEFORE same-day sells (matches the official 估值 engine): the day's
# WAVG basis applied to sells already includes that day's buys.
#   poolCost_d = costpool_{d-1} + buyCashGross_d + buyFee_d   (buy fee capitalized, FVTOCI)
#   poolQty_d  = qty_{d-1} + buyQty_d
#   wavg_d     = poolCost_d / poolQty_d        (basis for that day's sells)
#   costpool_d = poolCost_d − sellQty_d × wavg_d
#   day0 prev  = StartPos cost / qty.
ws=wb.create_sheet('CostPool'); write_matrix_frame(ws)
def _buyCash(r,L):  return f'SUMIFS({TR_CNY},{TR_CODE},$A{r},{TR_DATE},{L}$1,{TR_DIR},"买入")'
def _sellProc(r,L): return f'SUMIFS({TR_CNY},{TR_CODE},$A{r},{TR_DATE},{L}$1,{TR_DIR},"卖出")'
def _buyFee(r,L):   return f'SUMIFS({FEE_TOT},{FEE_CODE},$A{r},{FEE_DATE},{L}$1,{FEE_DIR},"买入")'
def _sellFee(r,L):  return f'SUMIFS({FEE_TOT},{FEE_CODE},$A{r},{FEE_DATE},{L}$1,{FEE_DIR},"卖出")'
def _sellQty(r,L):  return f'SUMIFS({TR_Q},{TR_CODE},$A{r},{TR_DATE},{L}$1,{TR_DIR},"卖出")'
def _buyQty(r,L):   return f'SUMIFS({TR_Q},{TR_CODE},$A{r},{TR_DATE},{L}$1,{TR_DIR},"买入")'
def wavg_today(r,i,sp):
    """WAVG per share for day i sells = (prevCost + sameday buyCash + buyFee)/(prevQty + buyQty)."""
    L=day_letter(i)
    if i==0:
        prevCost=f'StartPos!$E{sp}'; prevQty=f'StartPos!$D{sp}'
    else:
        Lp=day_letter(i-1); prevCost=f'CostPool!{Lp}{r}'; prevQty=f'QtyEnd!{Lp}{r}'
    poolCost=f'({prevCost}+{_buyCash(r,L)}+{_buyFee(r,L)})'
    poolQty=f'({prevQty}+{_buyQty(r,L)})'
    return f'IF({poolQty}=0,0,{poolCost}/{poolQty})', poolCost
for code in codes:
    r=MROW[code]; sp=sp_row[code]
    for i,day in enumerate(DAYS):
        L=day_letter(i)
        w,poolCost=wavg_today(r,i,sp)
        f=f'={poolCost}-{_sellQty(r,L)}*({w})'
        ws.cell(row=r,column=day_col(i),value=f).number_format=CNY2
style_header_row(ws,1,3)

# ============================================================ Daily_OCI_GY (unrealized)
# OCI公允_d = HoldFV_d − HoldFV_{d-1};  HoldFV_d = QtyEnd_d×PxCNY_d − CostPool_d
ws=wb.create_sheet('Daily_OCI_GY'); write_matrix_frame(ws)
GYTOT=day_col(NDAY)
ws.cell(row=1,column=GYTOT,value='月合计 OCI公允_Month').font=HDR; ws.cell(row=1,column=GYTOT).fill=HDRFILL
for code in codes:
    r=MROW[code]; sp=sp_row[code]
    for i,day in enumerate(DAYS):
        L=day_letter(i)
        hold=f'(QtyEnd!{L}{r}*PxCNY!{L}{r}-CostPool!{L}{r})'
        if i==0:
            holdp=f'(StartPos!$D{sp}*StartPos!$F{sp}-StartPos!$E{sp})'
        else:
            Lp=day_letter(i-1); holdp=f'(QtyEnd!{Lp}{r}*PxCNY!{Lp}{r}-CostPool!{Lp}{r})'
        ws.cell(row=r,column=day_col(i),value=f'={hold}-{holdp}').number_format=CNY
    ws.cell(row=r,column=GYTOT,value=f'=SUM({day_letter(0)}{r}:{day_letter(NDAY-1)}{r})').number_format=CNY
gy_trow=len(codes)+2
ws.cell(row=gy_trow,column=2,value='合计 TOTAL').font=BOLD
for i in range(NDAY):
    L=day_letter(i)
    ws.cell(row=gy_trow,column=day_col(i),value=f'=SUM({L}2:{L}{len(codes)+1})').number_format=CNY
    ws.cell(row=gy_trow,column=day_col(i)).font=BOLD
ws.cell(row=gy_trow,column=GYTOT,value=f'=SUM({get_column_letter(GYTOT)}2:{get_column_letter(GYTOT)}{len(codes)+1})').number_format=CNY
ws.cell(row=gy_trow,column=GYTOT).font=BOLD
ws.column_dimensions[get_column_letter(GYTOT)].width=18; style_header_row(ws,1,3)
GY_TOTCOL=get_column_letter(GYTOT)

# ============================================================ Daily_OCI_JC (realized)
# OCI价差_d = sellProceeds_d − sellFee_d − sellQty_d × wavg_d
#   wavg_d INCLUDES same-day buys (same basis CostPool applies to sells) so realized + unrealized
#   reconcile to the official 估值 engine on same-day buy+sell names (e.g. 03933 联邦制药 May-28).
ws=wb.create_sheet('Daily_OCI_JC'); write_matrix_frame(ws)
JCTOT=day_col(NDAY)
ws.cell(row=1,column=JCTOT,value='月合计 OCI价差_Month').font=HDR; ws.cell(row=1,column=JCTOT).fill=HDRFILL
for code in codes:
    r=MROW[code]; sp=sp_row[code]
    for i,day in enumerate(DAYS):
        L=day_letter(i)
        w,_pool=wavg_today(r,i,sp)
        f=f'={_sellProc(r,L)}-{_sellFee(r,L)}-{_sellQty(r,L)}*({w})'
        ws.cell(row=r,column=day_col(i),value=f).number_format=CNY
    ws.cell(row=r,column=JCTOT,value=f'=SUM({day_letter(0)}{r}:{day_letter(NDAY-1)}{r})').number_format=CNY
jc_trow=len(codes)+2
ws.cell(row=jc_trow,column=2,value='合计 TOTAL').font=BOLD
for i in range(NDAY):
    L=day_letter(i)
    ws.cell(row=jc_trow,column=day_col(i),value=f'=SUM({L}2:{L}{len(codes)+1})').number_format=CNY
    ws.cell(row=jc_trow,column=day_col(i)).font=BOLD
ws.cell(row=jc_trow,column=JCTOT,value=f'=SUM({get_column_letter(JCTOT)}2:{get_column_letter(JCTOT)}{len(codes)+1})').number_format=CNY
ws.cell(row=jc_trow,column=JCTOT).font=BOLD
ws.column_dimensions[get_column_letter(JCTOT)].width=18; style_header_row(ws,1,3)
JC_TOTCOL=get_column_letter(JCTOT)

# ============================================================ Jan official + daily recon
if IS_JAN:
    def _off_matrix(sheet, src, totlabel):
        wso=wb.create_sheet(sheet); write_matrix_frame(wso)
        TOT=day_col(NDAY)
        wso.cell(row=1,column=TOT,value=totlabel).font=HDR; wso.cell(row=1,column=TOT).fill=HDRFILL
        for code in codes:
            r=MROW[code]; series=src.get(code,{})
            for i,day in enumerate(DAYS):
                wso.cell(row=r,column=day_col(i),value=round(series.get(day,0.0),2)).number_format=CNY
            # month total uses COMPLETE official sum (incl any non-trading-day flow)
            full = off_oci_month.get(code,0.0) if 'GY' in sheet else off_jc_month.get(code,0.0)
            wso.cell(row=r,column=TOT,value=round(full,2)).number_format=CNY
        trow=len(codes)+2
        wso.cell(row=trow,column=2,value='合计 TOTAL').font=BOLD
        for i in range(NDAY):
            L=day_letter(i)
            wso.cell(row=trow,column=day_col(i),value=f'=SUM({L}2:{L}{len(codes)+1})').number_format=CNY
            wso.cell(row=trow,column=day_col(i)).font=BOLD
        wso.cell(row=trow,column=TOT,value=f'=SUM({get_column_letter(TOT)}2:{get_column_letter(TOT)}{len(codes)+1})').number_format=CNY
        wso.cell(row=trow,column=TOT).font=BOLD
        wso.column_dimensions[get_column_letter(TOT)].width=18; style_header_row(wso,1,3)
        return get_column_letter(TOT)
    OFFGY_TOT=_off_matrix('Off_OCI_GY', off_oci_daily, '月合计(全月官方) OCI公允')
    OFFJC_TOT=_off_matrix('Off_OCI_JC', off_jc_daily, '月合计(全月官方) OCI价差')

    def _recon_daily(sheet, ours, off, ourstot, offtot):
        wsr=wb.create_sheet(sheet); write_matrix_frame(wsr)
        TOT=day_col(NDAY)
        wsr.cell(row=1,column=TOT,value='月差异 (我方−官方全月)').font=HDR; wsr.cell(row=1,column=TOT).fill=HDRFILL
        for code in codes:
            r=MROW[code]
            for i,day in enumerate(DAYS):
                L=day_letter(i)
                wsr.cell(row=r,column=day_col(i),value=f'={ours}!{L}{r}-{off}!{L}{r}').number_format=CNY
            wsr.cell(row=r,column=TOT,value=f'={ours}!{ourstot}{r}-{off}!{offtot}{r}').number_format=CNY
        trow=len(codes)+2
        wsr.cell(row=trow,column=2,value='合计 TOTAL').font=BOLD
        for i in range(NDAY):
            L=day_letter(i)
            wsr.cell(row=trow,column=day_col(i),value=f'=SUM({L}2:{L}{len(codes)+1})').number_format=CNY
            wsr.cell(row=trow,column=day_col(i)).font=BOLD
        wsr.cell(row=trow,column=TOT,value=f'=SUM({get_column_letter(TOT)}2:{get_column_letter(TOT)}{len(codes)+1})').number_format=CNY
        wsr.cell(row=trow,column=TOT).font=BOLD
        wsr.column_dimensions[get_column_letter(TOT)].width=20; style_header_row(wsr,1,3)
        # green/red on diff
        from openpyxl.formatting.rule import CellIsRule
        rng=f'{day_letter(0)}2:{get_column_letter(TOT)}{len(codes)+1}'
        wsr.conditional_formatting.add(rng,CellIsRule(operator='between',formula=['-50','50'],fill=GRN))
        wsr.conditional_formatting.add(rng,CellIsRule(operator='greaterThan',formula=['3000'],fill=RED))
        wsr.conditional_formatting.add(rng,CellIsRule(operator='lessThan',formula=['-3000'],fill=RED))
    _recon_daily('Recon_Daily_GY','Daily_OCI_GY','Off_OCI_GY',GY_TOTCOL,OFFGY_TOT)
    _recon_daily('Recon_Daily_JC','Daily_OCI_JC','Off_OCI_JC',JC_TOTCOL,OFFJC_TOT)

# ============================================================ Monthly_Subjects ★
ws=wb.create_sheet('Monthly_Subjects')
hdr=['代码 code','名称 name','类别','OCI价差 已实现 Realized','OCI公允 未实现 Unrealized',
     '价格盈亏 Price (价差+公允)','股利 Dividend','交易费用 Fees(含于OCI,披露)','净损益 Net(价差+公允+股利)','GICS 行业']
if IS_JAN:
    hdr += ['官方OCI价差','官方OCI公允','差异 价差(我方−官方)','差异 公允(我方−官方)']
ws.append(hdr); style_header_row(ws,1,len(hdr))
MS_FIRST=2
for r,code in enumerate(codes,start=2):
    mr=MROW[code]
    ws.cell(row=r,column=1,value=code).number_format='@'
    ws.cell(row=r,column=2,value=start[code]['name']); ws.cell(row=r,column=3,value=cls[code])
    ws.cell(row=r,column=4,value=f'=Daily_OCI_JC!{JC_TOTCOL}{mr}').number_format=CNY   # D realized
    ws.cell(row=r,column=5,value=f'=Daily_OCI_GY!{GY_TOTCOL}{mr}').number_format=CNY   # E unrealized
    ws.cell(row=r,column=6,value=f'=D{r}+E{r}').number_format=CNY                       # F price
    ws.cell(row=r,column=7,value=f'=Dividends!{DIV_TOTCOL}{mr}').number_format=CNY       # G dividend
    ws.cell(row=r,column=8,value=f'=-SUMIFS({FEE_TOT},{FEE_CODE},$A{r})').number_format=CNY  # H fees (disclosure)
    ws.cell(row=r,column=9,value=f'=F{r}+G{r}').number_format=CNY                        # I net
    ws.cell(row=r,column=10,value=GICS.get(code,'其他 Other'))
    if IS_JAN:
        ws.cell(row=r,column=11,value=f'=Off_OCI_JC!{OFFJC_TOT}{mr}').number_format=CNY
        ws.cell(row=r,column=12,value=f'=Off_OCI_GY!{OFFGY_TOT}{mr}').number_format=CNY
        ws.cell(row=r,column=13,value=f'=D{r}-K{r}').number_format=CNY
        ws.cell(row=r,column=14,value=f'=E{r}-L{r}').number_format=CNY
MS_LAST=len(codes)+1
def ms_subtot(row,label,crit):
    ws.cell(row=row,column=2,value=label).font=BOLD
    cols=['D','E','F','G','H','I'] + (['K','L','M','N'] if IS_JAN else [])
    for col in cols:
        if crit is None: f=f'=SUM({col}{MS_FIRST}:{col}{MS_LAST})'
        else: f=f'=SUMIFS({col}${MS_FIRST}:{col}${MS_LAST},$C${MS_FIRST}:$C${MS_LAST},"{crit}")'
        cc=ws.cell(row=row,column=openpyxl.utils.column_index_from_string(col),value=f)
        cc.number_format=CNY; cc.font=BOLD; cc.fill=GREY
    for c in (1,2,3): ws.cell(row=row,column=c).fill=GREY
    ws.cell(row=row,column=3,value=('' if crit is None else crit)).font=BOLD
a_row=MS_LAST+1; hk_row=MS_LAST+2; tot_row=MS_LAST+3
ms_subtot(a_row,'A股小计 A-share','A')
ms_subtot(hk_row,'港股通小计 HK','HK'); ms_subtot(tot_row,'合计 GRAND TOTAL',None)
ncol=len(hdr)
for c in range(1,ncol+1): ws.cell(row=tot_row,column=c).fill=BLUE
ws.cell(row=tot_row,column=2).font=Font(bold=True,size=11)
for col,w in zip('ABCDEFGHIJ',[10,20,7,18,18,16,13,18,18,20]): ws.column_dimensions[col].width=w
if IS_JAN:
    for col,w in zip('KLMN',[14,14,16,16]): ws.column_dimensions[col].width=w
ws.freeze_panes='D2'
MS='Monthly_Subjects'; A_ROW,HK_ROW,T_ROW=a_row,hk_row,tot_row

# ============================================================ PnL_Analysis (segment) ★
ws=wb.create_sheet('PnL_Analysis')
TITLE=Font(bold=True,size=13,color='FFFFFF'); SEC=Font(bold=True,size=11,color='FFFFFF')
SECFILL=PatternFill('solid',fgColor='305496')
ws.cell(row=1,column=1,value=f'PnL 分析 (FVTOCI) / PnL Analysis — {MONTH_LABEL} (自营权益4)').font=TITLE
for c in range(1,8): ws.cell(row=1,column=c).fill=SECFILL
def msref(col,seg_row): return f'{MS}!{col}{seg_row}'
# opening MV by segment (StartPos)
SP='StartPos'
omvA=f'SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}="A")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})'
omvHK=f'SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}="HK")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})'
omvT=f'SUMPRODUCT({SP}!$D$2:$D${SP_LAST},{SP}!$F$2:$F${SP_LAST})'
r0=3
ws.cell(row=r0,column=1,value='① 损益拆解 (按分部, FVTOCI) / PnL by Segment').font=SEC
for c in range(1,6): ws.cell(row=r0,column=c).fill=SECFILL
hr=r0+1
for c,t in enumerate(['科目 Item','A股 (CNY)','港股通 (CNY)','合计 (CNY)','备注 Remark'],start=1):
    cc=ws.cell(row=hr,column=c,value=t); cc.font=HDR; cc.fill=HDRFILL
    cc.alignment=Alignment(horizontal='center',wrap_text=True); cc.border=BORD
bf=hr+1
rows_def=[
 ('已实现 OCI价差 Realized','D','卖出锁定(WAVG成本基准, 已扣卖出费)'),
 ('未实现 OCI公允 Unrealized','E','持仓浮动(已扣资本化买入费)'),
 ('价格盈亏 Price P&L',None,'＝ 已实现 + 未实现'),
 ('加: 股息 Dividend','G','现金分红(除权日计提)'),
 ('交易费用 Fees(披露)','H','已含于OCI(买入费入成本/卖出费冲已实现); 此处单列'),
 ('净损益 Net PnL',None,'＝ 价格盈亏 + 股息'),
 ('期初市值 Opening MV',None,'期初数量 × 期初市价(CNY)'),
 ('收益率 Return %',None,'＝ 净损益 / 期初市值'),
]
r_real=bf; r_unr=bf+1; r_price=bf+2; r_div=bf+3; r_fee=bf+4; r_net=bf+5; r_omv=bf+6; r_ret=bf+7
for k,(label,col,rmk) in enumerate(rows_def):
    rr=bf+k; emph=label.startswith(('价格','净损益','收益率'))
    ws.cell(row=rr,column=1,value=label).font=BOLD if emph else Font(size=10)
    if col:
        ws.cell(row=rr,column=2,value=f'={msref(col,A_ROW)}').number_format=CNY
        ws.cell(row=rr,column=3,value=f'={msref(col,HK_ROW)}').number_format=CNY
        ws.cell(row=rr,column=4,value=f'={msref(col,T_ROW)}').number_format=CNY
    elif label.startswith('价格'):
        for c,LL in [(2,'B'),(3,'C'),(4,'D')]: ws.cell(row=rr,column=c,value=f'={LL}{r_real}+{LL}{r_unr}').number_format=CNY
    elif label.startswith('净损益'):
        for c,LL in [(2,'B'),(3,'C'),(4,'D')]: ws.cell(row=rr,column=c,value=f'={LL}{r_price}+{LL}{r_div}').number_format=CNY
    elif label.startswith('期初市值'):
        ws.cell(row=rr,column=2,value=f'={omvA}').number_format=CNY
        ws.cell(row=rr,column=3,value=f'={omvHK}').number_format=CNY
        ws.cell(row=rr,column=4,value=f'={omvT}').number_format=CNY
    elif label.startswith('收益率'):
        for c,LL in [(2,'B'),(3,'C'),(4,'D')]: ws.cell(row=rr,column=c,value=f'=IF({LL}{r_omv}=0,"",{LL}{r_net}/{LL}{r_omv})').number_format=PCT
    ws.cell(row=rr,column=5,value=rmk).font=Font(size=9,color='666666')
    for c in range(1,5):
        ws.cell(row=rr,column=c).border=BORD
        if label.startswith('净损益') and c>1: ws.cell(row=rr,column=c).fill=GRN
        if label.startswith('价格') and c>1: ws.cell(row=rr,column=c).fill=GREY
        if emph: ws.cell(row=rr,column=c).font=BOLD
# Jan: official compare row
if IS_JAN:
    cr=r_ret+2
    ws.cell(row=cr,column=1,value='①b 对官方校验 / vs Official (Jan, 全月CNY)').font=SEC
    for c in range(1,6): ws.cell(row=cr,column=c).fill=SECFILL
    for c,t in enumerate(['科目','我方 Ours','官方 Official','差异 Diff'],start=1):
        cc=ws.cell(row=cr+1,column=c,value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD
    for k,(lab,our,off) in enumerate([('OCI价差 Realized','D','K'),('OCI公允 Unrealized','E','L')]):
        rr=cr+2+k
        ws.cell(row=rr,column=1,value=lab)
        ws.cell(row=rr,column=2,value=f'={msref(our,T_ROW)}').number_format=CNY
        ws.cell(row=rr,column=3,value=f'={msref(off,T_ROW)}').number_format=CNY
        ws.cell(row=rr,column=4,value=f'=B{rr}-C{rr}').number_format=CNY
        for c in range(1,5): ws.cell(row=rr,column=c).border=BORD
for col,w in zip('ABCDE',[26,16,16,16,52]): ws.column_dimensions[col].width=w
ws.sheet_view.showGridLines=False

# ============================================================ order + save
front=['Monthly_Subjects','PnL_Analysis','Daily_OCI_GY','Daily_OCI_JC']
if IS_JAN: front += ['Recon_Daily_GY','Recon_Daily_JC']
order=[s for s in front if s in wb.sheetnames]+[s for s in wb.sheetnames if s not in front]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active=0
wb.save(OUTFILE)
print('saved',OUTFILE)




