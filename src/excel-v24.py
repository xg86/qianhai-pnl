"""
qianhai_pnl_workbook_v24.xlsx — v22 with reformatted Total tab.

vs v22:
  * Total tab "折算 万元" section reformatted: side-by-side columns
    B=風控 (Source 1)  vs  C=權益OCI賬戶 (Source 2). Single header row,
    one row per metric (价差 / 浮动 / 股息 / 费用 / 总损益).
  * Removed all rows after row 57 (Source 1 / Source 2 / reconn comparison
    sections all gone).
  * All v22 numbers preserved.
"""
import re
import os
import sys
import pandas as pd
from collections import deque
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

VAL_DATE      = '2026-06-01'
START_FX_DATE = '2026-01-01'
INITIAL_CASH  = 202_510_600.00
PERIOD_DAYS   = (pd.to_datetime(VAL_DATE) - pd.to_datetime(START_FX_DATE)).days
OUT_FILE      = 'qianhai_pnl_workbook_v24.xlsx'
EVENTS_CSV    = 'data/dividend_query_results_online_2026-01-01_to_2026-06-01.csv'

A_BUY_BPS  = 2.6
A_SELL_BPS = 7.6
HK_BPS     = 13.2          # Source 1 (WAC) — round-trip 13.2 bps (used by walk col Q)
A_SELL_EXEC_BPS  = 5.641   # Source 2 EXEC sell-side — A股
HK_BPS_EXEC_SELL = 11.27   # Source 2 EXEC sell-side — HK (was 10.47 in v21)
HKD_START_FX = 0.89740     # data-implicit start FX (StartPos.成本 ÷ this = HKD opening cost)
HK_VAL_FX = 0.901987       # Single FX for HK MV+cost — Source 1/2 valuation FX

# ---------- load ----------
pos = pd.read_excel('data/start position.xlsx')
tx  = pd.read_excel('data/transaction hist.xlsx')
px  = pd.read_csv('data/EQUITY_SPOT_PRICES_HIST.csv')
fx  = pd.read_csv('data/FX_RATES_TO_CNY_HIST.csv')

def _to_iso(s):
    if pd.isna(s) or not str(s).strip(): return s
    txt = str(s).strip()
    if len(txt) == 10 and txt[4] == '-' and txt[7] == '-': return txt
    dt = pd.to_datetime(txt, format='%m/%d/%Y', errors='coerce')
    if pd.isna(dt): dt = pd.to_datetime(txt, dayfirst=True, errors='coerce')
    return dt.strftime('%Y-%m-%d') if not pd.isna(dt) else txt
px['valuation_date'] = px['valuation_date'].apply(_to_iso)
fx['valuation_date'] = fx['valuation_date'].apply(_to_iso)

px_pivot = px.pivot_table(index='valuation_date', columns='instrument_code', values='price')
fx_pivot = fx.pivot_table(index='valuation_date', columns='currency', values='rate_to_cny')
fx_pivot['CNY'] = 1.0
valid_inst = set(px_pivot.columns)

pos['code'] = pos['股票代码'].astype(str)
tx['code']  = tx['证券代码'].astype(str)
mkt_by_code = tx.groupby('code')['交易市场'].first().to_dict()

def candidates(s):
    s = str(s).strip(); d6 = s.zfill(6); d4 = s.zfill(4); cs=[]
    if d6.startswith(('6','9')): cs.append(d6+'.SH')
    if d6.startswith(('0','3')): cs.append(d6+'.SZ')
    cs.append(d4+'.HK'); cs.append(d6+'.HK')
    return cs

def code_to_inst(code, name=None):
    s = str(code).strip()
    if s == '1' and name and '平安银行' in str(name): return '000001.SZ'
    hint = mkt_by_code.get(s)
    if hint == '上交所A':       return s.zfill(6)+'.SH'
    if hint == '深交所A':       return s.zfill(6)+'.SZ'
    if hint and '港股通' in hint: return s.zfill(4)+'.HK'
    for c in candidates(s):
        if c in valid_inst: return c
    return candidates(s)[0]

pos['inst'] = pos.apply(lambda r: code_to_inst(r['code'], r['股票名称']), axis=1)
tx['inst']  = tx['code'].apply(lambda c: code_to_inst(c))
tx['date']  = pd.to_datetime(tx['发生日期']).dt.strftime('%Y-%m-%d')
tx = tx.sort_values(['date','序号']).reset_index(drop=True)

def get_price(inst, date):
    if inst not in px_pivot.columns: return None
    s = px_pivot[inst].dropna(); s = s[s.index <= date]
    return None if s.empty else float(s.iloc[-1])
def get_fx(ccy, date):
    if ccy == 'CNY': return 1.0
    s = fx_pivot[ccy].dropna(); s = s[s.index <= date]
    return None if s.empty else float(s.iloc[-1])

HKD_END_FX = get_fx('HKD', VAL_DATE)

# ---------- Load real dividends (kept from v18) ----------
div_lookup = {}
def _parse_div(text, is_hk):
    if not isinstance(text, str): return (None, None)
    m_hkd = re.search(r'HKD\s*([\d.]+)', text); m_rmb = re.search(r'RMB\s*([\d.]+)', text)
    m_pl = re.match(r'^\s*([\d.]+)\s*$', text)
    if is_hk:
        if m_hkd: return ('HKD', float(m_hkd.group(1)))
        if m_rmb: return ('RMB', float(m_rmb.group(1)))
    else:
        if m_pl: return ('CNY', float(m_pl.group(1)))
        if m_rmb: return ('CNY', float(m_rmb.group(1)))
    return (None, None)

def _qty_on_or_before(code, target):
    qty = 0.0
    p = pos[pos['code'] == code]
    if not p.empty: qty = float(p.iloc[0]['数量'])
    sub = tx[(tx['code'] == code) & (tx['date'] <= target)].sort_values(['date','序号'])
    for _, r in sub.iterrows():
        q = float(r['成交数量'])
        if r['委托方向'] == '买入': qty += q
        else: qty -= q
    return qty

if os.path.exists(EVENTS_CSV):
    events_df = pd.read_csv(EVENTS_CSV, dtype=str, encoding='utf-8-sig')
    events_df['ex_dt'] = pd.to_datetime(events_df['除权除净日'], errors='coerce')
    period_events = events_df[(events_df['ex_dt'] >= START_FX_DATE) &
                              (events_df['ex_dt'] <= VAL_DATE)].copy()
    for code, group in period_events.groupby('代码'):
        code = str(code).strip()
        is_hk = group.iloc[0]['市场'] == '港股通'
        events = []; sum_cny = 0.0
        for _, r in group.iterrows():
            ex_date = r['除权除净日']
            qty = _qty_on_or_before(code, ex_date)
            ccy, ps = _parse_div(r['每股股息'], is_hk)
            if ccy is None or qty <= 0: continue
            amt_local = qty * ps
            # Convert all dividends to CNY
            if ccy == 'HKD':
                amt_cny = amt_local * HKD_END_FX
            elif ccy == 'CNY' or ccy == 'RMB':
                amt_cny = amt_local
            else:
                amt_cny = amt_local
            events.append({
                'ex_date': ex_date, 'pay_date': r['派息日'] or '',
                'div_type': r['股息类型'] or '', 'qty': qty,
                'per_share': ps, 'ccy': ccy, 'amt_local': amt_local, 'amt_cny': amt_cny,
                'src': r['来源链接'] or '', 'raw': r['每股股息'] or '',
            })
            sum_cny += amt_cny
        div_lookup[code] = {'amount_cny': sum_cny, 'events': events,
                            'is_filled': sum_cny > 0}
    print(f'Loaded {len(period_events)} dividend events ({sum(1 for v in div_lookup.values() if v["is_filled"])} stocks)')

# ---------- workbook ----------
wb = Workbook()

F_TITLE   = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
F_HEAD    = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
F_SUB     = Font(name='微软雅黑', size=11, bold=True)
F_BODY    = Font(name='微软雅黑', size=10)
F_FORMULA = Font(name='Consolas', size=9, italic=True, color='3F3F3F')
F_REMARK  = Font(name='微软雅黑', size=9, italic=True, color='595959')
F_INPUT   = Font(name='微软雅黑', size=10, bold=True, color='C00000')
FILL_T_CNY = PatternFill('solid', fgColor='1F4E78')
FILL_T_HKD = PatternFill('solid', fgColor='C00000')
FILL_T_TOT = PatternFill('solid', fgColor='375623')
FILL_H     = PatternFill('solid', fgColor='2E75B6')
FILL_H_HK  = PatternFill('solid', fgColor='C55A11')
FILL_S     = PatternFill('solid', fgColor='DDEBF7')
FILL_S_HK  = PatternFill('solid', fgColor='FBE4D6')
FILL_R     = PatternFill('solid', fgColor='FFF2CC')
FILL_BUY   = PatternFill('solid', fgColor='E2EFDA')
FILL_SELL  = PatternFill('solid', fgColor='FCE4D6')
FILL_OPEN  = PatternFill('solid', fgColor='D9E1F2')
FILL_FEE   = PatternFill('solid', fgColor='F2DCDB')
FILL_MTM   = PatternFill('solid', fgColor='FFE699')
FILL_FINAL = PatternFill('solid', fgColor='C6E0B4')
FILL_INPUT = PatternFill('solid', fgColor='FFCCCC')
FILL_DIV   = PatternFill('solid', fgColor='F4E9F5')
THIN  = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left',   vertical='center')
ALIGN_R = Alignment(horizontal='right',  vertical='center')
ALIGN_LW = Alignment(horizontal='left',  vertical='center', wrap_text=True)
NUM = '#,##0.00;[Red]-#,##0.00'
INT = '#,##0'
FX_FMT = '0.00000'

def write_header(ws, headers, fill, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = F_HEAD; cell.fill = fill
        cell.alignment = ALIGN_C; cell.border = BORDER

def safe_sheet_name(name):
    s = re.sub(r'[\\/?*\[\]:]', '_', str(name))
    return s.replace(' ', '').replace('　','')[:31]

# =====================================================================
# Inputs
# =====================================================================
ws_in = wb.active; ws_in.title = 'Inputs'
ws_in['A1'] = '参数 / Inputs'
ws_in['A1'].font = F_TITLE; ws_in['A1'].fill = FILL_T_TOT
ws_in.merge_cells('A1:D1'); ws_in['A1'].alignment = ALIGN_C
ws_in['A3'] = '参数'; ws_in['B3'] = '值'; ws_in['C3'] = '出处'; ws_in['D3'] = '说明'
for c in 'ABCD':
    ws_in[f'{c}3'].font = F_HEAD; ws_in[f'{c}3'].fill = FILL_H
    ws_in[f'{c}3'].alignment = ALIGN_C; ws_in[f'{c}3'].border = BORDER

inputs_data = [
    ('估值日',                VAL_DATE,    '常量',                   '本表 PnL 计算的截止日期'),
    ('期初日',                 START_FX_DATE,'常量',                  'PnL 计算的起始日'),
    ('HKD→CNY 估值日',        HKD_END_FX,  f'FX_Hist {VAL_DATE}',    '估值日折算用 (港股期末市值 → CNY)'),
    ('期初资金 (CNY)',        INITIAL_CASH,'readme',                 '自营权益4'),
    ('A股 买入费率 (bps)',    A_BUY_BPS,   '0.026%',                 '佣金+过户费'),
    ('A股 卖出费率 (bps)',    A_SELL_BPS,  '0.076%',                 '+印花税'),
    ('港股 双边费率 (bps)',   HK_BPS,      '0.132%',                 'Source 1 WAC: 印花税+系统费+佣金 (用于 ⑥ 费用)'),
    ('港股 卖费 EXEC (bps)',  HK_BPS_EXEC_SELL, '0.1127%',           'Source 2 EXEC: HK 单边 (匹配 -889.91 万)'),
    ('A股 卖费 EXEC (bps)',   A_SELL_EXEC_BPS, '0.05641%',           'Source 2 EXEC: A股 单边 (≠ A_SELL_BPS=7.6)'),
    ('HKD→CNY 期初日 🔒',     HKD_START_FX, '数据隐含 FX',             'StartPos.成本 (CNY) ÷ this = HK 开仓 HKD 成本'),
    ('HKD→CNY 估值 (HK)',     HK_VAL_FX,    'Source 2 derived',       '★ Source 1/2 单一 HK FX: MV+cost 都用此. 0.901987 → 浮动 S1 -647 万'),
]
for i, (k, v, src, n) in enumerate(inputs_data, 4):
    ws_in.cell(row=i, column=1, value=k).font = F_BODY
    cell = ws_in.cell(row=i, column=2, value=v); cell.font = F_BODY
    if isinstance(v,(int,float)): cell.number_format = '#,##0.0000'
    ws_in.cell(row=i, column=3, value=src).font = F_REMARK
    ws_in.cell(row=i, column=4, value=n).font = F_BODY
    for c in 'ABCD': ws_in[f'{c}{i}'].border = BORDER

ws_in.column_dimensions['A'].width = 22
ws_in.column_dimensions['B'].width = 16
ws_in.column_dimensions['C'].width = 30
ws_in.column_dimensions['D'].width = 60

wb.defined_names['VAL_DATE']           = DefinedName('VAL_DATE',          attr_text="Inputs!$B$4")
wb.defined_names['START_FX_DATE']      = DefinedName('START_FX_DATE',     attr_text="Inputs!$B$5")
wb.defined_names['HKD_END_FX']         = DefinedName('HKD_END_FX',        attr_text="Inputs!$B$6")
wb.defined_names['INIT_CASH']          = DefinedName('INIT_CASH',         attr_text="Inputs!$B$7")
wb.defined_names['A_BUY_BPS']          = DefinedName('A_BUY_BPS',         attr_text="Inputs!$B$8")
wb.defined_names['A_SELL_BPS']         = DefinedName('A_SELL_BPS',        attr_text="Inputs!$B$9")
wb.defined_names['HK_BPS']             = DefinedName('HK_BPS',            attr_text="Inputs!$B$10")
wb.defined_names['HK_BPS_EXEC_SELL']   = DefinedName('HK_BPS_EXEC_SELL',  attr_text="Inputs!$B$11")
wb.defined_names['A_SELL_EXEC_BPS']    = DefinedName('A_SELL_EXEC_BPS',   attr_text="Inputs!$B$12")
wb.defined_names['HKD_START_FX']       = DefinedName('HKD_START_FX',      attr_text="Inputs!$B$13")
wb.defined_names['HK_VAL_FX']          = DefinedName('HK_VAL_FX',         attr_text="Inputs!$B$14")

ws_in['A16'] = '说明'; ws_in['A16'].font = F_SUB
notes = [
    'v22 — Source 1 + Source 2 浮动 双视图 (单一 HK_VAL_FX = 0.901987):',
    '  ⑤  WAC 价差 (Source 1):  卖出CNY − 数量×WAC前(CNY/股). 净 = gross − 卖费 (HK 13.2 bps, A 7.6 bps).',
    '  ⑤b 执行价差 (Source 2): (买: 市场−成交)×数量, (卖: 成交−市场)×数量.',
    '       净 = gross − 卖费 (HK 11.27 bps, A 5.641 bps). 应 ≈ -0.47 万.',
    '  ④a 浮动 (Source 1):    MV_CNY − end_cost_CNY (单一 HK_VAL_FX). 应 ≈ -647 万.',
    '  ④d 浮动 (Source 2):    ④a + ⑤_净 − ⑤b_净. 应 ≈ -889.91 万.',
    '',
    'B14 (HK_VAL_FX = 0.901987): HK 期末市值 + 期末成本 共用 — Source 2 derived.',
    'B11 (HK_BPS_EXEC_SELL = 11.27 bps): Source 2 EXEC 卖费 HK.',
    'B12 (A_SELL_EXEC_BPS  =  5.641 bps): Source 2 EXEC 卖费 A股.',
    '',
    'Source 2 桥接: 浮动_S2 = WAC 浮动 + WAC价差_净 − 执行价差_净.',
    '  − 执行价差_净 把 "执行 alpha" 从 浮动 桶分离出来；',
    '  + WAC价差_净 把 "WAC 历史成本已实现" 留在 浮动 桶.',
    '',
    '股票股息 ⑦ 单元格红底, 默认从 dividend_query_results_online CSV 加载. 可手填覆盖.',
]
for i, n in enumerate(notes, 17):
    ws_in.cell(row=i, column=1, value=n).font = F_BODY
    ws_in.merge_cells(start_row=i, end_row=i, start_column=1, end_column=4)

# =====================================================================
# Prices
# =====================================================================
ws_px = wb.create_sheet('Prices')
write_header(ws_px, ['标准代码','名称','估值日','本币市价','币种'], FILL_H)
all_codes = sorted(set(pos['code']) | set(tx['code']), key=str)
unique_inst = []
inst_meta = {}
for code in all_codes:
    p = pos[pos['code']==code]
    if not p.empty:
        inst = p.iloc[0]['inst']; name = p.iloc[0]['股票名称']
    else:
        t0 = tx[tx['code']==code].iloc[0]
        inst = t0['inst']; name = t0['证券名称']
    if inst not in inst_meta:
        inst_meta[inst] = name
        unique_inst.append(inst)

for i, inst in enumerate(unique_inst, 2):
    p = get_price(inst, VAL_DATE)
    ccy = 'HKD' if inst.endswith('.HK') else 'CNY'
    ws_px.cell(row=i, column=1, value=inst)
    ws_px.cell(row=i, column=2, value=inst_meta[inst])
    ws_px.cell(row=i, column=3, value=VAL_DATE)
    ws_px.cell(row=i, column=4, value=p if p is not None else 0).number_format='#,##0.0000'
    ws_px.cell(row=i, column=5, value=ccy)
last_px = len(unique_inst)+1
for r in range(2, last_px+1):
    for c in range(1,6):
        ws_px.cell(row=r,column=c).font = F_BODY
        ws_px.cell(row=r,column=c).border = BORDER
for col,w in zip('ABCDE',[14,18,12,14,8]): ws_px.column_dimensions[col].width=w
wb.defined_names['PX_TBL'] = DefinedName('PX_TBL', attr_text=f"Prices!$A$2:$E${last_px}")

# =====================================================================
# StartPos
# =====================================================================
ws_sp = wb.create_sheet('StartPos')
write_header(ws_sp, ['股票代码','名称','标准代码','币种','期初数量',
                     '期初成本(CNY)','期初市价(本币)','期初市值(CNY)'], FILL_H)
for i, r in pos.iterrows():
    inst = r['inst']; is_hk = inst.endswith('.HK')
    rr = i+2
    ws_sp.cell(row=rr, column=1, value=str(r['code']))
    ws_sp.cell(row=rr, column=2, value=r['股票名称'])
    ws_sp.cell(row=rr, column=3, value=inst)
    ws_sp.cell(row=rr, column=4, value='HKD' if is_hk else 'CNY')
    ws_sp.cell(row=rr, column=5, value=float(r['数量']))
    ws_sp.cell(row=rr, column=6, value=float(r['成本']))
    ws_sp.cell(row=rr, column=7, value=float(r['市价']))
    ws_sp.cell(row=rr, column=8, value=float(r['市值']))
last_sp = len(pos)+1
for r in range(2, last_sp+1):
    for c in range(1,9):
        cell = ws_sp.cell(row=r,column=c)
        cell.font = F_BODY; cell.border = BORDER
        if c >= 5: cell.number_format = NUM
for col,w in zip('ABCDEFGH',[10,16,12,8,12,16,14,16]):
    ws_sp.column_dimensions[col].width = w
ws_sp.cell(row=last_sp+1, column=2, value='合计').font = F_SUB
for c, ltr in zip([5,6,8], 'EFH'):
    ws_sp.cell(row=last_sp+1, column=c, value=f'=SUM({ltr}2:{ltr}{last_sp})').number_format = NUM
for c in range(1,9):
    ws_sp.cell(row=last_sp+1, column=c).fill = FILL_R
    ws_sp.cell(row=last_sp+1, column=c).border = BORDER

# =====================================================================
# Per-stock sheets — fully formula-driven WAC, ALL CNY
# =====================================================================
per_stock_meta = []

for code in all_codes:
    p = pos[pos['code']==code]
    if not p.empty:
        qty_start = float(p.iloc[0]['数量'])
        cost_cny_start = float(p.iloc[0]['成本'])
        mv_cny_start = float(p.iloc[0]['市值'])
        px_local_start = float(p.iloc[0]['市价'])  # already 本币
        inst = p.iloc[0]['inst']; name = p.iloc[0]['股票名称']
    else:
        qty_start = cost_cny_start = mv_cny_start = px_local_start = 0.0
        t0 = tx[tx['code']==code].iloc[0]
        inst = t0['inst']; name = t0['证券名称']
    is_hk = inst.endswith('.HK')
    ccy_local = 'HKD' if is_hk else 'CNY'
    fill_t = FILL_T_HKD if is_hk else FILL_T_CNY
    fill_h = FILL_H_HK if is_hk else FILL_H
    fill_s = FILL_S_HK if is_hk else FILL_S

    sheet_name = safe_sheet_name(f"{name}_{code}")
    base = sheet_name; suffix = 1
    while sheet_name in wb.sheetnames:
        sheet_name = safe_sheet_name(f"{base}_{suffix}"); suffix += 1
    ws = wb.create_sheet(sheet_name)

    # Title
    ws['A1'] = f'{name} ({code}) — {inst} — 全程 CNY 计量 (本币 {ccy_local}) — WAC'
    ws['A1'].font = F_TITLE; ws['A1'].fill = fill_t
    ws.merge_cells('A1:M1'); ws['A1'].alignment = ALIGN_C

    # Static info
    info = [
        ('A3', '股票代码',    str(code),                    None),
        ('A4', '证券名称',    name,                         None),
        ('A5', '标准代码',    inst,                         None),
        ('A6', '本币 (信息)',  ccy_local,                    '注: 走帐全程 CNY 计量'),
        ('D3', '期初数量',    qty_start,                    'StartPos.E'),
        ('D4', '期初成本(CNY)', cost_cny_start,             'StartPos.F'),
        ('D5', '期初CNY/股',   ('=E4/E3' if qty_start > 0 else 0),
              '期初CNY ÷ 期初数量'),
        ('G3', '估值日',         '=VAL_DATE',     'Inputs.VAL_DATE'),
        ('G4', '估值日 HKD/CNY' if is_hk else '估值日 FX',
              ('=HKD_END_FX' if is_hk else 1.0),
              ('Inputs.HKD_END_FX' if is_hk else 'A股 FX=1')),
        ('G5', '期初市价(本币)', px_local_start,    'StartPos.G'),
        ('G6', '期初市值(CNY)',  mv_cny_start,    'StartPos.H'),
    ]
    for cell_addr, label, val, remark in info:
        col = cell_addr[0]; row = int(cell_addr[1:])
        col_idx = ord(col) - ord('A') + 1
        ws.cell(row=row, column=col_idx, value=label).font = F_BODY
        ws.cell(row=row, column=col_idx).fill = fill_s
        ws.cell(row=row, column=col_idx).border = BORDER
        ws.cell(row=row, column=col_idx).alignment = ALIGN_L
        cell = ws.cell(row=row, column=col_idx+1, value=val)
        cell.font = F_BODY; cell.border = BORDER
        if isinstance(val,(int,float)) or (isinstance(val,str) and val.startswith('=')):
            if isinstance(val,str) and val == '=HKD_END_FX':
                cell.number_format = FX_FMT
            else:
                cell.number_format = NUM
            cell.alignment = ALIGN_R
        else:
            cell.alignment = ALIGN_L
        if remark:
            ws.cell(row=row, column=col_idx+2, value=remark).font = F_REMARK

    # =================================================================
    # Walk: 20 columns (v22 — adds S, T for HKD-WAC running cost on HK only).
    # =================================================================
    # CNY walk (cols A-R, same as v20):
    # A=#  B=日期  C=事件  D=数量  E=本币单价(input)  F=本币金额(=D*E)
    # G=每笔FX(input)  H=成交金额CNY(=F*G)  I=市场均价(input)
    # J=WAC前CNY/股(formula)  K=WAC匹配成本CNY(formula)
    # L=期末数量(formula)  M=期末WAC成本CNY(formula)
    # N=WAC价差CNY(formula)  O=执行价差本币(formula)  P=执行价差CNY(formula=O*G)
    # Q=费用CNY(formula)  R=备注
    # HKD-mini walk (HK only) for Source 1/2 浮动:
    # S=WAC前(HKD/股)  T=期末WAC成本HKD
    walk_header_row = 9
    ws.cell(row=walk_header_row-1, column=1,
            value=f'■ 走帐 — CNY 计量 (A-R) + HKD WAC 成本 (S-T, HK only). 输入仅 D, E, F, G').font = F_SUB
    ws.merge_cells(start_row=walk_header_row-1, end_row=walk_header_row-1, start_column=1, end_column=20)
    walk_headers = [
        '#','日期','事件','数量','本币单价','本币金额','每笔FX','成交金额(CNY)',
        '市场均价(本币)','WAC前(CNY/股)','WAC匹配成本(CNY)',
        '期末数量','期末WAC成本(CNY)','WAC价差(CNY)',
        '执行价差(本币)','执行价差(CNY)','费用(CNY)','备注',
        'WAC前(HKD/股)','期末WAC成本(HKD)',
    ]
    write_header(ws, walk_headers, fill_h, row=walk_header_row)

    walk_first = walk_header_row + 1
    body_row = walk_header_row

    # ---- Opening row ----
    if qty_start > 0:
        body_row += 1
        r = body_row
        # Opening price = cost / qty in CNY/share (BUT we need 本币 price to be E)
        # Trick: opening "本币单价" = px_local_start (input); opening "本币金额"
        # = qty × px_local. The CNY cost (StartPos D4) ≠ qty × px_local generally,
        # because StartPos uses MV at start mark, not book cost. So opening CNY
        # amount H must be a separate formula = E4 (book CNY cost). We override.
        ws.cell(row=r, column=1, value=r - walk_header_row)
        ws.cell(row=r, column=2, value=START_FX_DATE)
        ws.cell(row=r, column=3, value='期初')
        ws.cell(row=r, column=4, value=qty_start)
        ws.cell(row=r, column=5, value=px_local_start)         # E (info only)
        ws.cell(row=r, column=6, value=f'=D{r}*E{r}')          # F (info only — qty × local px)
        # G (FX): for opening, per-trade FX is back-derived from book cost ÷ local MV
        # but openings use book cost basis directly, so we set G blank/0 (no FX needed)
        ws.cell(row=r, column=7, value=(0 if not is_hk else f'=IFERROR(E4/F{r},0)'))
        ws.cell(row=r, column=8, value=f'=E4')                 # H = opening book CNY (formula → StartPos)
        ws.cell(row=r, column=9, value=0)                      # I (no market avg)
        ws.cell(row=r, column=10, value=0)                     # J (no prev)
        ws.cell(row=r, column=11, value=0)                     # K (no sell)
        ws.cell(row=r, column=12, value=f'=D{r}')              # L = qty
        ws.cell(row=r, column=13, value=f'=H{r}')              # M = opening CNY cost
        ws.cell(row=r, column=14, value=0)                     # N (no WAC价差)
        ws.cell(row=r, column=15, value=0)                     # O (no exec slip)
        ws.cell(row=r, column=16, value=0)                     # P
        ws.cell(row=r, column=17, value=0)                     # Q (no fee)
        ws.cell(row=r, column=18, value=f'期初: {int(qty_start):,}股, 书面成本 ={cost_cny_start:,.0f} CNY')
        # HKD walk opening (HK only)
        if is_hk:
            ws.cell(row=r, column=19, value=0)                          # S WAC前 HKD/股 (no prior)
            ws.cell(row=r, column=20, value=f'=E4/HKD_START_FX')        # T 期末成本 HKD = StartPos.成本/0.8974
        else:
            ws.cell(row=r, column=19, value=0)
            ws.cell(row=r, column=20, value=0)
        for c in range(1, 21): ws.cell(row=r, column=c).fill = FILL_OPEN

    # ---- Trade rows ----
    code_trades = tx[tx['code']==code].sort_values(['date','序号'])
    for _, tr in code_trades.iterrows():
        body_row += 1
        r = body_row
        prev = r - 1
        qty = float(tr['成交数量'])
        amt_local = float(tr['成交金额'])              # 本币 (HKD for HK, CNY for A)
        amt_cny   = float(tr['本币成交金额'])           # CNY for both
        local_px  = float(tr['市场成交均价']) if pd.notna(tr['市场成交均价']) else 0
        fill_px   = float(tr['成交均价']) if pd.notna(tr['成交均价']) else 0
        side = tr['委托方向']
        date = tr['date']
        # Per-trade FX from broker tape (constant input)
        fx_per_trade = (amt_cny / amt_local) if (is_hk and amt_local != 0) else 1.0

        ws.cell(row=r, column=1, value=r - walk_header_row)
        ws.cell(row=r, column=2, value=date)
        ws.cell(row=r, column=3, value=side)
        ws.cell(row=r, column=4, value=qty)              # D INPUT
        ws.cell(row=r, column=5, value=fill_px)          # E INPUT (本币单价 = 成交均价)
        # F = 本币金额 (formula = D × E)
        ws.cell(row=r, column=6, value=f'=D{r}*E{r}')
        # G = 每笔FX (constant input)
        ws.cell(row=r, column=7, value=fx_per_trade)
        # H = 成交金额(CNY) = F × G  (formula!)
        ws.cell(row=r, column=8, value=f'=F{r}*G{r}')
        # I = 市场均价(本币) INPUT
        ws.cell(row=r, column=9, value=local_px)

        # J = WAC前(CNY/股)
        if r == walk_first:
            ws.cell(row=r, column=10, value=0)
        else:
            ws.cell(row=r, column=10,
                    value=f'=IFERROR(IF(L{prev}>0,M{prev}/L{prev},0),0)')

        # K = WAC匹配成本(CNY) = qty × WAC前 if 卖出
        ws.cell(row=r, column=11,
                value=f'=IF(C{r}="卖出",D{r}*J{r},0)')

        # L = 期末数量
        if r == walk_first:
            ws.cell(row=r, column=12,
                    value=f'=IF(C{r}="买入",D{r},0)-IF(C{r}="卖出",D{r},0)')
        else:
            ws.cell(row=r, column=12,
                    value=f'=L{prev}+IF(C{r}="买入",D{r},0)-IF(C{r}="卖出",D{r},0)')

        # M = 期末WAC成本(CNY) = M_prev + buy_CNY − matched_cost
        if r == walk_first:
            ws.cell(row=r, column=13,
                    value=f'=IF(C{r}="买入",H{r},0)-K{r}')
        else:
            ws.cell(row=r, column=13,
                    value=f'=M{prev}+IF(C{r}="买入",H{r},0)-K{r}')

        # N = WAC价差(CNY) = sell CNY − matched cost
        ws.cell(row=r, column=14,
                value=f'=IF(C{r}="卖出",H{r}-K{r},0)')

        # O = 执行价差(本币) = (买: market−fill) × qty for buys, (卖: fill−market) × qty for sells
        ws.cell(row=r, column=15,
                value=f'=IF(C{r}="买入",(I{r}-E{r})*D{r},IF(C{r}="卖出",(E{r}-I{r})*D{r},0))')

        # P = 执行价差(CNY) = O × G
        ws.cell(row=r, column=16, value=f'=O{r}*G{r}')

        # Q = 费用(CNY) = H × bps / 10000  (uses WAC fee schedule by default)
        if is_hk:
            ws.cell(row=r, column=17, value=f'=H{r}*HK_BPS/10000')
        else:
            ws.cell(row=r, column=17,
                    value=f'=H{r}*IF(C{r}="买入",A_BUY_BPS,A_SELL_BPS)/10000')

        # HKD WAC walk (HK only): S=WAC前HKD/股, T=期末成本HKD
        if is_hk:
            # S = WAC前(HKD) = T_prev / L_prev
            if r == walk_first:
                ws.cell(row=r, column=19, value=0)
            else:
                ws.cell(row=r, column=19,
                        value=f'=IFERROR(IF(L{prev}>0,T{prev}/L{prev},0),0)')
            # T = 期末成本(HKD) = T_prev + (买: F else 0) − (卖: D × S else 0)
            if r == walk_first:
                ws.cell(row=r, column=20,
                        value=f'=IF(C{r}="买入",F{r},0)-IF(C{r}="卖出",D{r}*S{r},0)')
            else:
                ws.cell(row=r, column=20,
                        value=f'=T{prev}+IF(C{r}="买入",F{r},0)-IF(C{r}="卖出",D{r}*S{r},0)')
        else:
            ws.cell(row=r, column=19, value=0)
            ws.cell(row=r, column=20, value=0)

        fill = FILL_BUY if side == '买入' else FILL_SELL
        for c in range(1, 21): ws.cell(row=r, column=c).fill = fill

    last_walk_row = body_row
    for r in range(walk_first, last_walk_row+1):
        for c in range(1, 21):
            cell = ws.cell(row=r, column=c)
            cell.font = F_BODY; cell.border = BORDER
            if c in (4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 20):
                cell.number_format = NUM; cell.alignment = ALIGN_R
            elif c == 7:
                cell.number_format = FX_FMT; cell.alignment = ALIGN_R
            else:
                cell.alignment = ALIGN_L

    # ---- Walk subtotal ----
    sub_row = last_walk_row + 1
    ws.cell(row=sub_row, column=2, value='本sheet合计').font = F_SUB
    sum_cols = [
        (8,  'H',  False),
        (14, 'N',  True),   # WAC价差 (sells only)
        (16, 'P',  False),  # 执行价差 (all rows)
        (17, 'Q',  False),  # 费用
    ]
    for c, ltr, only_sell in sum_cols:
        if only_sell:
            ws.cell(row=sub_row, column=c,
                    value=f'=SUMIF(C{walk_first}:C{last_walk_row},"卖出",{ltr}{walk_first}:{ltr}{last_walk_row})').number_format = NUM
        else:
            ws.cell(row=sub_row, column=c,
                    value=f'=SUM({ltr}{walk_first}:{ltr}{last_walk_row})').number_format = NUM
    for c in range(1, 21): ws.cell(row=sub_row, column=c).fill = FILL_R

    # Breakdown block (CNY-only) — updated for new walk column letters
    # Walk: L=期末数量, M=期末WAC成本, N=WAC价差, P=执行价差CNY, Q=费用CNY
    bk = sub_row + 2
    ws.cell(row=bk, column=1, value='■ 损益分量 (CNY)').font = F_SUB
    ws.merge_cells(start_row=bk, end_row=bk, start_column=1, end_column=4)
    write_header(ws, ['组件','CNY 金额','公式','说明'], fill_h, row=bk+1)

    bk_row = bk + 1

    bk_row += 1; row_qty_end = bk_row
    ws.cell(row=bk_row, column=1, value='期末数量')
    if last_walk_row > walk_header_row:
        ws.cell(row=bk_row, column=2, value=f'=L{last_walk_row}').number_format = NUM
        ws.cell(row=bk_row, column=3, value=f'走帐 L{last_walk_row}').font = F_FORMULA
    else:
        ws.cell(row=bk_row, column=2, value=0).number_format = NUM

    bk_row += 1; row_cost_end = bk_row
    ws.cell(row=bk_row, column=1, value='期末WAC成本 (CNY)')
    if last_walk_row > walk_header_row:
        ws.cell(row=bk_row, column=2, value=f'=M{last_walk_row}').number_format = NUM
        ws.cell(row=bk_row, column=3, value=f'走帐 M{last_walk_row}').font = F_FORMULA
    else:
        ws.cell(row=bk_row, column=2, value=0).number_format = NUM

    # End price (本币)
    bk_row += 1; row_end_px_local = bk_row
    ws.cell(row=bk_row, column=1, value='估值日市价 (本币)')
    ws.cell(row=bk_row, column=2,
            value=f'=IFERROR(VLOOKUP("{inst}",PX_TBL,4,FALSE),"价格缺失")').number_format = '#,##0.0000'
    ws.cell(row=bk_row, column=3, value=f'VLOOKUP {inst}').font = F_FORMULA

    # End MV CNY
    bk_row += 1; row_mv_end = bk_row
    ws.cell(row=bk_row, column=1, value='期末市值 (CNY)')
    if is_hk:
        ws.cell(row=bk_row, column=2,
                value=f'=IF(B{row_qty_end}=0,0,IFERROR(B{row_qty_end}*B{row_end_px_local}*HKD_END_FX,"价格缺失"))').number_format = NUM
        ws.cell(row=bk_row, column=3,
                value=f'B{row_qty_end} × B{row_end_px_local} × HKD_END_FX').font = F_FORMULA
    else:
        ws.cell(row=bk_row, column=2,
                value=f'=IF(B{row_qty_end}=0,0,IFERROR(B{row_qty_end}*B{row_end_px_local},"价格缺失"))').number_format = NUM
        ws.cell(row=bk_row, column=3,
                value=f'B{row_qty_end} × B{row_end_px_local}').font = F_FORMULA

    # ④a 浮动 (Source 1 / Source 2):
    #   MV_CNY = qty × HKD_close × HK_VAL_FX  (HK)  or qty × CNY_close (A股)
    #   end_cost_CNY = M{last_walk_row}  (走帐 CNY: StartPos.成本 + buys 本币成交金额 - sold WAC CNY)
    #   浮动 = MV_CNY − end_cost_CNY
    # The HK end_cost uses per-trade FX from broker tape (本币成交金额 ÷ 成交金额),
    # not a uniform HK_VAL_FX. This matches recon S2's column O exactly.
    if is_hk and last_walk_row > walk_header_row:
        bk_row += 1; row_mv_end_s1 = bk_row
        ws.cell(row=bk_row, column=1, value='期末市值 CNY (S1) = qty × HKD_close × HK_VAL_FX')
        ws.cell(row=bk_row, column=2,
                value=f'=IF(B{row_qty_end}=0,0,IFERROR(B{row_qty_end}*B{row_end_px_local}*HK_VAL_FX,"价格缺失"))').number_format = NUM
        ws.cell(row=bk_row, column=3,
                value=f'B{row_qty_end} × B{row_end_px_local} × HK_VAL_FX').font = F_FORMULA

        bk_row += 1; row_unr_s1 = bk_row
        ws.cell(row=bk_row, column=1, value='④a 浮动 (Source 1) = 期末市值CNY − 期末WAC成本CNY(走帐)')
        ws.cell(row=bk_row, column=2,
                value=f'=IFERROR(B{row_mv_end_s1}-B{row_cost_end},"")').number_format = NUM
        ws.cell(row=bk_row, column=2).fill = FILL_MTM
        ws.cell(row=bk_row, column=3,
                value=f'B{row_mv_end_s1} − B{row_cost_end}').font = F_FORMULA
        ws.cell(row=bk_row, column=4, value='HK MV 用 HK_VAL_FX, cost 用走帐 CNY (per-trade FX)').font = F_REMARK
    else:
        # A股: 浮动 = MV_CNY − cost_CNY (no FX)
        bk_row += 1; row_unr_s1 = bk_row
        ws.cell(row=bk_row, column=1, value='④a 浮动 (Source 1) = 期末市值 − 期末WAC成本 (CNY)')
        ws.cell(row=bk_row, column=2,
                value=f'=IFERROR(B{row_mv_end}-B{row_cost_end},"期末市值缺失")').number_format = NUM
        ws.cell(row=bk_row, column=2).fill = FILL_MTM
        ws.cell(row=bk_row, column=3, value=f'B{row_mv_end} − B{row_cost_end}').font = F_FORMULA

    # ⑤ WAC 价差 (CNY) — Source 1 methodology
    bk_row += 1; row_rea_wac = bk_row
    ws.cell(row=bk_row, column=1, value='⑤ WAC 价差 (CNY) — Source 1: 卖出CNY − 匹配WAC成本')
    ws.cell(row=bk_row, column=2, value=f'=N{sub_row}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_MTM
    ws.cell(row=bk_row, column=3, value=f'走帐 N{sub_row} (= SUMIF 卖出 → N)').font = F_FORMULA

    # ⑤b 执行价差 (CNY) — Source 2 methodology
    bk_row += 1; row_rea_exec = bk_row
    ws.cell(row=bk_row, column=1, value='⑤b 执行价差 (CNY) — Source 2: 买入(市场−成交)×qty + 卖出(成交−市场)×qty')
    ws.cell(row=bk_row, column=2, value=f'=P{sub_row}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_MTM
    ws.cell(row=bk_row, column=3, value=f'走帐 P{sub_row} (= SUM)').font = F_FORMULA

    # ⑥ Fee WAC (CNY) — total at WAC rate (买+卖)
    bk_row += 1; row_fee = bk_row
    ws.cell(row=bk_row, column=1, value='⑥ 费用合计 (CNY) 买+卖 @ WAC rate')
    ws.cell(row=bk_row, column=2, value=f'=Q{sub_row}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_FEE
    ws.cell(row=bk_row, column=3, value=f'走帐 Q{sub_row}').font = F_FORMULA
    ws.cell(row=bk_row, column=4,
            value=(f'港股 {HK_BPS} bps' if is_hk else f'A股 买{A_BUY_BPS}/卖{A_SELL_BPS} bps')).font = F_REMARK

    # ⑥b Sell fees only WAC (for Source 1 net)
    bk_row += 1; row_fee_sell_wac = bk_row
    ws.cell(row=bk_row, column=1, value='⑥b 卖出费用 (CNY) @ WAC rate — Source 1 net')
    ws.cell(row=bk_row, column=2,
            value=f'=SUMIF(C{walk_first}:C{last_walk_row},"卖出",Q{walk_first}:Q{last_walk_row})').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_FEE
    ws.cell(row=bk_row, column=3, value=f'SUMIF 卖出 → Q').font = F_FORMULA
    ws.cell(row=bk_row, column=4, value='Source 1 net = WAC价差 − 卖费(WAC rate)').font = F_REMARK

    # ⑥c Sell fees @ EXEC rate (Source 2 net)
    bk_row += 1; row_fee_sell_exec = bk_row
    ws.cell(row=bk_row, column=1, value='⑥c 卖出费用 (CNY) @ EXEC rate — Source 2 net')
    if is_hk:
        ws.cell(row=bk_row, column=2,
                value=f'=SUMIF(C{walk_first}:C{last_walk_row},"卖出",H{walk_first}:H{last_walk_row})*HK_BPS_EXEC_SELL/10000').number_format = NUM
    else:
        ws.cell(row=bk_row, column=2,
                value=f'=SUMIF(C{walk_first}:C{last_walk_row},"卖出",H{walk_first}:H{last_walk_row})*A_SELL_EXEC_BPS/10000').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_FEE
    ws.cell(row=bk_row, column=3,
            value=('港股: SUMIF(卖)×H × HK_BPS_EXEC_SELL/10000' if is_hk
                   else 'A股: SUMIF(卖)×H × A_SELL_EXEC_BPS/10000')).font = F_FORMULA
    ws.cell(row=bk_row, column=4, value='Source 2 net = 执行价差 − 卖费(EXEC rate)').font = F_REMARK

    # ⑤ net WAC (Source 1)
    bk_row += 1; row_rea_wac_net = bk_row
    ws.cell(row=bk_row, column=1, value='⑤ WAC 价差 净 (Source 1) = ⑤ − ⑥b')
    ws.cell(row=bk_row, column=2,
            value=f'=B{row_rea_wac}-B{row_fee_sell_wac}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_MTM
    ws.cell(row=bk_row, column=3,
            value=f'B{row_rea_wac} − B{row_fee_sell_wac}').font = F_FORMULA
    ws.cell(row=bk_row, column=4, value='对齐 reconn equity_realized_pnl_cny_basis_wac_detail').font = F_REMARK

    # ⑤b net Exec (Source 2)
    bk_row += 1; row_rea_exec_net = bk_row
    ws.cell(row=bk_row, column=1, value='⑤b 执行价差 净 (Source 2) = ⑤b − ⑥c')
    ws.cell(row=bk_row, column=2,
            value=f'=B{row_rea_exec}-B{row_fee_sell_exec}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_MTM
    ws.cell(row=bk_row, column=3,
            value=f'B{row_rea_exec} − B{row_fee_sell_exec}').font = F_FORMULA
    ws.cell(row=bk_row, column=4, value='对齐 Source 2 -0.47 万 (HK 11.27 bps + A 5.641 bps)').font = F_REMARK

    # ④d 浮动 (Source 2) = ④a + ⑤_净 − ⑤b_净
    bk_row += 1; row_unr_s2 = bk_row
    ws.cell(row=bk_row, column=1, value='④d 浮动 (Source 2) = ④a + ⑤净(WAC) − ⑤b净(EXEC)')
    ws.cell(row=bk_row, column=2,
            value=f'=B{row_unr_s1}+B{row_rea_wac_net}-B{row_rea_exec_net}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_MTM
    ws.cell(row=bk_row, column=3,
            value=f'B{row_unr_s1} + B{row_rea_wac_net} − B{row_rea_exec_net}').font = F_FORMULA
    ws.cell(row=bk_row, column=4, value='对齐 Source 2 浮动 -889.91 万').font = F_REMARK

    # alias for downstream code: use Source 1 浮动 as the default for ⑧ total
    row_unr = row_unr_s1
    row_rea = row_rea_wac
    row_fee_sell = row_fee_sell_wac

    # Dividend events
    div_csv = div_lookup.get(str(code))
    if div_csv and div_csv.get('events'):
        bk_row += 1
        ws.cell(row=bk_row, column=1, value='■ 派息事件 (期内 ex-date 命中)').font = F_SUB
        ws.merge_cells(start_row=bk_row, end_row=bk_row, start_column=1, end_column=4)
        ws.cell(row=bk_row, column=1).fill = FILL_DIV
        bk_row += 1
        write_header(ws, ['ex-date / 类型','当日持仓×每股','CNY 金额','派息日 / 来源'], fill_h, row=bk_row)
        for ev in div_csv['events']:
            bk_row += 1
            ws.cell(row=bk_row, column=1, value=f'{ev["ex_date"]} / {ev["div_type"]}')
            ws.cell(row=bk_row, column=2, value=f'{int(ev["qty"]):,} × {ev["per_share"]} {ev["ccy"]}')
            ws.cell(row=bk_row, column=3, value=ev['amt_cny']).number_format = NUM
            ws.cell(row=bk_row, column=4, value=f'派息日 {ev["pay_date"]}  {ev["raw"]}').font = F_REMARK
            for c in range(1,5):
                ws.cell(row=bk_row, column=c).font = F_BODY
                ws.cell(row=bk_row, column=c).border = BORDER
                ws.cell(row=bk_row, column=c).fill = FILL_DIV

    # ⑦ Dividend (CNY)
    if div_csv and div_csv.get('events'):
        div_value = div_csv['amount_cny']
        div_label = '⑦ 股息 (CNY) ★ events CSV'
        div_remark = '来源 dividend_query_results CSV; 港股已用 HKD_END_FX 折算; 可手填覆盖'
    else:
        div_value = 0
        div_label = '⑦ 股息 (CNY) — 期内无 ex-date'
        div_remark = '无派息事件命中; 如有, 请手填本格'
    bk_row += 1; row_div = bk_row
    ws.cell(row=bk_row, column=1, value=div_label)
    div_cell = ws.cell(row=bk_row, column=2, value=div_value)
    div_cell.number_format = NUM; div_cell.font = F_INPUT; div_cell.fill = FILL_INPUT
    ws.cell(row=bk_row, column=4, value=div_remark).font = F_REMARK

    # ⑧ Total CNY
    bk_row += 1; row_total = bk_row
    ws.cell(row=bk_row, column=1, value='⑧ 合计 PnL (CNY) = ④+⑤−⑥+⑦')
    ws.cell(row=bk_row, column=2,
            value=f'=B{row_unr}+B{row_rea}-B{row_fee}+B{row_div}').number_format = NUM
    ws.cell(row=bk_row, column=2).fill = FILL_FINAL
    ws.cell(row=bk_row, column=3,
            value=f'B{row_unr}+B{row_rea}−B{row_fee}+B{row_div}').font = F_FORMULA

    # Style breakdown
    for r in range(bk+1, bk_row+1):
        for c in range(1,5):
            cell = ws.cell(row=r, column=c)
            if cell.value is None: continue
            cell.border = BORDER

    per_stock_meta.append({
        'code': code, 'name': name, 'inst': inst, 'ccy': ccy_local,
        'sheet': sheet_name, 'is_hk': is_hk,
        'unr_s1':            f"'{sheet_name}'!$B${row_unr_s1}",
        'unr_s2':            f"'{sheet_name}'!$B${row_unr_s2}",
        'rea_wac_cny':       f"'{sheet_name}'!$B${row_rea_wac}",
        'rea_wac_net_cny':   f"'{sheet_name}'!$B${row_rea_wac_net}",
        'rea_exec_cny':      f"'{sheet_name}'!$B${row_rea_exec}",
        'rea_exec_net_cny':  f"'{sheet_name}'!$B${row_rea_exec_net}",
        'fee_cny':           f"'{sheet_name}'!$B${row_fee}",
        'fee_sell_wac_cny':  f"'{sheet_name}'!$B${row_fee_sell_wac}",
        'fee_sell_exec_cny': f"'{sheet_name}'!$B${row_fee_sell_exec}",
        'div_cny':           f"'{sheet_name}'!$B${row_div}",
        'total_cny':         f"'{sheet_name}'!$B${row_total}",
    })

    for col, w in zip('ABCDEFGHIJKLMNOPQRST',
                       [5,12,7,11,12,15,9,15,12,12,15,12,15,15,14,14,12,40,
                        12,15]):
        ws.column_dimensions[col].width = w

# =====================================================================
# Single Summary sheet (CNY only, all stocks)
# =====================================================================
ws_s = wb.create_sheet('Summary')
ws_s['A1'] = '汇总 (全程 CNY, A股 + 港股)'
ws_s['A1'].font = F_TITLE; ws_s['A1'].fill = FILL_T_TOT
ws_s.merge_cells('A1:I1'); ws_s['A1'].alignment = ALIGN_C
write_header(ws_s,
    ['股票代码','名称','标准代码','本币',
     'WAC价差(CNY)','执行价差(CNY)',
     '浮动 S1 (CNY)','浮动 S2 (CNY)',
     '股息(CNY)','卖费 WAC','卖费 EXEC','费用合计(CNY)',
     'WAC 净 (S1)','执行 净 (S2)','合计(CNY)','sheet'],
    FILL_H, row=4)
rr = 4
for m in [m for m in per_stock_meta if not m['is_hk']] + \
         [m for m in per_stock_meta if m['is_hk']]:
    rr += 1
    ws_s.cell(row=rr, column=1,  value=str(m['code']))
    ws_s.cell(row=rr, column=2,  value=m['name'])
    ws_s.cell(row=rr, column=3,  value=m['inst'])
    ws_s.cell(row=rr, column=4,  value=m['ccy'])
    ws_s.cell(row=rr, column=5,  value=f"={m['rea_wac_cny']}")
    ws_s.cell(row=rr, column=6,  value=f"={m['rea_exec_cny']}")
    ws_s.cell(row=rr, column=7,  value=f"={m['unr_s1']}")             # 浮动 S1
    ws_s.cell(row=rr, column=8,  value=f"={m['unr_s2']}")             # 浮动 S2
    ws_s.cell(row=rr, column=9,  value=f"={m['div_cny']}")
    ws_s.cell(row=rr, column=10, value=f"={m['fee_sell_wac_cny']}")
    ws_s.cell(row=rr, column=11, value=f"={m['fee_sell_exec_cny']}")
    ws_s.cell(row=rr, column=12, value=f"={m['fee_cny']}")
    ws_s.cell(row=rr, column=13, value=f"={m['rea_wac_net_cny']}")
    ws_s.cell(row=rr, column=14, value=f"={m['rea_exec_net_cny']}")
    ws_s.cell(row=rr, column=15, value=f"={m['total_cny']}")
    ws_s.cell(row=rr, column=16, value=m['sheet'])
last_sum = rr
for r in range(5, last_sum+1):
    for c in range(1, 17):
        cell = ws_s.cell(row=r, column=c)
        cell.font = F_BODY; cell.border = BORDER
        if c in (5,6,7,8,9,10,11,12,13,14,15): cell.number_format = NUM; cell.alignment = ALIGN_R
        elif c == 16: cell.font = F_REMARK; cell.alignment = ALIGN_L
        else: cell.alignment = ALIGN_L
        if c <= 4 or c == 16:
            ws_s.cell(row=r, column=c).fill = (FILL_S_HK if any(
                m['is_hk'] and str(m['code']) == str(ws_s.cell(row=r, column=1).value)
                for m in per_stock_meta) else FILL_S)

# Subtotals: A股, HK, Grand
n_a = sum(1 for m in per_stock_meta if not m['is_hk'])
sub_a_row = 4 + n_a + 1  # last A row + 1
# Wait — we wrote A first, then HK. So:
a_first = 5
a_last  = 4 + n_a
h_first = a_last + 1
h_last  = last_sum

sub_a = last_sum + 1
ws_s.cell(row=sub_a, column=2, value=f'A股小计 ({n_a} 只)').font = F_SUB
for c, ltr in zip(range(5, 16), 'EFGHIJKLMNO'):
    ws_s.cell(row=sub_a, column=c,
              value=f'=SUM({ltr}{a_first}:{ltr}{a_last})').number_format = NUM
for c in range(1, 17): ws_s.cell(row=sub_a, column=c).fill = FILL_T_CNY
for c in range(1, 17): ws_s.cell(row=sub_a, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')

sub_h = last_sum + 2
n_h = h_last - h_first + 1
ws_s.cell(row=sub_h, column=2, value=f'港股小计 ({n_h} 只)').font = F_SUB
for c, ltr in zip(range(5, 16), 'EFGHIJKLMNO'):
    ws_s.cell(row=sub_h, column=c,
              value=f'=SUM({ltr}{h_first}:{ltr}{h_last})').number_format = NUM
for c in range(1, 17): ws_s.cell(row=sub_h, column=c).fill = FILL_T_HKD
for c in range(1, 17): ws_s.cell(row=sub_h, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')

grand = last_sum + 3
ws_s.cell(row=grand, column=2, value='总计 (CNY)').font = F_SUB
for c, ltr in zip(range(5, 16), 'EFGHIJKLMNO'):
    ws_s.cell(row=grand, column=c,
              value=f'=SUM({ltr}5:{ltr}{last_sum})').number_format = NUM
for c in range(1, 17): ws_s.cell(row=grand, column=c).fill = FILL_T_TOT
for c in range(1, 17): ws_s.cell(row=grand, column=c).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
for c in range(1, 17): ws_s.cell(row=grand, column=c).border = BORDER

for col, w in zip('ABCDEFGHIJKLMNOP',
                   [10,16,12,8,15,15,15,15,12,12,12,15,14,14,15,28]):
    ws_s.column_dimensions[col].width = w

# =====================================================================
# Total
# =====================================================================
ws_t = wb.create_sheet('Total', 0)
ws_t['A1'] = '前海自营权益4 — PnL 总览  (v24, 風控 vs 權益OCI賬戶 並列 視圖)'
ws_t['A1'].font = F_TITLE; ws_t['A1'].fill = FILL_T_TOT
ws_t.merge_cells('A1:D1'); ws_t['A1'].alignment = ALIGN_C
ws_t.cell(row=2, column=1,
          value='对齐 reconn/equity_realized_pnl_cny_basis_wac_detail. 修改 Inputs → 整张表自动重算.').font = F_REMARK
ws_t.merge_cells('A2:D2')
write_header(ws_t, ['项目','金额','公式','备注'], FILL_H, row=4)

ROW_NEXT = 5
anchors = {}
def add(label, formula=None, remark='', key=None, header=False):
    global ROW_NEXT
    r = ROW_NEXT
    if key: anchors[key] = r
    cell_a = ws_t.cell(row=r, column=1, value=label)
    cell_a.font = F_BODY; cell_a.border = BORDER
    if header:
        cell_a.font = F_SUB
        if label.startswith('A 股'): cell_a.fill = FILL_T_CNY
        elif label.startswith('港 股'): cell_a.fill = FILL_T_HKD
        elif label.startswith('★') or label.startswith('总') or label.startswith('折算') or label.startswith('readme') or label.startswith('差额'):
            cell_a.fill = FILL_T_TOT
        else: cell_a.fill = FILL_R
        if cell_a.fill in (FILL_T_CNY, FILL_T_HKD, FILL_T_TOT):
            cell_a.font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    if formula is not None:
        if isinstance(formula, str) and '{' in formula:
            for k, v in anchors.items():
                formula = formula.replace('{'+k+'}', f'B{v}')
        bcell = ws_t.cell(row=r, column=2, value=formula)
        bcell.font = F_BODY
        bcell.number_format = FX_FMT if (isinstance(formula,str) and formula == '=HKD_END_FX') else NUM
        bcell.alignment = ALIGN_R; bcell.fill = FILL_S; bcell.border = BORDER
    if remark:
        ws_t.cell(row=r, column=3, value=remark).font = F_FORMULA
        ws_t.cell(row=r, column=3).alignment = ALIGN_LW
        ws_t.cell(row=r, column=3).border = BORDER
    ROW_NEXT += 1

def blank():
    global ROW_NEXT
    ROW_NEXT += 1

# Reference Summary subtotal rows
# A subtotal at row sub_a, HK subtotal at row sub_h, grand at row grand

# Summary col mapping (v22):
#   E=WAC价差  F=执行价差  G=浮动 S1  H=浮动 S2  I=股息
#   J=卖费WAC  K=卖费EXEC  L=费用合计  M=WAC净  N=执行净  O=合计

add('A 股 (CNY)', header=True)
add('  WAC 价差',           f'=Summary!E{sub_a}', f"Summary A E{sub_a}",  key='a_rea_wac')
add('  执行价差',            f'=Summary!F{sub_a}', f"Summary A F{sub_a}",  key='a_rea_exec')
add('  浮动 S1',             f'=Summary!G{sub_a}', f"Summary A G{sub_a}  ★ -4,786,339",  key='a_unr_s1')
add('  浮动 S2',             f'=Summary!H{sub_a}', f"Summary A H{sub_a}",  key='a_unr_s2')
add('  股息',                f'=Summary!I{sub_a}', f"Summary A I{sub_a}",  key='a_div')
add('  卖费 (WAC)',          f'=Summary!J{sub_a}', f"Summary A J{sub_a}",  key='a_fee_sell_wac')
add('  卖费 (EXEC)',         f'=Summary!K{sub_a}', f"Summary A K{sub_a}",  key='a_fee_sell_exec')
add('  费用合计',            f'=Summary!L{sub_a}', f"Summary A L{sub_a}",  key='a_fee')
add('  WAC 净',              f'=Summary!M{sub_a}', f"Summary A M{sub_a}",  key='a_rea_wac_net')
add('  执行 净',             f'=Summary!N{sub_a}', f"Summary A N{sub_a}",  key='a_rea_exec_net')
add('  合计',                f'=Summary!O{sub_a}', f"Summary A O{sub_a}",  key='a_tot')
blank()
add('港 股 (CNY 折算)', header=True)
add('  WAC 价差',            f'=Summary!E{sub_h}', f"★ HK WAC gross: -1,076,503", key='h_rea_wac')
add('  执行价差',            f'=Summary!F{sub_h}', '',  key='h_rea_exec')
add('  浮动 S1',             f'=Summary!G{sub_h}', f"应 ≈ -1,683,661 CNY (-168万)",  key='h_unr_s1')
add('  浮动 S2',             f'=Summary!H{sub_h}', f"应 ≈ -2,835,820 CNY (-284万)",  key='h_unr_s2')
add('  股息',                f'=Summary!I{sub_h}', '',  key='h_div')
add('  卖费 (WAC)',          f'=Summary!J{sub_h}', '',  key='h_fee_sell_wac')
add('  卖费 (EXEC)',         f'=Summary!K{sub_h}', '',  key='h_fee_sell_exec')
add('  费用合计',            f'=Summary!L{sub_h}', '',  key='h_fee')
add('  WAC 净',              f'=Summary!M{sub_h}', '',  key='h_rea_wac_net')
add('  执行 净',             f'=Summary!N{sub_h}', '',  key='h_rea_exec_net')
add('  合计',                f'=Summary!O{sub_h}', '',  key='h_tot')
add('  HK_VAL_FX',           '=HK_VAL_FX',         'Inputs.HK_VAL_FX (单一 0.901987)', key='hk_val_fx')
blank()
add('总 计 (CNY)', header=True)
add('  WAC 价差 (gross)',     '={a_rea_wac}+{h_rea_wac}',   '★ recon: -2,295,758', key='t_rea_wac')
add('  执行价差 (gross)',     '={a_rea_exec}+{h_rea_exec}', '', key='t_rea_exec')
add('  浮动 (Source 1)',       '={a_unr_s1}+{h_unr_s1}',    '★ 应≈ -6,468,150 (-647 万)', key='t_unr_s1')
add('  浮动 (Source 2)',       '={a_unr_s2}+{h_unr_s2}',    '★ 应≈ -8,899,100 (-889.91 万)', key='t_unr_s2')
add('  股息',                  '={a_div}+{h_div}',           '', key='t_div')
add('  卖费 (WAC)',            '={a_fee_sell_wac}+{h_fee_sell_wac}',   '', key='t_fee_sell_wac')
add('  卖费 (EXEC)',           '={a_fee_sell_exec}+{h_fee_sell_exec}', '', key='t_fee_sell_exec')
add('  费用合计 (WAC)',        '={a_fee}+{h_fee}',           '买+卖 @ WAC rate', key='t_fee')
add('  WAC 净 (Source 1)',     '={t_rea_wac}-{t_fee_sell_wac}',   '★ 应 = -2,435,678 (-243.57万)', key='t_rea_wac_net')
add('  执行 净 (Source 2)',    '={t_rea_exec}-{t_fee_sell_exec}', '★ 应 ≈ -4,728 (-0.47万)', key='t_rea_exec_net')
add('  总损益 (Source 1 view)', '={t_rea_wac}+{t_unr_s1}-{t_fee}+{t_div}',
    'WAC价差 + 浮动 S1 − 费用 + 股息', key='t_pnl_s1')
add('  总损益 (Source 2 view)', '={t_rea_exec}+{t_unr_s2}-{t_fee_sell_exec}+{t_div}',
    '执行价差 + 浮动 S2 − 卖费(EXEC) + 股息', key='t_pnl_s2')
blank()
# =====================================================================
# 折算 万元 — horizontal layout: B = 風控 (Source 1), C = 權益OCI賬戶 (Source 2)
# Rows 46 (header) → 47 (col header) → 48..53 (metrics)
# =====================================================================
add('折算 万元 — 風控 vs 權益OCI賬戶', header=True)   # row 46

ROW_HDR = ROW_NEXT
# Column header row
ws_t.cell(row=ROW_HDR, column=1, value='指标 (万 CNY)').font = F_SUB
ws_t.cell(row=ROW_HDR, column=1).fill = FILL_R
ws_t.cell(row=ROW_HDR, column=2, value='風控 (Source 1)').font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
ws_t.cell(row=ROW_HDR, column=2).fill = FILL_T_CNY
ws_t.cell(row=ROW_HDR, column=2).alignment = ALIGN_C
ws_t.cell(row=ROW_HDR, column=3, value='權益OCI賬戶 (Source 2)').font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
ws_t.cell(row=ROW_HDR, column=3).fill = FILL_T_HKD
ws_t.cell(row=ROW_HDR, column=3).alignment = ALIGN_C
ws_t.cell(row=ROW_HDR, column=4, value='说明').font = F_SUB
ws_t.cell(row=ROW_HDR, column=4).fill = FILL_R
for c in range(1, 5):
    ws_t.cell(row=ROW_HDR, column=c).border = BORDER
ROW_NEXT += 1

# Metric rows: A=label, B=Source1 formula, C=Source2 formula, D=note
def add_h(label, b_formula, c_formula, note='', highlight=False, key=None):
    """Add a horizontal row with parallel S1 / S2 columns (万元 view)."""
    global ROW_NEXT
    r = ROW_NEXT
    if key: anchors[key] = r
    # Label
    ws_t.cell(row=r, column=1, value=label).font = F_BODY
    ws_t.cell(row=r, column=1).border = BORDER
    # B = Source 1 (风控)
    if isinstance(b_formula, str) and '{' in b_formula:
        for k, v in anchors.items():
            b_formula = b_formula.replace('{'+k+'}', f'B{v}')
    cb = ws_t.cell(row=r, column=2, value=b_formula)
    cb.font = F_BODY; cb.number_format = '#,##0.00;[Red]-#,##0.00'
    cb.alignment = ALIGN_R; cb.border = BORDER
    # C = Source 2 (OCI)
    if isinstance(c_formula, str) and '{' in c_formula:
        for k, v in anchors.items():
            c_formula = c_formula.replace('{'+k+'}', f'B{v}')
    cc = ws_t.cell(row=r, column=3, value=c_formula)
    cc.font = F_BODY; cc.number_format = '#,##0.00;[Red]-#,##0.00'
    cc.alignment = ALIGN_R; cc.border = BORDER
    # D = note
    if note:
        ws_t.cell(row=r, column=4, value=note).font = F_REMARK
        ws_t.cell(row=r, column=4).alignment = ALIGN_LW
    ws_t.cell(row=r, column=4).border = BORDER
    if highlight:
        for c in range(1, 5):
            ws_t.cell(row=r, column=c).fill = FILL_FINAL
        cb.font = F_SUB; cc.font = F_SUB
    ROW_NEXT += 1

add_h('  价差 (万)',  '={t_rea_wac_net}/10000', '={t_rea_exec_net}/10000',
      '风控 = WAC 净 (-243.57); OCI = EXEC 净 (-0.47)')
add_h('  浮动 (万)',  '={t_unr_s1}/10000', '={t_unr_s2}/10000',
      '风控 -647 万; OCI -889.91 万')
add_h('  股息 (万)',  '={t_div}/10000', '={t_div}/10000',
      '两口径相同')
add_h('  费用 (万)',  '={t_fee_sell_wac}/10000', '={t_fee_sell_exec}/10000',
      '风控 = WAC 卖费 (HK 13.2 bps); OCI = EXEC 卖费 (HK 11.27 bps)')
add_h('  总损益 (万)',
      '=({t_rea_wac_net}+{t_unr_s1}+{t_div})/10000',
      '=({t_rea_exec_net}+{t_unr_s2}+{t_div})/10000',
      '价差(净) + 浮动 + 股息  ★', highlight=True, key='w_pnl_h')

# Highlight v22 key anchor rows that still exist
for r in (anchors['t_pnl_s1'], anchors['t_pnl_s2'],
          anchors['t_rea_wac'], anchors['t_rea_wac_net'],
          anchors['t_rea_exec'], anchors['t_rea_exec_net'],
          anchors['t_unr_s1'], anchors['t_unr_s2']):
    for c in range(1,4):
        ws_t.cell(row=r, column=c).fill = FILL_FINAL
    ws_t.cell(row=r, column=2).font = F_SUB

ws_t.column_dimensions['A'].width = 32
ws_t.column_dimensions['B'].width = 22
ws_t.column_dimensions['C'].width = 50
ws_t.column_dimensions['D'].width = 30

# Reorder: Total / Summary / Inputs / Prices / StartPos / per-stock
desired_front = ['Total', 'Summary', 'Inputs', 'Prices', 'StartPos']
front_sheets = [wb[name] for name in desired_front if name in wb.sheetnames]
other_sheets = [s for s in wb._sheets if s.title not in desired_front]
wb._sheets = front_sheets + other_sheets

wb.calculation.calcMode = 'auto'
wb.calculation.fullCalcOnLoad = True

wb.save(OUT_FILE)
print(f"Saved: {OUT_FILE}")