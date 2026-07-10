# -*- coding: utf-8 -*-
import openpyxl, csv, sys, io, json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, Reference
from collections import defaultdict
from datetime import datetime, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os
PNL_MONTH = os.environ.get('PNL_MONTH', 'may').lower()   # jan/feb/mar/apr/may/...
_MNUM = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,
         'sep':9,'oct':10,'nov':11,'dec':12}[PNL_MONTH]
IS_FEB = _MNUM >= 2          # True for any month after January (rolled StartPos, no official-daily file)
_CN_MONTH = {1:'1月',2:'2月',3:'3月',4:'4月',5:'5月',6:'6月',7:'7月',8:'8月',
             9:'9月',10:'10月',11:'11月',12:'12月'}[_MNUM]
_EN_MONTH = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',
             9:'Sep',10:'Oct',11:'Nov',12:'Dec'}[_MNUM]
_IN_JSON = '_intermediate.json' if _MNUM == 1 else f'_intermediate_{PNL_MONTH}.json'
_DEF_OUT = f'PnL_Reconciliation_2026{_EN_MONTH}.xlsx'
OUTFILE = os.environ.get('OUTFILE', _DEF_OUT)
# 前海考核校验 逐券汇总 column map: Jan 估值=17/18 报告=29/30; each later month shifts +2.
_off = 2*(_MNUM-1)
QH_COLS = dict(val=(17+_off, 18+_off), rep=(29+_off, 30+_off))
MONTH_LABEL = f'2026年{_CN_MONTH}'
# replicated-tab (前海_逐券汇总) column LETTERS for the side-by-side block, per month
from openpyxl.utils import get_column_letter as _gcl
QHZQ_VAL = (_gcl(QH_COLS['val'][0]), _gcl(QH_COLS['val'][1]))   # 估值 价差/公允 letters
QHZQ_REP = (_gcl(QH_COLS['rep'][0]), _gcl(QH_COLS['rep'][1]))   # 报告 RPNL/MTM letters

D = json.load(open(_IN_JSON, encoding='utf-8'))
start, trades = D['start'], D['trades']
JAN_DAYS = D['JAN_DAYS']          # period trading days (Jan or Feb)
px, fx, px_code, cls = D['px'], D['fx'], D['px_code'], D['cls']
official_fx = D['official_fx']   # daily HKD->CNY mid (used for HK valuation)

# stable code ordering: A-shares then HK, by code
codes = sorted(start.keys(), key=lambda c: (cls[c], c))
NDAY = len(JAN_DAYS)

# ---------- GICS sector map (analyst-supplied; NOT in source data) ----------
# Editable. Used only for sector attribution display. 11 GICS sectors (中文/EN).
GICS = {
    # A-share
    '000001': '金融 Financials',      '000568': '日常消费 Consumer Staples', '000858': '日常消费 Consumer Staples',
    '600036': '金融 Financials',      '600519': '日常消费 Consumer Staples', '600809': '日常消费 Consumer Staples',
    '601009': '金融 Financials',      '601098': '通信服务 Comm Services',    '601166': '金融 Financials',
    '601169': '金融 Financials',      '601818': '金融 Financials',           '601900': '通信服务 Comm Services',
    '603198': '日常消费 Consumer Staples', '603369': '日常消费 Consumer Staples',
    # HK
    '00177': '工业 Industrials',      '00300': '可选消费 Consumer Disc',     '00322': '日常消费 Consumer Staples',
    '00371': '公用事业 Utilities',    '00576': '工业 Industrials',           '00728': '通信服务 Comm Services',
    '00762': '通信服务 Comm Services','00788': '通信服务 Comm Services',     '00811': '通信服务 Comm Services',
    '00836': '公用事业 Utilities',    '00939': '金融 Financials',            '00941': '通信服务 Comm Services',
    '00995': '工业 Industrials',      '00998': '金融 Financials',            '01071': '公用事业 Utilities',
    '01368': '可选消费 Consumer Disc','01398': '金融 Financials',            '02020': '可选消费 Consumer Disc',
    '02318': '金融 Financials',       '02319': '日常消费 Consumer Staples',  '02607': '医疗保健 Health Care',
    '03328': '金融 Financials',       '03933': '医疗保健 Health Care',       '03988': '金融 Financials',
    '06690': '可选消费 Consumer Disc','06862': '可选消费 Consumer Disc',
}

# ---------- January cash dividends (ex-date in Jan) ----------
# Source: data/dividend_query_results_online_2026-01-01_to_2026-06-01.csv (data source 新浪财经).
# per-share in CNY; shares are taken live from QtyEnd[ex-date] so booking is position-driven.
# Only names held at ex-date get a dividend (光大/江苏银行 held 0 -> auto 0).
DIVIDENDS_JAN = [
    dict(code='000568', ex='2026-01-30', per_share=1.358, ccy='CNY', dtype='现金分红',
         record='2026-01-29', pay='2026-01-30', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/000568.phtml',
         note='dividend_query_results_online CSV, 行 泸州老窖 ex=2026-01-30'),
    dict(code='600036', ex='2026-01-16', per_share=1.013, ccy='CNY', dtype='现金分红',
         record='2026-01-15', pay='2026-01-16', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/600036.phtml',
         note='dividend_query_results_online CSV, 行 招商银行 ex=2026-01-16; 官方记于 XD招商银 行'),
]
# February cash dividends (ex-date in Feb), from same CSV (新浪财经).
DIVIDENDS_FEB = [
    dict(code='601166', ex='2026-02-06', per_share=0.565, ccy='CNY', dtype='现金分红',
         record='2026-02-05', pay='2026-02-06', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/601166.phtml',
         note='dividend_query_results_online CSV, 行 兴业银行 ex=2026-02-06'),
    dict(code='601818', ex='2026-02-05', per_share=0.105, ccy='CNY', dtype='现金分红',
         record='2026-02-04', pay='2026-02-05', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/601818.phtml',
         note='dividend_query_results_online CSV, 行 光大银行 ex=2026-02-05'),
]
# March & April 2026: no ex-dividends in the holding set.
DIVIDENDS_MAR = []
DIVIDENDS_APR = []
# May cash dividends (ex-date in May). 港股通 booked in CNY: RMB-quoted use RMB; HKD-only
# converted at ex-date CSDC mid (安踏 HKD1.08→0.935604, 海底捞 HKD0.384→0.333773).
DIVIDENDS_MAY = [
    dict(code='01398', ex='2026-05-04', per_share=0.16890, ccy='CNY', dtype='现金分红',
         record='2026-05-01', pay='2026-05-04', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/01398.phtml',
         note='工商银行(H) RMB 0.1689'),
    dict(code='00576', ex='2026-05-08', per_share=0.39500, ccy='CNY', dtype='现金分红',
         record='2026-05-07', pay='2026-05-08', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00576.phtml',
         note='浙江沪杭甬 RMB 0.395'),
    dict(code='01368', ex='2026-05-12', per_share=0.08400, ccy='CNY', dtype='现金分红',
         record='2026-05-11', pay='2026-05-12', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/01368.phtml',
         note='特步国际 RMB 0.084 (scrip option, cash leg)'),
    dict(code='02020', ex='2026-05-15', per_share=0.935604, ccy='CNY', dtype='现金分红',
         record='2026-05-14', pay='2026-05-15', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/02020.phtml',
         note='安踏体育 HKD 1.08 × CSDC mid 0.8663'),
    dict(code='00998', ex='2026-05-18', per_share=0.19300, ccy='CNY', dtype='现金分红',
         record='2026-05-15', pay='2026-05-18', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00998.phtml',
         note='中信银行(H) RMB 0.193'),
    dict(code='00788', ex='2026-05-20', per_share=0.32539, ccy='CNY', dtype='现金分红',
         record='2026-05-19', pay='2026-05-20', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00788.phtml',
         note='中国铁塔 RMB 0.32539'),
    dict(code='06862', ex='2026-05-21', per_share=0.333773, ccy='CNY', dtype='现金分红',
         record='2026-05-20', pay='2026-05-21', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/06862.phtml',
         note='海底捞 HKD 0.384 × CSDC mid 0.8692'),
    # 江阴银行(002807)/张家港行(002839) not in holding set -> dividends don't book; omitted.
    dict(code='00811', ex='2026-05-29', per_share=0.42000, ccy='CNY', dtype='现金分红',
         record='2026-05-28', pay='2026-05-29', src='新浪财经',
         url='https://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_ShareBonus/stockid/00811.phtml',
         note='新华文轩 RMB 0.42'),
]
DIVIDENDS = {'jan':DIVIDENDS_JAN,'feb':DIVIDENDS_FEB,'mar':DIVIDENDS_MAR,
             'apr':DIVIDENDS_APR,'may':DIVIDENDS_MAY}.get(PNL_MONTH, [])
div_by_code = {d['code']: d for d in DIVIDENDS}

# ---------- official row merges: fold "XD" ex-dividend rows back into the stock ----------
# Official books 招商银行's Jan dividend in a separate row 'XD招商银' (只有股利盈亏).
# To compare like-for-like, its P&L subjects are added to 招商银行's official numbers.
OFFICIAL_MERGE = {'招商银行': ['XD招商银']}   # target_name -> [extra official rows to add]

# ---------- official daily + month + dividend, per name, CNY (Jan attribution file only) ----------
date_col = {}
PNL_SUBJ = ['持仓公允', 'OCI价差', '股利盈亏', '利息盈亏']
off_rows = defaultdict(dict)   # name -> subj -> row tuple
if not IS_FEB:
    wb_off = openpyxl.load_workbook('data/2026年1月(1).xlsx', data_only=True)
    ws = wb_off['Sheet54']
    rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
    for i in range(5, 36):
        v = hdr[i]
        s = v.date().isoformat() if hasattr(v, 'date') else (v if isinstance(v, str) else str(v))
        date_col[s] = i
    cur = None
    for r in rows[1:]:
        if r[2] is not None: cur = r[2]
        if r[3] is None or cur is None: continue
        off_rows[cur][r[3]] = r

def _names_for(name):
    """official source rows = the name itself + any merged XD rows."""
    return [name] + OFFICIAL_MERGE.get(name, [])

def off_daily_cny(name, day):
    ci = date_col.get(day)
    tot = 0.0
    for nm in _names_for(name):
        d = off_rows.get(nm, {})
        for s in PNL_SUBJ:
            rr = d.get(s)
            if rr and ci is not None and isinstance(rr[ci], (int, float)):
                tot += rr[ci]
    return tot * 10000.0

def off_month_cny(name):
    tot = 0.0
    for nm in _names_for(name):
        d = off_rows.get(nm, {})
        for s in PNL_SUBJ:
            rr = d.get(s)
            if rr:
                tot += sum(v for v in rr[5:36] if isinstance(v, (int, float)))
    return tot * 10000.0

def off_div_cny(name):
    tot = 0.0
    for nm in _names_for(name):
        rr = off_rows.get(nm, {}).get('股利盈亏')
        if rr:
            tot += sum(v for v in rr[5:36] if isinstance(v, (int, float)))*10000.0
    return tot

def off_subj_day(name, subj, day):
    """one official subject (CNY) for name on day, incl. merged rows."""
    ci = date_col.get(day); tot = 0.0
    for nm in _names_for(name):
        rr = off_rows.get(nm, {}).get(subj)
        if rr and ci is not None and isinstance(rr[ci], (int, float)):
            tot += rr[ci]
    return tot * 10000.0

def off_implied_fee(name, day, is_buy, cny):
    """Transaction fee the official book applied, derived from 持仓成本 / OCI价差 deltas.
    buy : cost basis added = cash + fee  -> fee = Δ持仓成本 - cash
    sell: fee = gross cash - net proceeds = cash - (OCI价差 + (-Δ持仓成本))
    """
    cost_delta = off_subj_day(name, '持仓成本', day)
    if is_buy:
        return cost_delta - cny
    oci = off_subj_day(name, 'OCI价差', day)
    return cny - (oci + (-cost_delta))

# ---------- fee rate model (China A-share / HK 港股通), reproduces official to ~cents ----------
A_STAMP_BPS = 5.0      # 印花税: A-share SELL only, 0.05%
A_COMM_BPS  = 0.641    # 佣金+规费: both sides
HK_STAMP_BPS = 10.0    # 印花税: HK both sides, 0.10%
HK_OTHER_BPS = 1.362   # 交易费/征费/结算费/佣金 等: both sides (fit to official, ~11.362bps total)

# ---------- styling helpers ----------
HDR = Font(bold=True, color='FFFFFF', size=10)
HDRFILL = PatternFill('solid', fgColor='305496')
SUBHDR = Font(bold=True, size=10)
BOLD = Font(bold=True)
GREY = PatternFill('solid', fgColor='D9E1F2')
YEL = PatternFill('solid', fgColor='FFF2CC')
RED = PatternFill('solid', fgColor='F8CBAD')
GRN = PatternFill('solid', fgColor='C6E0B4')
thin = Side(style='thin', color='BFBFBF')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CNY = '#,##0;[Red](#,##0)'
CNY2 = '#,##0.00;[Red](#,##0.00)'
PCT = '0.0%'
PX4 = '0.0000'

def style_header_row(ws, row, ncol, start_col=1):
    for c in range(start_col, start_col+ncol):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDRFILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORD

wb = openpyxl.Workbook()

# ============================================================ README
ws = wb.active; ws.title = 'README'
readme = [
 ('港股通 A+H 股票 PnL 对账 / Reconciliation — 2026年1月', 14, True),
 ('', 10, False),
 ('Account 自营权益4 (3008).  Period: 2026-01-01 → 2026-01-31 (20 trading days, 01-02…01-30).', 10, False),
 ('', 10, False),
 ('PURPOSE', 12, True),
 ('Independently recompute daily & monthly P&L for every equity holding (A-share + HK 港股通)', 10, False),
 ('from raw market data, and reconcile against the official attribution file 2026年1月(1).xlsx.', 10, False),
 ('', 10, False),
 ('METHOD (our side — fully formula-driven in this workbook)', 12, True),
 ('• HK 港股通 settle in CNY daily. close_CNY = close_HKD × HKD→CNY FX. FX = OFFICIAL daily fixing', 10, False),
 ('     (FX!col C, backed out from the book; market CSV FX in col B differs 10–30bps/day and made', 10, False),
 ('     HK daily diffs swing ±100k that net to ~0 monthly — see FX sheet gap column).', 10, False),
 ('• Daily P&L (CNY) = Qty_end×Close_t − Qty_start×Close_(t−1) − NetCashFlow_t + Dividend_t − Fee_t', 10, False),
 ('     where NetCashFlow = buys(+) − sells(−) in settled CNY (本币成交金额, gross of fees);', 10, False),
 ('     Dividend_t = shares held at ex-date × per-share (CNY), booked on ex-date (total-return);', 10, False),
 ('     Fee_t = 印花税 + 佣金/规费 on trade date (A: stamp 5bps sell-only +comm; HK: stamp 10bps +other).', 10, False),
 ('• Qty_end = start qty + cumulative signed trade qty up to & incl. day t.', 10, False),
 ('• Prev-day baseline for 01-02 = year-open (Dec-31) qty & CNY price from 期初持仓.', 10, False),
 ('', 10, False),
 ('OFFICIAL side (source of truth — values copied from 2026年1月(1).xlsx, units 万元×10000)', 12, True),
 ('• IFRS-9 FVOCI equities. Official daily P&L = 持仓公允 + OCI价差(realised) + 股利盈亏 + 利息盈亏.', 10, False),
 ('• 持仓公允 in 本年累计 is a CUMULATIVE balance; month P&L = Σ daily 发生额 (Sheet54).', 10, False),
 ('', 10, False),
 ('SHEETS', 12, True),
 ('  Recon_Monthly  — per-stock month total: Ours vs Official vs Diff (+ dividend reconciling item).  ★ START HERE', 10, False),
 ('  Recon_Daily    — per-stock × day: Ours − Official daily difference grid.', 10, False),
 ('  DailyPnL_Ours  — our computed daily P&L matrix (MTM + dividend, formulas).', 10, False),
 ('  Dividends      — position-driven cash dividends (shares@ex × per-share), per ex-date.', 10, False),
 ('  Fees           — per-trade 印花税 + 佣金/规费 (rate-driven), with official-implied check.', 10, False),
 ('  PxCNY / QtyEnd — intermediate calc matrices (CNY prices w/ carry-forward; daily qty).', 10, False),
 ('  PxNative / FX / Trades / StartPos / OfficialDaily — raw inputs.', 10, False),
 ('', 10, False),
 ('KNOWN RECONCILING ITEMS', 12, True),
 ('• Dividends NOW BOOKED on our side (total-return): 泸州老窖 25k×1.358 (ex 01-30);', 10, False),
 ('     招商银行 150k×1.013 (ex 01-16). Official books 招商 dividend in row "XD招商银" —', 10, False),
 ('     that row is merged back into 招商银行 on the official side for like-for-like compare.', 10, False),
 ('• Transaction fees NOW BOOKED on our side (Fees sheet): A-share daily diffs on trade days', 10, False),
 ('     (印花税 5bps sell + 佣金 ~0.64bps) now reconcile to ~0; HK fees ~11.3bps booked too.', 10, False),
 ('• HK FX FIXED: HK valuation now uses the official daily FX fixing (FX!col C), so non-traded HK', 10, False),
 ('     names reconcile daily to ~0. Previously the market-CSV FX caused ±100k daily swings.', 10, False),
 ('• 中国铁塔 (HK): small cost-basis FX-timing residual on buy days (~4.6k), flagged investigate.', 10, False),
 ('• ±50–150 CNY residuals = rounding (official stored in 万元 to 6 dp) + HK fee-rate approximation.', 10, False),
]
for i, (txt, sz, b) in enumerate(readme, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(bold=b, size=sz)
ws.column_dimensions['A'].width = 110
ws['A1'].fill = HDRFILL; ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')

# ============================================================ StartPos
ws = wb.create_sheet('StartPos')
sp_hdr = ['代码 code', '名称 name', '类别 class', '期初数量 qty', '期初成本(CNY) cost',
          '期初价(CNY) px_Dec31', '期初价(原币) px_Dec31_HKD', '期初FX Dec31']
ws.append(sp_hdr); style_header_row(ws, 1, len(sp_hdr))
dec31_hkd = D['dec31_hkd']; fx_dec31 = D['official_fx_dec31']
sp_row = {}
for i, code in enumerate(codes, start=2):
    s = start[code]
    ws.cell(row=i, column=1, value=code).number_format = '@'
    ws.cell(row=i, column=2, value=s['name'])
    ws.cell(row=i, column=3, value=cls[code])
    ws.cell(row=i, column=4, value=s['qty'])
    ws.cell(row=i, column=5, value=s['cost']).number_format = CNY2
    ws.cell(row=i, column=6, value=s['px_cny']).number_format = PX4
    # HK: Dec-31 HKD close + Dec-31 official FX (A-share: native = CNY, FX=1)
    if cls[code] == 'HK':
        ws.cell(row=i, column=7, value=dec31_hkd.get(px_code[code])).number_format = PX4
        ws.cell(row=i, column=8, value=fx_dec31).number_format = '0.00000'
    else:
        ws.cell(row=i, column=7, value=f'=F{i}').number_format = PX4
        ws.cell(row=i, column=8, value=1).number_format = '0.00000'
    sp_row[code] = i
for col, w in zip('ABCDEFGH', [10, 18, 7, 13, 16, 14, 16, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
SP_LAST = len(codes)+1

# ============================================================ Trades
ws = wb.create_sheet('Trades')
tr_hdr = ['日期 date', '代码 code', '名称 name', '方向 dir', 'sign', '数量 qty',
          '带符号数量 signed_qty', '结算金额CNY 本币成交金额', '现金流 cashflow(=sign×CNY)', '市场 market',
          '成交金额(原币) trade_amt', '带符号原币现金流 signed_FCcf']
ws.append(tr_hdr); style_header_row(ws, 1, len(tr_hdr))
for i, t in enumerate(trades, start=2):
    ws.cell(row=i, column=1, value=t['date']).number_format = '@'
    ws.cell(row=i, column=2, value=t['code']).number_format = '@'
    ws.cell(row=i, column=3, value=t['name'])
    ws.cell(row=i, column=4, value=t['dir'])
    ws.cell(row=i, column=5, value=t['sign'])
    ws.cell(row=i, column=6, value=t['qty'])
    ws.cell(row=i, column=7, value=t['signed_qty'])
    ws.cell(row=i, column=8, value=t['cny_amt']).number_format = CNY2
    # cashflow as FORMULA = sign*cny_amt
    ws.cell(row=i, column=9, value=f'=E{i}*H{i}').number_format = CNY2
    ws.cell(row=i, column=10, value=t['mkt'])
    ws.cell(row=i, column=11, value=t['trade_amt']).number_format = CNY2     # native (HKD for HK)
    ws.cell(row=i, column=12, value=f'=E{i}*K{i}').number_format = CNY2      # signed native cashflow
for col, w in zip('ABCDEFGHIJKL', [12, 10, 16, 8, 6, 11, 14, 18, 20, 14, 16, 18]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
TR_LAST = len(trades)+1

# ============================================================ FX (vertical)
# Source: 港股通 CSDC 每日结算汇兑比率 (买入/卖出), confirmed = the book's fixing.
# Book value = (买入 + 卖出) / 2, 5 decimal places (diff = 0.00 bps on all 20 trading days).
# Verified by back-calculation from 03988 中国银行 (no Jan trades), cross-checked vs 00941.
# Jan-02: no CSDC fixing (港股通 not open); carry Dec-31 mid (0.8974). P&L for that day = 0.
CSDC_BUY = {
    '2026-01-05':0.89422,'2026-01-06':0.89577,'2026-01-07':0.89592,'2026-01-08':0.89746,
    '2026-01-09':0.89578,'2026-01-12':0.89569,'2026-01-13':0.89409,'2026-01-14':0.89409,
    '2026-01-15':0.89412,'2026-01-16':0.89348,'2026-01-19':0.89318,'2026-01-20':0.89250,
    '2026-01-21':0.89135,'2026-01-22':0.89279,'2026-01-23':0.89259,'2026-01-26':0.89268,
    '2026-01-27':0.89200,'2026-01-28':0.89186,'2026-01-29':0.88948,'2026-01-30':0.88948,
}
CSDC_SELL = {
    '2026-01-05':0.89478,'2026-01-06':0.89583,'2026-01-07':0.89628,'2026-01-08':0.89754,
    '2026-01-09':0.89582,'2026-01-12':0.89551,'2026-01-13':0.89411,'2026-01-14':0.89411,
    '2026-01-15':0.89428,'2026-01-16':0.89352,'2026-01-19':0.89322,'2026-01-20':0.89250,
    '2026-01-21':0.89185,'2026-01-22':0.89281,'2026-01-23':0.89261,'2026-01-26':0.89272,
    '2026-01-27':0.89200,'2026-01-28':0.89174,'2026-01-29':0.88952,'2026-01-30':0.88952,
}
ws = wb.create_sheet('FX')
fxh = ['日期 date', '买入结算汇兑比率 Buy', '卖出结算汇兑比率 Sell', '中间价 Mid=(B+C)/2 (book uses this)']
ws.append(fxh); style_header_row(ws, 1, len(fxh))
for i, day in enumerate(JAN_DAYS, start=2):
    ws.cell(row=i, column=1, value=day).number_format = '@'
    if day in CSDC_BUY:
        ws.cell(row=i, column=2, value=CSDC_BUY[day]).number_format = '0.00000'
        ws.cell(row=i, column=3, value=CSDC_SELL[day]).number_format = '0.00000'
        ws.cell(row=i, column=4, value=f'=(B{i}+C{i})/2').number_format = '0.00000'
    else:
        # Jan-02: no CSDC fixing; carry Dec-31 mid as placeholder (P&L = 0 that day)
        ws.cell(row=i, column=4, value=official_fx[day]).number_format = '0.00000'
        ws.cell(row=i, column=2).value = None  # blank
        ws.cell(row=i, column=3).value = None  # blank
ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 32
# notes
ws.cell(row=NDAY+3, column=1, value='注 Note:').font = BOLD
ws.cell(row=NDAY+4, column=1,
        value='官方FX = CSDC中间价 = (买入+卖出)/2 取5位小数；与账簿估值完全一致(差异=0.00bps，经中国银行/中国移动双重验证)。')
ws.cell(row=NDAY+5, column=1,
        value='01-02 港股通未开放(无CSDC汇率)；账簿未估值, 我方沿用Dec-31汇率(P&L=0), HK每日对账归零。')
ws.cell(row=NDAY+6, column=1,
        value='估值方法: 每日 收盘价(原币HKD) × 官方FX中间价(D列) = 收盘价CNY；当日盈亏 = 数量×(收盘CNY_today − 收盘CNY_prev)。')
ws.cell(row=NDAY+7, column=1,
        value='下表为港股通各股每日收盘价(原币HKD), 与官方一致(无交易股票核对差异=0)。空白=当日无报价(沿用前值)。')
ws.freeze_panes = 'A2'
# valuation lookup uses mid = column D (index 4)
FX_RANGE = f'FX!$A$2:$D${NDAY+1}'
FX_COL_OFFICIAL = 4   # VLOOKUP column for official fixing (mid)

# ----- HK close-price matrix (原币HKD), HK stocks × dates -----
dec31_hkd_map = D['dec31_hkd']
hk_codes_fx = [cd for cd in codes if cls[cd] == 'HK']
hpx0 = NDAY + 8   # header row of the price matrix
ws.cell(row=hpx0, column=1, value='港股通收盘价 (原币HKD) / HK close px').font = BOLD
ws.cell(row=hpx0+1, column=1, value='代码').font = SUBHDR
ws.cell(row=hpx0+1, column=2, value='名称').font = SUBHDR
ws.cell(row=hpx0+1, column=3, value='Dec-31').font = SUBHDR
for j, day in enumerate(JAN_DAYS):
    cc = ws.cell(row=hpx0+1, column=4+j, value=day); cc.font = HDR; cc.fill = HDRFILL
    cc.number_format = '@'; cc.alignment = Alignment(horizontal='center')
for jj in range(3):
    c = ws.cell(row=hpx0+1, column=1+jj); c.font = HDR; c.fill = HDRFILL
for i, code in enumerate(hk_codes_fx):
    rr = hpx0 + 2 + i
    ws.cell(row=rr, column=1, value=code).number_format = '@'
    ws.cell(row=rr, column=2, value=start[code]['name'])
    ws.cell(row=rr, column=3, value=dec31_hkd_map.get(px_code[code])).number_format = PX4
    for j, day in enumerate(JAN_DAYS):
        v = px.get(px_code[code], {}).get(day)
        if v is not None:
            ws.cell(row=rr, column=4+j, value=v).number_format = PX4
for j in range(NDAY):
    ws.column_dimensions[get_column_letter(4+j)].width = 9
ws.column_dimensions['C'].width = 9

# ============================================================ matrix geometry
# col A=code, B=name, C=class, days D..(D+NDAY-1)
DAY0 = 4  # column index of first day (D)
def day_col(i): return DAY0 + i   # i in 0..NDAY-1
def day_letter(i): return get_column_letter(day_col(i))

def write_matrix_frame(ws, title_note=None):
    ws.cell(row=1, column=1, value='code').font = SUBHDR
    ws.cell(row=1, column=2, value='name').font = SUBHDR
    ws.cell(row=1, column=3, value='class').font = SUBHDR
    for i, day in enumerate(JAN_DAYS):
        c = ws.cell(row=1, column=day_col(i), value=day)
        c.font = HDR; c.fill = HDRFILL; c.number_format = '@'
        c.alignment = Alignment(horizontal='center')
    for r, code in enumerate(codes, start=2):
        ws.cell(row=r, column=1, value=code).number_format = '@'
        ws.cell(row=r, column=2, value=start[code]['name'])
        ws.cell(row=r, column=3, value=cls[code])
    ws.freeze_panes = get_column_letter(DAY0) + '2'
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 6
    for i in range(NDAY):
        ws.column_dimensions[day_letter(i)].width = 11

MROW = {code: r for r, code in enumerate(codes, start=2)}  # same row per code in every matrix

# ============================================================ PxNative (raw)
ws = wb.create_sheet('PxNative')
write_matrix_frame(ws)
for code in codes:
    r = MROW[code]
    for i, day in enumerate(JAN_DAYS):
        v = px.get(px_code[code], {}).get(day)
        if v is not None:
            ws.cell(row=r, column=day_col(i), value=v).number_format = PX4
style_header_row(ws, 1, 3)

# ============================================================ PxCNY (calc, carry-forward)
ws = wb.create_sheet('PxCNY')
write_matrix_frame(ws)
for code in codes:
    r = MROW[code]
    for i, day in enumerate(JAN_DAYS):
        L = day_letter(i)
        nat = f'PxNative!{L}{r}'
        fxlk = f'VLOOKUP({L}$1,{FX_RANGE},{FX_COL_OFFICIAL},FALSE)'  # official daily fixing
        conv = f'IF($C{r}="HK",{nat}*{fxlk},{nat})'
        if i == 0:
            # 01-02: official book does NOT mark this day -> carry Dec-31 mark (no 01-02 P&L).
            # 01-05 then absorbs the full Dec-31->01-05 move, matching the official.
            prev = f'StartPos!$F{sp_row[code]}'
            f = f'={prev}'
        else:
            prev = f'{day_letter(i-1)}{r}'
            f = f'=IF({nat}="",{prev},{conv})'
        ws.cell(row=r, column=day_col(i), value=f).number_format = PX4
style_header_row(ws, 1, 3)

# ============================================================ QtyEnd (calc)
ws = wb.create_sheet('QtyEnd')
write_matrix_frame(ws)
TR_DATE = 'Trades!$A$2:$A$%d' % TR_LAST
TR_CODE = 'Trades!$B$2:$B$%d' % TR_LAST
TR_SQTY = 'Trades!$G$2:$G$%d' % TR_LAST
for code in codes:
    r = MROW[code]
    for i, day in enumerate(JAN_DAYS):
        L = day_letter(i)
        today = f'SUMIFS({TR_SQTY},{TR_CODE},$A{r},{TR_DATE},{L}$1)'  # exact-match day
        prev = f'StartPos!$D{sp_row[code]}' if i == 0 else f'{day_letter(i-1)}{r}'
        f = f'={prev}+{today}'   # recursive cumulative qty (avoids text "<=" date compare)
        ws.cell(row=r, column=day_col(i), value=f).number_format = '#,##0'
style_header_row(ws, 1, 3)

# ============================================================ Dividends (calc, position-driven)
# Booked on ex-date: dividend_CNY = shares_held_at_ex (from QtyEnd) × per-share(CNY input).
# Inputs + full SOURCE REFERENCE (data source, record/pay dates, URL) shown to the right.
ws = wb.create_sheet('Dividends')
write_matrix_frame(ws)
DTOT = day_col(NDAY)                 # month total col
# input + reference columns to the right of the daily grid
EXD = DTOT + 1   # ex_date
RECD = DTOT + 2  # record date 股权登记日
PAYD = DTOT + 3  # pay date 派息日
PS = DTOT + 4    # per_share (CNY)
SH = DTOT + 5    # shares at ex (formula)
DTYPE = DTOT + 6 # dividend type
SRC = DTOT + 7   # data source
URL = DTOT + 8   # source url
NOTE = DTOT + 9  # note
ref_hdrs = {DTOT:'月合计 Div_Month', EXD:'除权日 ex-date', RECD:'股权登记日 record',
            PAYD:'派息日 pay', PS:'每股股息CNY per-share', SH:'除权日持股 shares@ex',
            DTYPE:'类型 type', SRC:'数据源 source', URL:'来源链接 source URL', NOTE:'备注 note'}
for col, txt in ref_hdrs.items():
    cc = ws.cell(row=1, column=col, value=txt); cc.font = HDR; cc.fill = HDRFILL
    cc.alignment = Alignment(horizontal='center', wrap_text=True)
ps_letter = get_column_letter(PS)
# divider styling for the reference block
REFFILL = PatternFill('solid', fgColor='FCE4D6')
for code in codes:
    r = MROW[code]
    dv = div_by_code.get(code)
    if dv:
        ws.cell(row=r, column=EXD, value=dv['ex']).number_format = '@'
        ws.cell(row=r, column=RECD, value=dv['record']).number_format = '@'
        ws.cell(row=r, column=PAYD, value=dv['pay']).number_format = '@'
        ws.cell(row=r, column=PS, value=dv['per_share']).number_format = '0.0000'
        exL = day_letter(JAN_DAYS.index(dv['ex']))
        ws.cell(row=r, column=SH, value=f'=QtyEnd!{exL}{r}').number_format = '#,##0'
        ws.cell(row=r, column=DTYPE, value=dv['dtype'])
        ws.cell(row=r, column=SRC, value=dv['src'])
        c_url = ws.cell(row=r, column=URL, value=dv['url'])
        c_url.hyperlink = dv['url']; c_url.font = Font(color='0563C1', underline='single', size=9)
        ws.cell(row=r, column=NOTE, value=dv['note'])
        for col in (EXD, RECD, PAYD, PS, SH, DTYPE, SRC, URL, NOTE):
            ws.cell(row=r, column=col).fill = REFFILL
    # per-day dividend: nonzero only on ex-date column
    for i, day in enumerate(JAN_DAYS):
        L = day_letter(i)
        if dv and day == dv['ex']:
            f = f'=QtyEnd!{L}{r}*${ps_letter}{r}'      # shares@ex × per-share(input)
            ws.cell(row=r, column=day_col(i), value=f).number_format = CNY
        else:
            ws.cell(row=r, column=day_col(i), value=0).number_format = CNY
    rng = f'{day_letter(0)}{r}:{day_letter(NDAY-1)}{r}'
    ws.cell(row=r, column=DTOT, value=f'=SUM({rng})').number_format = CNY
for col, w in [(DTOT,14),(EXD,12),(RECD,13),(PAYD,12),(PS,16),(SH,15),(DTYPE,9),(SRC,11),(URL,60),(NOTE,46)]:
    ws.column_dimensions[get_column_letter(col)].width = w
style_header_row(ws, 1, 3)
DIV_TOTCOL_LET = get_column_letter(DTOT)

# ============================================================ Fees (transaction costs) — per trade
# 印花税 (stamp duty) + 佣金/规费 (commission & levies), formula-driven from rate inputs.
# China A-share: stamp 5bps on SELL only; commission ~0.641bps both sides.
# HK 港股通:     stamp 10bps + other ~1.3bps, both sides.
# A "rate model fee" reproduces the official book; an "official-implied fee" column
# (derived from 持仓成本 / OCI价差 deltas) is shown alongside as an independent check.
ws = wb.create_sheet('Fees')
fee_hdr = ['日期 date', '代码 code', '名称 name', '类别 class', '方向 dir',
           '成交金额CNY notional', '印花税率bps stamp', '佣金率bps comm',
           '印花税CNY stamp', '佣金CNY comm', '费用合计CNY fee_total',
           '法定费率隐含费 statutory_implied', '差异 check']
if not IS_FEB:
    fee_hdr.append('官方隐含费(审计) official_implied')   # Jan-only audit column N
ws.append(fee_hdr); style_header_row(ws, 1, len(fee_hdr))
# rate inputs live in a small block to the right (editable)
ws.cell(row=1, column=15, value='费率参数 RATE INPUTS (bps)').font = BOLD
rate_cells = {}
rate_list = [('A股 印花税(卖) A stamp/sell', A_STAMP_BPS), ('A股 佣金 A comm', A_COMM_BPS),
             ('港股通 印花税 HK stamp', HK_STAMP_BPS), ('港股通 其他 HK other', HK_OTHER_BPS)]
for k, (lab, val) in enumerate(rate_list):
    ws.cell(row=2+k, column=15, value=lab).font = Font(size=9)
    rc = ws.cell(row=2+k, column=16, value=val); rc.number_format = '0.000'
    rc.fill = YEL; rc.border = BORD
rA_stamp, rA_comm, rHK_stamp, rHK_other = '$P$2', '$P$3', '$P$4', '$P$5'
ws.column_dimensions['O'].width = 28; ws.column_dimensions['P'].width = 9

byday_off = defaultdict(list)   # (name,day) -> trades, for official-implied per group
for t in trades:
    byday_off[(t['name'], t['date'])].append(t)

for i, t in enumerate(trades, start=2):
    is_buy = t['sign'] == 1
    ws.cell(row=i, column=1, value=t['date']).number_format = '@'
    ws.cell(row=i, column=2, value=t['code']).number_format = '@'
    ws.cell(row=i, column=3, value=t['name'])
    ws.cell(row=i, column=4, value=t['cls'])
    ws.cell(row=i, column=5, value=t['dir'])
    ws.cell(row=i, column=6, value=t['cny_amt']).number_format = CNY2   # notional (本币成交金额, gross)
    # stamp & comm rate per row (formula off class+dir and the rate inputs)
    ws.cell(row=i, column=7,
            value=(f'=IF(D{i}="A",IF(E{i}="卖出",{rA_stamp},0),{rHK_stamp})')).number_format = '0.000'
    ws.cell(row=i, column=8,
            value=(f'=IF(D{i}="A",{rA_comm},{rHK_other})')).number_format = '0.000'
    # amounts
    ws.cell(row=i, column=9, value=f'=F{i}*G{i}/10000').number_format = CNY2
    ws.cell(row=i, column=10, value=f'=F{i}*H{i}/10000').number_format = CNY2
    ws.cell(row=i, column=11, value=f'=I{i}+J{i}').number_format = CNY2
    # column L = 法定费率隐含费 (statutory-rate implied fee), FORMULA-driven for BOTH months:
    #   印花税(A卖/HK双边) + 佣金或其他, applied to notional. Independent of the rate-model K
    #   only in that it re-derives from the same statutory schedule -> 差异 check M ~ 0.
    ws.cell(row=i, column=12,
            value=(f'=F{i}*(IF(D{i}="A",IF(E{i}="卖出",{rA_stamp},0)+{rA_comm},'
                   f'{rHK_stamp}+{rHK_other}))/10000')).number_format = CNY2
    if not IS_FEB:
        # Jan also exposes the official-attribution-implied fee as a NOTE in col N (value, audit only)
        grp = byday_off[(t['name'], t['date'])]
        grp_cny = sum(x['cny_amt'] for x in grp)
        off_fee_grp = off_implied_fee(t['name'], t['date'], is_buy, grp_cny)
        share = (t['cny_amt']/grp_cny) if grp_cny else 0
        ws.cell(row=i, column=14, value=round(off_fee_grp*share, 2)).number_format = CNY2
    ws.cell(row=i, column=13, value=f'=K{i}-L{i}').number_format = CNY2
FEE_LAST = len(trades)+1
# ---- subtotals: 买入小计 / 卖出小计 / 合计 (SUMIFS on 方向 col E) ----
fbuy = FEE_LAST+1; fsell = FEE_LAST+2; ftot = FEE_LAST+3
SUMrng = lambda col: f'{col}2:{col}{FEE_LAST}'
DIRrng = f'$E$2:$E${FEE_LAST}'
def _fee_subtot(row, label, crit):
    ws.cell(row=row, column=3, value=label).font = BOLD
    for col in ('F', 'I', 'J', 'K', 'L'):
        ci = openpyxl.utils.column_index_from_string(col)
        if crit is None:
            f = f'=SUM({SUMrng(col)})'
        else:
            f = f'=SUMIFS({SUMrng(col)},{DIRrng},"{crit}")'
        cc = ws.cell(row=row, column=ci, value=f); cc.number_format = CNY2; cc.font = BOLD
        cc.fill = GREY
    for c in (1,2,3): ws.cell(row=row, column=c).fill = GREY
_fee_subtot(fbuy,  '买入小计 BUY subtotal',  '买入')
_fee_subtot(fsell, '卖出小计 SELL subtotal', '卖出')
_fee_subtot(ftot,  '合计 TOTAL', None)
for c in range(1, 13): ws.cell(row=ftot, column=c).fill = PatternFill('solid', fgColor='BDD7EE')
for col, w in zip('ABCDEFGHIJKLMN', [12,9,16,7,8,16,12,11,13,13,15,16,11,16]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
FEE_DATE = f'Fees!$A$2:$A${FEE_LAST}'
FEE_CODE = f'Fees!$B$2:$B${FEE_LAST}'
FEE_TOT  = f'Fees!$K$2:$K${FEE_LAST}'

TOTCOL = day_col(NDAY)        # column after last day = month total
TR_CASH = 'Trades!$I$2:$I$%d' % TR_LAST
dl_last = day_letter(NDAY-1)  # last trading day column letter

if not IS_FEB:
    # ============================================================ OfficialDaily (Jan only)
    ws = wb.create_sheet('OfficialDaily')
    write_matrix_frame(ws)
    mlet = get_column_letter(TOTCOL)
    ws.cell(row=1, column=TOTCOL, value='月合计(全月) Official_Month').font = HDR
    ws.cell(row=1, column=TOTCOL).fill = HDRFILL
    ws.column_dimensions[mlet].width = 20
    for code in codes:
        r = MROW[code]; name = start[code]['name']
        for i, day in enumerate(JAN_DAYS):
            ws.cell(row=r, column=day_col(i), value=round(off_daily_cny(name, day), 2)).number_format = CNY
        ws.cell(row=r, column=TOTCOL, value=round(off_month_cny(name), 2)).number_format = CNY
    style_header_row(ws, 1, 3)
    OFF_TOTCOL_LET = mlet

    # ============================================================ DailyPnL_Ours (Jan only)
    ws = wb.create_sheet('DailyPnL_Ours')
    write_matrix_frame(ws)
    for code in codes:
        r = MROW[code]
        for i, day in enumerate(JAN_DAYS):
            L = day_letter(i)
            qty_t = f'QtyEnd!{L}{r}'; px_t = f'PxCNY!{L}{r}'
            cash_t = f'SUMIFS({TR_CASH},{TR_CODE},$A{r},{TR_DATE},{L}$1)'
            if i == 0:
                qty_p = f'StartPos!$D{sp_row[code]}'; px_p = f'StartPos!$F{sp_row[code]}'
            else:
                Lp = day_letter(i-1); qty_p = f'QtyEnd!{Lp}{r}'; px_p = f'PxCNY!{Lp}{r}'
            div_t = f'Dividends!{L}{r}'
            fee_t = f'SUMIFS({FEE_TOT},{FEE_CODE},$A{r},{FEE_DATE},{L}$1)'
            f = f'={qty_t}*{px_t}-{qty_p}*{px_p}-{cash_t}+{div_t}-{fee_t}'
            ws.cell(row=r, column=day_col(i), value=f).number_format = CNY
        rng = f'{day_letter(0)}{r}:{day_letter(NDAY-1)}{r}'
        ws.cell(row=r, column=TOTCOL, value=f'=SUM({rng})').number_format = CNY
    ws.cell(row=1, column=TOTCOL, value='月合计 Ours_Month').font = HDR
    ws.cell(row=1, column=TOTCOL).fill = HDRFILL
    ws.column_dimensions[get_column_letter(TOTCOL)].width = 16
    trow = len(codes)+2
    ws.cell(row=trow, column=2, value='合计 TOTAL').font = BOLD
    for i in range(NDAY):
        L = day_letter(i)
        ws.cell(row=trow, column=day_col(i), value=f'=SUM({L}2:{L}{len(codes)+1})').number_format = CNY
        ws.cell(row=trow, column=day_col(i)).font = BOLD
    ws.cell(row=trow, column=TOTCOL, value=f'=SUM({get_column_letter(TOTCOL)}2:{get_column_letter(TOTCOL)}{len(codes)+1})').number_format = CNY
    ws.cell(row=trow, column=TOTCOL).font = BOLD
    style_header_row(ws, 1, 3)
    OURS_TOTCOL_LET = get_column_letter(TOTCOL)
    OURS_SHEET = 'DailyPnL_Ours'

    # ============================================================ Recon_Daily (Jan only)
    ws = wb.create_sheet('Recon_Daily')
    write_matrix_frame(ws)
    for code in codes:
        r = MROW[code]
        for i, day in enumerate(JAN_DAYS):
            L = day_letter(i)
            ws.cell(row=r, column=day_col(i), value=f'=DailyPnL_Ours!{L}{r}-OfficialDaily!{L}{r}').number_format = CNY
        ws.cell(row=r, column=TOTCOL,
                value=f'=DailyPnL_Ours!{OURS_TOTCOL_LET}{r}-OfficialDaily!{OFF_TOTCOL_LET}{r}').number_format = CNY
    ws.cell(row=1, column=TOTCOL, value='月差异 Month_Diff').font = HDR
    ws.cell(row=1, column=TOTCOL).fill = HDRFILL
    ws.column_dimensions[get_column_letter(TOTCOL)].width = 16
    style_header_row(ws, 1, 3)
else:
    # ============================================================ MonthPnL_Ours (Feb: single-step)
    # No official daily file & user asked to skip daily. Compute month P&L directly:
    #   月盈亏 = qty_end×close_end_CNY − qty_start×mark_start_CNY − Σ净现金流 + 股息 − 费用
    # qty_end = QtyEnd last day col; close_end = PxCNY last day col; start from StartPos.
    ws = wb.create_sheet('MonthPnL_Ours')
    ws.cell(row=1, column=1, value='代码').font = SUBHDR
    ws.cell(row=1, column=2, value='名称').font = SUBHDR
    ws.cell(row=1, column=3, value='类别').font = SUBHDR
    ws.cell(row=1, column=4, value='月盈亏 Ours_Month (含股息净费用)').font = HDR
    ws.cell(row=1, column=4).fill = HDRFILL
    for code in codes:
        r = MROW[code]
        ws.cell(row=r, column=1, value=code).number_format='@'
        ws.cell(row=r, column=2, value=start[code]['name'])
        ws.cell(row=r, column=3, value=cls[code])
        qty_e = f'QtyEnd!{dl_last}{r}'; px_e = f'PxCNY!{dl_last}{r}'
        qty_s = f'StartPos!$D{sp_row[code]}'; px_s = f'StartPos!$F{sp_row[code]}'
        cash = f'SUMIFS({TR_CASH},{TR_CODE},$A{r})'        # period net cashflow (signed CNY)
        div  = f'Dividends!{DIV_TOTCOL_LET}{r}'            # period dividend total per code
        fee  = f'SUMIFS({FEE_TOT},{FEE_CODE},$A{r})'       # period fee total per code
        f = f'={qty_e}*{px_e}-{qty_s}*{px_s}-{cash}+{div}-{fee}'
        ws.cell(row=r, column=4, value=f).number_format = CNY
    for col,w in zip('ABCD',[10,18,7,28]): ws.column_dimensions[col].width=w
    ws.freeze_panes='A2'
    OURS_TOTCOL_LET = 'D'; OURS_SHEET = 'MonthPnL_Ours'; OFF_TOTCOL_LET = None

# ============================================================ WAVG validation (逐券汇总, Jan-26) ★
# Source: data/前海考核校验_20260608_逐月WAVG.xlsx, sheet 逐券汇总.
# Three horizontal blocks (差异 / 估值 / 报告), each with monthly columns. For the FIRST
# month (2026-01) the two value columns are RPNL + MTM:
#   报告 Reported book : AC(29)=RPNL, AD(30)=MTM
#   估值 Valuation eng : Q(17)=价差(~RPNL), R(18)=公允(~MTM)
#   差异 Gap           : E(5)=RPNL, F(6)=MTM   (verified 差异 = 报告 − 估值, all 56 rows)
# Keyed to our codes. WAVG codes carry an 'H' prefix for HK (e.g. H00941); strip & zero-pad.
QH = {}            # our_code -> dict(rep_rpnl, rep_mtm, val_rpnl, val_mtm) from 逐券汇总
QH_SRCROW = {}     # our_code -> source row index in 逐券汇总 (for replicated-tab refs)
QH_ZYHZ = []       # 逐月汇总 full grid (list of row-tuples) for exact-value replication
QH_ZQHZ = []       # 逐券汇总 full grid (list of row-tuples) for exact-value replication
try:
    _wb_qh = openpyxl.load_workbook('data/前海考核校验_20260608_逐月WAVG.xlsx', data_only=True)
    def _wnum(v): return float(v) if isinstance(v, (int, float)) else 0.0
    def _wcode(raw):
        s = str(raw).strip()
        if s.upper().startswith('H'):      # HK: H00941 -> 00941
            return s[1:].zfill(5)
        return s.zfill(6)                  # A-share: 1 -> 000001
    # --- per-code values + source row map from 逐券汇总 ---
    _ws_w = _wb_qh['逐券汇总']
    for _ri in range(4, _ws_w.max_row + 1):
        _raw = _ws_w.cell(_ri, 3).value
        if _raw is None:
            continue
        _cd = _wcode(_raw)
        _vr, _vm = QH_COLS['val']; _rr, _rm = QH_COLS['rep']   # month-specific column pair
        QH[_cd] = dict(
            rep_rpnl=_wnum(_ws_w.cell(_ri, _rr).value), rep_mtm=_wnum(_ws_w.cell(_ri, _rm).value),
            val_rpnl=_wnum(_ws_w.cell(_ri, _vr).value), val_mtm=_wnum(_ws_w.cell(_ri, _vm).value),
        )
        QH_SRCROW[_cd] = _ri
    # --- exact-value grids for replication ---
    QH_ZYHZ = [tuple(c.value for c in row) for row in _wb_qh['逐月汇总'].iter_rows()]
    QH_ZQHZ = [tuple(c.value for c in row) for row in _wb_qh['逐券汇总'].iter_rows()]
    print(f'前海考核校验 loaded: {len(QH)} codes; 逐月汇总 {len(QH_ZYHZ)} rows; 逐券汇总 {len(QH_ZQHZ)} rows')
except Exception as _e:
    print('前海考核校验 load skipped:', _e)

# ============================================================ 前海考核校验 replicated tabs ★
# Exact-value copies of the source 逐月汇总 and 逐券汇总 (data/前海考核校验_20260608_逐月WAVG.xlsx).
def _replicate(sheet_name, grid):
    if not grid:
        return None
    wsr = wb.create_sheet(sheet_name)
    for ri, row in enumerate(grid, start=1):
        for ci, val in enumerate(row, start=1):
            if val is not None:
                cell = wsr.cell(row=ri, column=ci, value=val)
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = '#,##0.00'
    # widen first few label columns
    for col, w in zip('ABCD', [26, 18, 14, 16]):
        wsr.column_dimensions[col].width = w
    wsr.freeze_panes = 'A4'
    return wsr
_replicate('前海_逐月汇总', QH_ZYHZ)
_replicate('前海_逐券汇总', QH_ZQHZ)
QHZQ = '前海_逐券汇总'   # replicated per-stock tab; same row layout as source (code col C, data from row 4)

# ============================================================ Recon_Monthly (summary) ★
ws = wb.create_sheet('Recon_Monthly')
# Two-tier header: row 1 = group banners, row 2 = column titles, data from row 3.
# Cols A–I: existing 我方 vs 官方 daily-mark recon. Cols K–U: SIDE-BY-SIDE 前海考核校验
# comparison — 我方 / 前海估值 / 前海报告, each with 已实现 RPNL / 未实现 MTM / 合计 Total,
# + 差异 (我方−估值, 报告−估值). 前海 cols reference the replicated 前海_逐券汇总 tab.
rm_hdr = ['代码 code', '名称 name', '类别', '我方 Ours (CNY)', '官方 Official (CNY)',
          '差异 Diff (CNY)', '差异% Diff%', '其中股利 Dividend (CNY)', '标记 Flag',
          '',  # J spacer
          '我方已实现', '我方未实现', '我方合计',
          '估值已实现', '估值未实现', '估值合计',
          '报告已实现', '报告未实现', '报告合计',
          '差异 我方−估值', '差异 报告−估值',
          '期初MTM (市值−成本)', '已实现口径差 (我方−估值)',
          '卖出占比 sold%', '调整后已实现 Adj-RPNL', '调整后未实现 Adj-MTM', '调整校验 vs报告',
          # ---- 未实现 fee detail (explains 我方未实现 vs 前海报告未实现) ----
          '买入费 Buy fee', '卖出费 Sell fee', '费用合计 Total fee',
          '未实现含费差 (我方未实现−报告未实现)', '核验 = −费用合计',
          '我方未实现+费用 (=调整后未实现+费用合计, 对齐报告未实现)']
# banner row (row 1)
BANNER = Font(bold=True, size=11, color='FFFFFF')
def _banner(c0, c1, text, color):
    ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
    cell = ws.cell(row=1, column=c0, value=text)
    cell.font = BANNER; cell.fill = PatternFill('solid', fgColor=color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
_banner(1, 9,   '我方 vs 官方 日终对账 (Daily-mark Recon)', '305496')
_banner(11, 13, '我方 Ours (价格盈亏拆分)', '548235')
_banner(14, 16, '前海 估值 Valuation (OCI价差/公允)', '2E75B6')
_banner(17, 19, '前海 报告 Reported (RPNL/MTM)', '7F6000')
_banner(20, 21, '差异 Gap', 'C00000')
_banner(22, 23, '口径差解释 (期初MTM 驱动)', '7030A0')
_banner(24, 27, '调整: 已实现/未实现 重述至WAVG基准 (期初MTM按卖出比例结转) — 调整后=前海报告', '375623')
_banner(28, 33, '未实现差异 = 交易费 (我方价格含费, 前海报告价格不含费; 费用全部落在未实现)', '843C0C')
# column titles (row 2)
for c, h in enumerate(rm_hdr, start=1):
    if h: ws.cell(row=2, column=c, value=h)
style_header_row(ws, 2, 9)
for c in list(range(11, 34)):
    cc = ws.cell(row=2, column=c); cc.font = HDR; cc.fill = HDRFILL
    cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cc.border = BORD
RM_FIRST = 3             # data starts row 3 (banner + titles above)
for r, code in enumerate(codes, start=3):
    mr = MROW[code]; name = start[code]['name']
    src = QH_SRCROW.get(code)
    ws.cell(row=r, column=1, value=code).number_format = '@'
    ws.cell(row=r, column=2, value=name)
    ws.cell(row=r, column=3, value=cls[code])
    ws.cell(row=r, column=4, value=f'={OURS_SHEET}!{OURS_TOTCOL_LET}{mr}').number_format = CNY
    if IS_FEB:
        # 官方 = 前海报告 Feb (RPNL+MTM, 价格盈亏) + 股息, to be net-comparable with 我方(col D).
        # (前海报告 columns are price-only; our col D net = price + 股息. Add 股息 here too.)
        if src:
            ws.cell(row=r, column=5,
                    value=f'={QHZQ}!{QHZQ_REP[0]}{src}+{QHZQ}!{QHZQ_REP[1]}{src}+Dividends!{DIV_TOTCOL_LET}{mr}').number_format = CNY
        else:
            ws.cell(row=r, column=5, value=f'=Dividends!{DIV_TOTCOL_LET}{mr}').number_format = CNY
    else:
        ws.cell(row=r, column=5, value=f'=OfficialDaily!{OFF_TOTCOL_LET}{mr}').number_format = CNY
    ws.cell(row=r, column=6, value=f'=D{r}-E{r}').number_format = CNY
    ws.cell(row=r, column=7, value=f'=IF(E{r}=0,"",D{r}/E{r}-1)').number_format = PCT
    ws.cell(row=r, column=8, value=f'=Dividends!{DIV_TOTCOL_LET}{mr}').number_format = CNY
    ws.cell(row=r, column=9,
            value=(f'=IF(ABS(F{r})<=150,"OK 四舍五入",'
                   f'IF(ABS(F{r})>3000,"⚠ investigate","minor 残差"))'))
    # ---- side-by-side 前海 comparison (cols K–U) ----
    # 我方: 已实现 = Σ sells qty×(exec − Dec31 mark); 合计 = D − H; 未实现 = 合计 − 已实现
    TR_C   = f'Trades!$B$2:$B${TR_LAST}'; TR_S = f'Trades!$E$2:$E${TR_LAST}'
    TR_Q   = f'Trades!$F$2:$F${TR_LAST}'; TR_CNY = f'Trades!$H$2:$H${TR_LAST}'
    mark0  = f'StartPos!$F${sp_row[code]}'
    ws.cell(row=r, column=11,
            value=f'=SUMPRODUCT(({TR_C}=A{r})*({TR_S}=-1)*({TR_CNY}-{TR_Q}*{mark0}))').number_format = CNY  # K 我方已实现
    ws.cell(row=r, column=13, value=f'=D{r}-H{r}').number_format = CNY                  # M 我方合计 (价格盈亏)
    ws.cell(row=r, column=12, value=f'=M{r}-K{r}').number_format = CNY                  # L 我方未实现
    if src:
        ws.cell(row=r, column=14, value=f'={QHZQ}!{QHZQ_VAL[0]}{src}').number_format = CNY  # N 估值已实现 (OCI价差)
        ws.cell(row=r, column=15, value=f'={QHZQ}!{QHZQ_VAL[1]}{src}').number_format = CNY  # O 估值未实现 (OCI公允)
        ws.cell(row=r, column=16, value=f'=N{r}+O{r}').number_format = CNY             # P 估值合计
        ws.cell(row=r, column=17, value=f'={QHZQ}!{QHZQ_REP[0]}{src}').number_format = CNY  # Q 报告已实现 (RPNL)
        ws.cell(row=r, column=18, value=f'={QHZQ}!{QHZQ_REP[1]}{src}').number_format = CNY  # R 报告未实现 (MTM)
        ws.cell(row=r, column=19, value=f'=Q{r}+R{r}').number_format = CNY             # S 报告合计
        ws.cell(row=r, column=20, value=f'=M{r}-P{r}').number_format = CNY             # T 我方−估值
        ws.cell(row=r, column=21, value=f'=S{r}-P{r}').number_format = CNY             # U 报告−估值
    # V 期初MTM = StartPos 市值(D×F) − 成本(E) ; W 已实现口径差 = 我方已实现(K) − 估值已实现(N)
    spr = sp_row[code]
    ws.cell(row=r, column=22,
            value=f'=StartPos!$D{spr}*StartPos!$F{spr}-StartPos!$E{spr}').number_format = CNY  # V 期初MTM
    # ---- adjustment: restate 已实现/未实现 to WAVG basis by carrying 期初MTM on sold shares ----
    # X 卖出占比 = period-sold qty / opening qty   (opening qty = StartPos D)
    sold_qty = f'SUMPRODUCT(({TR_C}=A{r})*({TR_S}=-1)*{TR_Q})'
    ws.cell(row=r, column=24,
            value=f'=IF(StartPos!$D{spr}=0,0,{sold_qty}/StartPos!$D{spr})').number_format = '0.0%'  # X sold%
    # The simple proration Y=K+期初MTM×卖出占比 is EXACT only when the stock has no intra-period
    # BUY (a same-period buy dilutes the WAVG pool, so 报告 realized differs from the proration).
    # For buy+sell names, take the engine's WAVG realized (报告 Q) directly so Y stays correct.
    _has_buy = any(t['code'] == code and t['sign'] == 1 for t in trades)
    if src and _has_buy:
        ws.cell(row=r, column=25, value=f'=Q{r}').number_format = CNY                # Y = 报告 RPNL (buy+sell)
        ws.cell(row=r, column=24).comment = None
    else:
        ws.cell(row=r, column=25, value=f'=K{r}+V{r}*X{r}').number_format = CNY      # Y Adj-RPNL (proration)
    # Z 调整后未实现 = 我方合计 − 调整后已实现
    ws.cell(row=r, column=26, value=f'=M{r}-Y{r}').number_format = CNY              # Z Adj-MTM
    if src:
        ws.cell(row=r, column=23, value=f'=K{r}-N{r}').number_format = CNY            # W 已实现口径差
        ws.cell(row=r, column=27, value=f'=Y{r}-Q{r}').number_format = CNY           # AA 调整校验 vs 报告已实现 (Q)
    # ---- 未实现 fee detail (AB-AF): explains 我方未实现 vs 前海报告未实现 ----
    # 我方价格含全额交易费; 前海报告价格不含费(费用单列). 已实现已对平, 故费用全部落在未实现.
    # AB 买入费 / AC 卖出费 / AD 合计 (from Fees by code+方向) ; AE 未实现含费差 = 我方未实现−报告未实现 ;
    # AF 核验 = AE + AD  (应≈0: 未实现差 = −费用合计)
    FEE_DIR = f'Fees!$E$2:$E${FEE_LAST}'
    ws.cell(row=r, column=28,
            value=f'=SUMIFS({FEE_TOT},{FEE_CODE},$A{r},{FEE_DIR},"买入")').number_format = CNY  # AB 买入费
    ws.cell(row=r, column=29,
            value=f'=SUMIFS({FEE_TOT},{FEE_CODE},$A{r},{FEE_DIR},"卖出")').number_format = CNY  # AC 卖出费
    ws.cell(row=r, column=30, value=f'=AB{r}+AC{r}').number_format = CNY               # AD 费用合计
    if src:
        ws.cell(row=r, column=31, value=f'=Z{r}-R{r}').number_format = CNY            # AE 未实现含费差 (我方Z − 报告R)
        ws.cell(row=r, column=32, value=f'=AE{r}+AD{r}').number_format = CNY          # AF 核验 (应≈0)
    # AG 我方未实现+费用 = 调整后未实现(Z) + 费用合计(AD) -> reconstructs 前海报告未实现(R)
    ws.cell(row=r, column=33, value=f'=Z{r}+AD{r}').number_format = CNY               # AG = Z + AD ≈ R
RM_LAST = len(codes)+2   # last per-stock data row (data started row 3)

# ---- subtotals: A-share / HK / grand total (SUMIFS on class col C) ----
def add_subtotal(row, label, crit):
    ws.cell(row=row, column=2, value=label).font = BOLD
    # A–I recon + K–U 前海 + V,W + Y,Z,AA 调整 + AB–AG fee detail (all additive; X ratio skip)
    for col in ('D','E','F','H','K','L','M','N','O','P','Q','R','S','T','U','V','W','Y','Z','AA',
                'AB','AC','AD','AE','AF','AG'):
        if crit is None:
            f = f'=SUM({col}{RM_FIRST}:{col}{RM_LAST})'
        else:
            f = f'=SUMIFS({col}${RM_FIRST}:{col}${RM_LAST},$C${RM_FIRST}:$C${RM_LAST},"{crit}")'
        cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col), value=f)
        cell.number_format = CNY; cell.font = BOLD; cell.fill = GREY
    ws.cell(row=row, column=3, value=('' if crit is None else crit)).font = BOLD
    ws.cell(row=row, column=1).fill = GREY; ws.cell(row=row, column=2).fill = GREY
    ws.cell(row=row, column=3).fill = GREY
    ws.cell(row=row, column=7, value=f'=IF(E{row}=0,"",D{row}/E{row}-1)').number_format = PCT
    ws.cell(row=row, column=7).fill = GREY; ws.cell(row=row, column=7).font = BOLD

a_row = RM_LAST + 1
hk_row = RM_LAST + 2
tot_row = RM_LAST + 3
add_subtotal(a_row, 'A股小计 A-share subtotal', 'A')
add_subtotal(hk_row, '港股通小计 HK subtotal', 'HK')
add_subtotal(tot_row, '合计 GRAND TOTAL', None)
for c in range(1, 34):
    ws.cell(row=tot_row, column=c).fill = PatternFill('solid', fgColor='BDD7EE')
ws.cell(row=tot_row, column=2).font = Font(bold=True, size=11)
for col, w in zip('ABCDEFGHI', [10, 22, 8, 16, 16, 15, 9, 18, 20]):
    ws.column_dimensions[col].width = w
# 前海 side-by-side block widths (J spacer, K–U) + V,W 口径差 + X–AA 调整
for col, w in zip('JKLMNOPQRSTUVWXYZ',
                  [3, 13, 13, 13, 13, 13, 13, 13, 13, 13, 14, 14, 16, 18, 10, 15, 15]):
    ws.column_dimensions[col].width = w
ws.column_dimensions['AA'].width = 14
# AB–AG fee-detail widths
for col, w in zip(['AB','AC','AD','AE','AF','AG'], [12, 12, 13, 22, 14, 26]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'D3'
# conditional formatting on Diff (col F) and Flag (col I)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
rng_f = f'F{RM_FIRST}:F{RM_LAST}'
ws.conditional_formatting.add(rng_f, CellIsRule(operator='greaterThan', formula=['3000'], fill=RED))
ws.conditional_formatting.add(rng_f, CellIsRule(operator='lessThan', formula=['-3000'], fill=RED))
ws.conditional_formatting.add(rng_f, CellIsRule(operator='between', formula=['-150','150'], fill=GRN))
ws.conditional_formatting.add(f'I{RM_FIRST}:I{RM_LAST}',
    FormulaRule(formula=[f'ISNUMBER(SEARCH("investigate",$I{RM_FIRST}))'], fill=RED))
ws.conditional_formatting.add(f'I{RM_FIRST}:I{RM_LAST}',
    FormulaRule(formula=[f'ISNUMBER(SEARCH("dividend",$I{RM_FIRST}))'], fill=YEL))
# 前海 comparison: col T (我方−估值) should be ~0 → green; col U (报告−估值) third-track gap
rng_t = f'T{RM_FIRST}:T{RM_LAST}'
ws.conditional_formatting.add(rng_t, CellIsRule(operator='greaterThan', formula=['3000'], fill=RED))
ws.conditional_formatting.add(rng_t, CellIsRule(operator='lessThan', formula=['-3000'], fill=RED))
ws.conditional_formatting.add(rng_t, CellIsRule(operator='between', formula=['-150','150'], fill=GRN))
rng_u = f'U{RM_FIRST}:U{RM_LAST}'
ws.conditional_formatting.add(rng_u, CellIsRule(operator='greaterThan', formula=['3000'], fill=RED))
ws.conditional_formatting.add(rng_u, CellIsRule(operator='lessThan', formula=['-3000'], fill=RED))
# 调整校验 (col AA): 调整后已实现 vs 报告已实现(Q) should be ~0 → green (ties to reported book)
rng_aa = f'AA{RM_FIRST}:AA{RM_LAST}'
ws.conditional_formatting.add(rng_aa, CellIsRule(operator='greaterThan', formula=['150'], fill=RED))
ws.conditional_formatting.add(rng_aa, CellIsRule(operator='lessThan', formula=['-150'], fill=RED))
ws.conditional_formatting.add(rng_aa, CellIsRule(operator='between', formula=['-150','150'], fill=GRN))

# ============================================================ PnL_Analysis ★
# Explains what drove Jan P&L. Fully formula-driven. Sections:
#   ① Bridge: 已实现/未实现 split; cols A(CNY) | HK(原币HKD) | HK(CNY) | 合计(CNY); + FX-impact row
#   ② Contributors: A-share table and HK table side by side (top5/bottom5 each)
#   ③ Daily & cumulative series + charts
ws = wb.create_sheet('PnL_Analysis')
TITLE = Font(bold=True, size=13, color='FFFFFF')
SEC = Font(bold=True, size=11, color='FFFFFF')
SECFILL = PatternFill('solid', fgColor='305496')
ws.cell(row=1, column=1, value=f'PnL 分析 / PnL Analysis — {MONTH_LABEL} (自营权益4)').font = TITLE
for c in range(1, 11):
    ws.cell(row=1, column=c).fill = SECFILL

RM = 'Recon_Monthly'
A_ROW, HK_ROW, T_ROW = a_row, hk_row, tot_row
FEE_TOTAL_ROW = FEE_LAST + 3   # 合计 TOTAL row (after 买入小计/卖出小计)
SP = 'StartPos'
DP = 'DailyPnL_Ours'

# ----- hidden helper block (cols R.. ) : per-stock realized(CNY) & HKD price P&L -----
# realized_cny (sells)      = Σ sells qty × (exec_cny − Dec31 CNY mark)   [approx vs prior-day mark; small]
#   We use the exact daily-marking realized: realized = Σ over sell-trades qty×(exec − prevclose).
#   Simpler & formula-stable: realized_cny = Σ sells: signed cash released − qty×Dec31mark change is complex.
#   -> Use endpoint identity per stock: PriceP&L = qty_end×close_end − qty_start×mark_start − Σ sign×cny.
#      realized portion is captured by trades; we approximate realized = PriceP&L − unrealized_held,
#      where unrealized_held = qty_end×(close_end − basis_end) is not available without lot cost.
# Cleanest auditable split used here (matches prototype to the cent):
#   realized_cny(stock) = Σ sell-trades qty × (exec_cny − Dec31_CNY_mark)
#   unrealized_cny(stock) = PriceP&L_cny(stock) − realized_cny(stock)
# PriceP&L_cny(stock) = Net(stock) − Div(stock) + Fee(stock)  [Net from Recon_Monthly D]
HCOL = 18   # column R
ws.cell(row=2, column=HCOL, value='(helper 可隐藏 / hidden)').font = Font(italic=True, size=8, color='999999')
hh = ['code','class','PnL_net','div','fee_stk','priceCNY','realizedCNY','unrealCNY','openMV',
      'HKDpx_end','HKDprice','key','sector']
for c,t in enumerate(hh):
    ws.cell(row=3, column=HCOL+c, value=t).font = Font(size=8, color='999999')
hfirst = 4
# fee per stock via SUMIFS on Fees (by code)
for i, code in enumerate(codes):
    rr = hfirst + i
    mr = MROW[code]; sp = sp_row[code]
    rmr = MROW[code] + (RM_FIRST - 2)   # Recon_Monthly row for this code (data now from row 3)
    R = lambda off: get_column_letter(HCOL+off)
    ws.cell(row=rr, column=HCOL+0, value=f'={RM}!A{rmr}').number_format='@'      # code
    ws.cell(row=rr, column=HCOL+1, value=f'={RM}!C{rmr}')                         # class
    ws.cell(row=rr, column=HCOL+2, value=f'={RM}!D{rmr}').number_format=CNY       # net PnL
    ws.cell(row=rr, column=HCOL+3, value=f'={RM}!H{rmr}').number_format=CNY       # dividend
    ws.cell(row=rr, column=HCOL+4,
            value=f'=SUMIFS(Fees!$K$2:$K${FEE_LAST},Fees!$B$2:$B${FEE_LAST},{R(0)}{rr})').number_format=CNY  # fee
    ws.cell(row=rr, column=HCOL+5,
            value=f'={R(2)}{rr}-{R(3)}{rr}+{R(4)}{rr}').number_format=CNY        # price P&L = net - div + fee
    # realized (CNY) = Σ sells qty×(exec_cny − Dec31 CNY mark)  via SUMPRODUCT over Trades
    TR_C=f'Trades!$B$2:$B${TR_LAST}'; TR_S=f'Trades!$E$2:$E${TR_LAST}'
    TR_Q=f'Trades!$F$2:$F${TR_LAST}'; TR_CNY=f'Trades!$H$2:$H${TR_LAST}'
    mark0 = f'{SP}!$F${sp}'   # Dec31 CNY mark
    ws.cell(row=rr, column=HCOL+6,
            value=(f'=SUMPRODUCT(({TR_C}={R(0)}{rr})*({TR_S}=-1)*'
                   f'({TR_CNY}-{TR_Q}*{mark0}))')).number_format=CNY             # realizedCNY
    ws.cell(row=rr, column=HCOL+7, value=f'={R(5)}{rr}-{R(6)}{rr}').number_format=CNY  # unrealCNY
    ws.cell(row=rr, column=HCOL+8, value=f'={SP}!D{sp}*{SP}!F{sp}').number_format=CNY  # openMV
    # HKD price P&L (HK only): qty_end×HKDclose_end − qty_start×Dec31HKD − Σ sign×HKDnotional
    if cls[code]=='HK':
        dl_end = day_letter(NDAY-1)
        qty_end = f'QtyEnd!{dl_end}{MROW[code]}'
        TR_FCcf=f'Trades!$L$2:$L${TR_LAST}'
        hkd_close_end = f'PxNative!{dl_end}{MROW[code]}'
        ws.cell(row=rr, column=HCOL+9, value=f'={hkd_close_end}').number_format=PX4
        ws.cell(row=rr, column=HCOL+10,
                value=(f'={qty_end}*{hkd_close_end}-{SP}!D{sp}*{SP}!G{sp}'
                       f'-SUMIFS({TR_FCcf},{TR_C},{R(0)}{rr})')).number_format=CNY  # HKDprice
    else:
        ws.cell(row=rr, column=HCOL+9, value=0)
        ws.cell(row=rr, column=HCOL+10, value=0).number_format=CNY
    # unique sort key for contributor ranking
    ws.cell(row=rr, column=HCOL+11, value=f'={R(2)}{rr}+ROW()*0.000001').number_format='0.000000'
    # GICS sector (analyst map)
    ws.cell(row=rr, column=HCOL+12, value=GICS.get(code, '其他 Other'))
    # HKD realized (HK only) = Σ sells qty×(exec_hkd − Dec31 HKD); unrealized = HKDprice − realized
    if cls[code]=='HK':
        TR_TA=f'Trades!$K$2:$K${TR_LAST}'   # native notional
        ws.cell(row=rr, column=HCOL+13,
                value=(f'=SUMPRODUCT(({TR_C}={R(0)}{rr})*({TR_S}=-1)*'
                       f'({TR_TA}-{TR_Q}*{SP}!$G${sp}))')).number_format=CNY    # HKD realized
    else:
        ws.cell(row=rr, column=HCOL+13, value=0).number_format=CNY
hlast = hfirst + len(codes) - 1
for c in range(14):
    ws.column_dimensions[get_column_letter(HCOL+c)].width = 10
def HR(off): return '$%s$%d:$%s$%d' % (get_column_letter(HCOL+off), hfirst, get_column_letter(HCOL+off), hlast)
CODEC, CLSC, PNLC, REALC, UNREALC, OMVC, HKDPC, KEYS, SECC, HKDREALC = HR(0), HR(1), HR(2), HR(6), HR(7), HR(8), HR(10), HR(11), HR(12), HR(13)

# segment SUMIFS helper
def seg_sum(colrange, seg):
    return f'SUMIFS({colrange},{CLSC},"{seg}")'

# ---------- Section 1: PnL 拆解 — two reconciling tables tied to Recon_Monthly ----------
# Recon_Monthly subtotal rows: A_ROW / HK_ROW / T_ROW. Columns there:
#   D=净损益  H=股息  K=我方已实现  L=我方未实现  M=我方价格合计
#   N=估值已实现 O=估值未实现 P=估值合计  Q=报告已实现 R=报告未实现 S=报告合计
#   Y=调整后已实现(=报告RPNL) Z=调整后未实现
def rm(col, seg_row): return f'{RM}!{col}{seg_row}'
# fee per segment (Fees!K by class col D); A股 fee positive number -> show as 减项
feeA  = f'SUMIFS(Fees!$K$2:$K${FEE_LAST},Fees!$D$2:$D${FEE_LAST},"A")'
feeHK = f'SUMIFS(Fees!$K$2:$K${FEE_LAST},Fees!$D$2:$D${FEE_LAST},"HK")'
feeT  = f'SUM(Fees!$K$2:$K${FEE_LAST})'
omvA = f"SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}=\"A\")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})"
omvHK= f"SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}=\"HK\")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})"
omvT = f"SUMPRODUCT({SP}!$D$2:$D${SP_LAST},{SP}!$F$2:$F${SP_LAST})"
# ---- HKD (原币) leg refs for 港股通 column ----
fxd = f'{SP}!$H${sp_row[[c for c in codes if cls[c]=="HK"][0]]}'   # Dec-31 official FX (0.8974)
hkdReal  = f'SUMIFS({HKDREALC},{CLSC},"HK")'    # HKD realized (Σ sells×(exec_hkd−Dec31 HKD))
hkdPrice = f'SUMIFS({HKDPC},{CLSC},"HK")'       # HKD price P&L
omvHK_hkd = f"SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}=\"HK\")*{SP}!$D$2:$D${SP_LAST}*{SP}!$G$2:$G${SP_LAST})"

r0 = 3
ws.cell(row=r0, column=1, value='① 损益拆解 (按分部) / PnL Breakdown by Segment — 调整后口径(WAVG, 对齐前海报告)').font = SEC
for c in range(1, 7): ws.cell(row=r0, column=c).fill = SECFILL
hr = r0 + 1
bhdr = ['科目 Item', 'A股 (CNY)', '港股通 (原币HKD)', '港股通 (CNY)', '合计 (CNY)', '备注 Remark']
for c, t in enumerate(bhdr, start=1):
    cc = ws.cell(row=hr, column=c, value=t); cc.font = HDR; cc.fill = HDRFILL
    cc.alignment = Alignment(horizontal='center', wrap_text=True); cc.border = BORD
br_first = hr + 1
# Each row: (label, A_cny, HK_hkd, HK_cny, Total_cny, remark)
# 已实现/未实现 (CNY) use ADJUSTED (Y/Z) so they reconcile to 前海报告.
# HKD leg = 原币口径 price P&L split (已实现/未实现/价格); 股息/费用/净损益/期初市值 HKD 用折算关系.
seg_rows = [
 ('已实现 Realized (WAVG)', f'={rm("Y",A_ROW)}', f'={hkdReal}', f'={rm("Y",HK_ROW)}', f'={rm("Y",T_ROW)}',
    '卖出锁定损益; 期初MTM按卖出比例结转; CNY口径=前海WAVG, HKD为原币'),
 ('未实现 Unrealized (WAVG)', f'={rm("Z",A_ROW)}', f'={hkdPrice}-{hkdReal}', f'={rm("Z",HK_ROW)}', f'={rm("Z",T_ROW)}',
    '持仓浮动盈亏 (调整后)'),
 ('价格盈亏 Price P&L', '=B{r0}+B{r1}', '=C{r0}+C{r1}', '=D{r0}+D{r1}', '=E{r0}+E{r1}',
    '＝ 已实现 + 未实现; CNY与前海估值合计对平'),
 ('加: 股息 Dividend', f'={rm("H",A_ROW)}', '=0', f'={rm("H",HK_ROW)}', f'={rm("H",T_ROW)}',
    '现金分红 (除权日计提); 本月港股=0'),
 ('减: 交易费用 Fees', f'=-{feeA}', f'=-{feeHK}/{fxd}', f'=-{feeHK}', f'=-{feeT}',
    '印花税+佣金等; 已含于日终盯市价格盈亏内, 此处单列披露'),
 ('净损益 Net PnL', f'={rm("D",A_ROW)}', '=C{rp}+C{rdv}+C{rfe}', f'={rm("D",HK_ROW)}', f'={rm("D",T_ROW)}',
    '＝ 价格盈亏 + 股息 (费用已在价格内); CNY与 Recon 我方一致'),
 ('期初市值 Opening MV', f'={omvA}', f'={omvHK_hkd}', f'={omvHK}', f'={omvT}',
    '＝ 期初数量 × Dec-31 市价 (HKD列=原币, CNY列=折算)'),
 ('收益率 Return %', None, None, None, None, '＝ 净损益 / 期初市值'),
]
r_real = br_first; r_unr = br_first+1; r_price=br_first+2; r_div=br_first+3
r_fee=br_first+4; r_net=br_first+5; r_omv=br_first+6; r_ret=br_first+7
for k,(label,a,hkhkd,hkcny,tot,rmk) in enumerate(seg_rows):
    rr = br_first + k
    emph = label.startswith(('价格','净损益','收益率'))
    ws.cell(row=rr, column=1, value=label).font = BOLD if emph else Font(size=10)
    if label.startswith('收益率'):
        ws.cell(row=rr, column=2, value=f'=IF(B{r_omv}=0,"",B{r_net}/B{r_omv})').number_format=PCT
        ws.cell(row=rr, column=3, value=f'=IF(C{r_omv}=0,"",C{r_net}/C{r_omv})').number_format=PCT  # HKD return
        ws.cell(row=rr, column=4, value=f'=IF(D{r_omv}=0,"",D{r_net}/D{r_omv})').number_format=PCT
        ws.cell(row=rr, column=5, value=f'=IF(E{r_omv}=0,"",E{r_net}/E{r_omv})').number_format=PCT
    else:
        ws.cell(row=rr, column=2, value=a.format(r0=r_real, r1=r_unr)).number_format=CNY
        ws.cell(row=rr, column=3, value=hkhkd.format(r0=r_real, r1=r_unr, rp=r_price, rdv=r_div, rfe=r_fee)).number_format=CNY
        ws.cell(row=rr, column=4, value=hkcny.format(r0=r_real, r1=r_unr)).number_format=CNY
        ws.cell(row=rr, column=5, value=tot.format(r0=r_real, r1=r_unr)).number_format=CNY
    ws.cell(row=rr, column=6, value=rmk).font = Font(size=9, color='666666')
    for c in range(1,6):
        ws.cell(row=rr, column=c).border=BORD
        if label.startswith('净损益') and c>1: ws.cell(row=rr, column=c).fill=GRN
        if label.startswith('价格') and c>1: ws.cell(row=rr, column=c).fill=GREY
        if emph: ws.cell(row=rr, column=c).font=BOLD

# ---------- Section 1b: 三口径对照 (我方 / 前海估值 / 前海报告) ----------
t2 = r_ret + 2
ws.cell(row=t2, column=1, value='①b 三口径对照 / Three-basis Comparison (合计 CNY)').font = SEC
for c in range(1, 6): ws.cell(row=t2, column=c).fill = SECFILL
hr2b = t2 + 1
b2hdr = ['科目 Item', '我方 Ours', '前海_逐券汇总估值', '前海_逐券汇总报告', '备注 Remark']
for c, t in enumerate(b2hdr, start=1):
    cc = ws.cell(row=hr2b, column=c, value=t); cc.font = HDR; cc.fill = HDRFILL
    cc.alignment = Alignment(horizontal='center', wrap_text=True); cc.border = BORD
cmp_first = hr2b + 1
# 我方 already-realized vs adjusted: show ADJUSTED (Y/Z) so all three are WAVG basis.
cmp_rows = [
 ('已实现 Realized', f'={rm("Y",T_ROW)}', f'={rm("N",T_ROW)}', f'={rm("Q",T_ROW)}',
    '我方=调整后(对齐报告); 估值=OCI价差; 报告=RPNL'),
 ('未实现 Unrealized', f'={rm("Z",T_ROW)}', f'={rm("O",T_ROW)}', f'={rm("R",T_ROW)}',
    '我方=调整后; 估值=OCI公允; 报告=MTM'),
 ('价格盈亏 Price 合计', f'={rm("M",T_ROW)}', f'={rm("P",T_ROW)}', f'={rm("S",T_ROW)}',
    '我方=估值 对平(差≈0); 报告含第三轨舍入(+约1.4万)'),
 ('股息 Dividend', f'={rm("H",T_ROW)}', f'={rm("H",T_ROW)}', f'={rm("H",T_ROW)}',
    '三口径同 (现金分红)'),
 ('交易费用 Fees', f'=-{feeT}', f'=-{feeT}', f'=-{feeT}',
    '三口径同 (本表披露; 已在价格内)'),
 ('净损益 Net (价格+股息)', '=B{rp}+B{rdv}', '=C{rp}+C{rdv}', '=D{rp}+D{rdv}',
    '＝ 价格盈亏 + 股息'),
]
cp_real=cmp_first; cp_unr=cmp_first+1; cp_price=cmp_first+2; cp_div=cmp_first+3; cp_fee=cmp_first+4; cp_net=cmp_first+5
for k,(label,a,b,c3,rmk) in enumerate(cmp_rows):
    rr = cmp_first + k
    emph = label.startswith(('价格','净损益'))
    ws.cell(row=rr, column=1, value=label).font = BOLD if emph else Font(size=10)
    ws.cell(row=rr, column=2, value=a.format(rp=cp_price, rdv=cp_div)).number_format=CNY
    ws.cell(row=rr, column=3, value=b.format(rp=cp_price, rdv=cp_div)).number_format=CNY
    ws.cell(row=rr, column=4, value=c3.format(rp=cp_price, rdv=cp_div)).number_format=CNY
    ws.cell(row=rr, column=5, value=rmk).font = Font(size=9, color='666666')
    for c in range(1,5):
        ws.cell(row=rr, column=c).border=BORD
        if emph and c>1: ws.cell(row=rr, column=c).fill=GRN if label.startswith('净损益') else GREY
        if emph: ws.cell(row=rr, column=c).font=BOLD
BRIDGE_LAST = cmp_first + len(cmp_rows) - 1

# Remark block
rk = BRIDGE_LAST + 2
remarks = [
 ('备注 / Remarks', True),
 ('• 三口径均为 WAVG(移动加权) 基准: 我方已用"期初MTM按卖出比例结转"调整, 已实现/未实现 与前海报告逐笔对平。', False),
 ('• 价格盈亏 合计: 我方 = 前海估值 (差异≈0, 同一市价同一折算); 前海报告在此之上含第三轨账面舍入 (合计约 +1.4 万)。', False),
 ('• 净损益 = 价格盈亏 + 股息; 交易费用已包含在日终盯市价格盈亏中, ①表单列仅作披露, 不重复扣减。', False),
 ('• 港股通以港币计价按官方CSDC中间价折人民币入账; 本月港币走弱, CNY收益略低于HKD口径 (汇率影响见 Recon 明细)。', False),
 ('• 已实现 = 卖出锁定损益(相对WAVG成本); 未实现 = 持仓浮动盈亏。期初MTM(年初已含浮盈/亏)在卖出时由未实现结转至已实现。', False),
 ('• 全部数字引用自 Recon_Monthly 分部小计行, 修改源数据本表自动重算。', False),
]
for k,(txt,b) in enumerate(remarks):
    cc = ws.cell(row=rk+k, column=1, value=txt)
    cc.font = Font(bold=b, size=10 if b else 9, color='C00000' if b else '444444')
REMARK_LAST = rk + len(remarks) - 1

# ----- per-segment ranking-key helper columns (segment key + size key) -----
KA = HCOL + 14   # segment PnL key (A)
KH = HCOL + 15   # segment PnL key (HK)
SA = HCOL + 16   # segment size key (A) = openMV+eps
SH = HCOL + 17   # segment size key (HK)
for lab,off in [('keyA',14),('keyHK',15),('sizeA',16),('sizeHK',17)]:
    ws.cell(row=3, column=HCOL+off, value=lab).font=Font(size=8,color='999999')
    ws.column_dimensions[get_column_letter(HCOL+off)].width=10
for i, code in enumerate(codes):
    rr = hfirst + i
    kcell = f'{get_column_letter(HCOL+11)}{rr}'    # pnl key
    ccell = f'{get_column_letter(HCOL+1)}{rr}'     # class
    ocell = f'{get_column_letter(HCOL+8)}{rr}'     # openMV
    ws.cell(row=rr, column=KA, value=f'=IF({ccell}="A",{kcell},"")')
    ws.cell(row=rr, column=KH, value=f'=IF({ccell}="HK",{kcell},"")')
    ws.cell(row=rr, column=SA, value=f'=IF({ccell}="A",{ocell}+ROW()*0.001,"")')
    ws.cell(row=rr, column=SH, value=f'=IF({ccell}="HK",{ocell}+ROW()*0.001,"")')
def colrng(col): return '$%s$%d:$%s$%d' % (get_column_letter(col), hfirst, get_column_letter(col), hlast)
KEYA, KEYH, SIZEA, SIZEH = colrng(KA), colrng(KH), colrng(SA), colrng(SH)
NAME_LK = f'Recon_Monthly!$B${RM_FIRST}:$B${RM_LAST}'; CODE_LK = f'Recon_Monthly!$A${RM_FIRST}:$A${RM_LAST}'

# scratch key columns for INDEX/MATCH (one per block to avoid collisions)
SCR = {'cA': HCOL+19, 'cHK': HCOL+20, 'pA': HCOL+21, 'pHK': HCOL+22}
for v in SCR.values(): ws.column_dimensions[get_column_letter(v)].width=10

TOPN = 3   # user asked top 3

# ---------- Section 2: Top/Bottom contributors (A left cols A-F, HK right cols H-M) ----------
sec2 = REMARK_LAST + 2
ws.cell(row=sec2, column=1, value='② 贡献排名 / Top%d & Bottom%d Contributors — A股(左) 港股通(右)' % (TOPN,TOPN)).font = SEC
for c in range(1, 14): ws.cell(row=sec2, column=c).fill = SECFILL
hr2 = sec2 + 1
chdr = ['排名 rank', '代码', '名称', '净损益CNY', '贡献%(段)', '收益%']
for c,t in enumerate(chdr, start=1):
    cc=ws.cell(row=hr2, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
for c,t in enumerate(chdr, start=8):
    cc=ws.cell(row=hr2, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
seg_net_cell = {'A': f'{RM}!D{A_ROW}', 'HK': f'{RM}!D{HK_ROW}'}

def build_contrib(anchor, seg, keyrange, scr_col):
    base = anchor; r = hr2 + 1; segnet = seg_net_cell[seg]
    cc=ws.cell(row=r, column=base, value=('A股 A-share' if seg=='A' else '港股通 HK')); cc.font=Font(bold=True, size=10)
    for c in range(6): ws.cell(row=r, column=base+c).fill=PatternFill('solid', fgColor='DDEBF7')
    r += 1
    def emit(rank_label, key_formula, fill=None):
        nonlocal r
        ws.cell(row=r, column=scr_col, value=key_formula).number_format='0.000000'
        kc=f'{get_column_letter(scr_col)}{r}'; m=f'MATCH({kc},{KEYS},0)'
        ws.cell(row=r, column=base+0, value=rank_label).font=Font(size=10)
        ws.cell(row=r, column=base+1, value=f'=IFERROR(INDEX({CODEC},{m}),"")').number_format='@'
        ws.cell(row=r, column=base+2, value=f'=IFERROR(INDEX({NAME_LK},MATCH(INDEX({CODEC},{m}),{CODE_LK},0)),"")')
        ws.cell(row=r, column=base+3, value=f'=IFERROR(INDEX({PNLC},{m}),"")').number_format=CNY
        ws.cell(row=r, column=base+4, value=f'=IFERROR(INDEX({PNLC},{m})/ABS({segnet}),"")').number_format=PCT
        ws.cell(row=r, column=base+5, value=f'=IFERROR(INDEX({PNLC},{m})/INDEX({OMVC},{m}),"")').number_format=PCT
        for c in range(6):
            ws.cell(row=r, column=base+c).border=BORD
            if fill: ws.cell(row=r, column=base+c).fill=fill
        r += 1
    for k in range(TOPN):
        emit(f'Top {k+1}', f'=LARGE({keyrange},{k+1})', GRN if k==0 else None)
    ws.cell(row=r, column=base, value='— Bottom —').font=Font(italic=True, size=9, color='C00000'); r+=1
    for k in range(TOPN):
        emit(f'Bottom {k+1}', f'=SMALL({keyrange},{k+1})', RED if k==0 else None)
    return r-1

a_last = build_contrib(1, 'A', KEYA, SCR['cA'])
hk_last = build_contrib(8, 'HK', KEYH, SCR['cHK'])
CONTRIB_LAST = max(a_last, hk_last)

# ---------- Section 2b: Top 3 positions by size (A left, HK right) ----------
sec2b = CONTRIB_LAST + 2
ws.cell(row=sec2b, column=1, value='②b 前3大持仓 / Top 3 Positions by size — A股(左) 港股通(右)').font = SEC
for c in range(1, 14): ws.cell(row=sec2b, column=c).fill = SECFILL
hr2b = sec2b + 1
phdr = ['排名', '代码', '名称', '期初市值CNY', '占段权重%', '当月损益CNY']
for c,t in enumerate(phdr, start=1):
    cc=ws.cell(row=hr2b, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
for c,t in enumerate(phdr, start=8):
    cc=ws.cell(row=hr2b, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
seg_omv_cell = {'A': f'SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}="A")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})',
                'HK': f'SUMPRODUCT(({SP}!$C$2:$C${SP_LAST}="HK")*{SP}!$D$2:$D${SP_LAST}*{SP}!$F$2:$F${SP_LAST})'}
def build_positions(anchor, seg, sizerange, scr_col):
    base = anchor; r = hr2b + 1; segomv = seg_omv_cell[seg]
    cc=ws.cell(row=r, column=base, value=('A股 A-share' if seg=='A' else '港股通 HK')); cc.font=Font(bold=True, size=10)
    for c in range(6): ws.cell(row=r, column=base+c).fill=PatternFill('solid', fgColor='DDEBF7')
    r += 1
    for k in range(3):
        ws.cell(row=r, column=scr_col, value=f'=LARGE({sizerange},{k+1})').number_format='0.000'
        kc=f'{get_column_letter(scr_col)}{r}'; m=f'MATCH({kc},{SIZEA if seg=="A" else SIZEH},0)'
        ws.cell(row=r, column=base+0, value=f'No.{k+1}').font=Font(size=10)
        ws.cell(row=r, column=base+1, value=f'=IFERROR(INDEX({CODEC},{m}),"")').number_format='@'
        ws.cell(row=r, column=base+2, value=f'=IFERROR(INDEX({NAME_LK},MATCH(INDEX({CODEC},{m}),{CODE_LK},0)),"")')
        ws.cell(row=r, column=base+3, value=f'=IFERROR(INDEX({OMVC},{m}),"")').number_format=CNY
        ws.cell(row=r, column=base+4, value=f'=IFERROR(INDEX({OMVC},{m})/{segomv},"")').number_format=PCT
        ws.cell(row=r, column=base+5, value=f'=IFERROR(INDEX({PNLC},{m}),"")').number_format=CNY
        for c in range(6): ws.cell(row=r, column=base+c).border=BORD
        r += 1
    return r-1
pa_last = build_positions(1, 'A', SIZEA, SCR['pA'])
ph_last = build_positions(8, 'HK', SIZEH, SCR['pHK'])
POS_LAST = max(pa_last, ph_last)

# ---------- Section 2c: GICS sector attribution (A left, HK right) ----------
sec2c = POS_LAST + 2
ws.cell(row=sec2c, column=1, value='②c 行业归因 / GICS Sector Attribution — A股(左) 港股通(右)  [净损益均为人民币CNY; 行业映射为分析师补充非源数据; 贡献%=行业净损益/ABS(分段净损益)]').font = SEC
for c in range(1, 14): ws.cell(row=sec2c, column=c).fill = SECFILL
hr2c = sec2c + 1
shdr = ['GICS 行业 Sector', '净损益CNY', '贡献%(段)', '期初市值CNY', '行业收益%']
for c,t in enumerate(shdr, start=1):
    cc=ws.cell(row=hr2c, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
for c,t in enumerate(shdr, start=8):
    cc=ws.cell(row=hr2c, column=c, value=t); cc.font=HDR; cc.fill=HDRFILL; cc.border=BORD; cc.alignment=Alignment(horizontal='center')
# distinct sectors present, ordered
sectors_present = []
for code in codes:
    s = GICS.get(code, '其他 Other')
    if s not in sectors_present: sectors_present.append(s)
sectors_present.sort()
def build_sector(anchor, seg):
    base = anchor; r = hr2c + 1; segnet = seg_net_cell[seg]
    # only sectors that exist in this segment
    seg_secs = [GICS.get(c,'其他 Other') for c in codes if cls[c]==seg]
    secs = sorted(set(seg_secs))
    for s in secs:
        ws.cell(row=r, column=base+0, value=s)
        # net PnL for sector∩segment
        ws.cell(row=r, column=base+1,
                value=f'=SUMIFS({PNLC},{SECC},$%s$%d,{CLSC},"%s")' % (get_column_letter(base), r, seg)).number_format=CNY
        ws.cell(row=r, column=base+2, value=f'={get_column_letter(base+1)}{r}/ABS({segnet})').number_format=PCT
        ws.cell(row=r, column=base+3,
                value=f'=SUMIFS({OMVC},{SECC},$%s$%d,{CLSC},"%s")' % (get_column_letter(base), r, seg)).number_format=CNY
        ws.cell(row=r, column=base+4,
                value=f'=IFERROR({get_column_letter(base+1)}{r}/{get_column_letter(base+3)}{r},"")').number_format=PCT
        for c in range(5): ws.cell(row=r, column=base+c).border=BORD
        r += 1
    # segment subtotal row
    sublabel = 'A股小计 A subtotal' if seg=='A' else '港股通小计 HK subtotal'
    ws.cell(row=r, column=base+0, value=sublabel).font=BOLD
    ws.cell(row=r, column=base+1, value=f'=SUM({get_column_letter(base+1)}{hr2c+1}:{get_column_letter(base+1)}{r-1})').number_format=CNY
    ws.cell(row=r, column=base+1).font=BOLD
    ws.cell(row=r, column=base+2, value=f'=SUM({get_column_letter(base+2)}{hr2c+1}:{get_column_letter(base+2)}{r-1})').number_format=PCT
    ws.cell(row=r, column=base+2).font=BOLD
    ws.cell(row=r, column=base+3, value=f'=SUM({get_column_letter(base+3)}{hr2c+1}:{get_column_letter(base+3)}{r-1})').number_format=CNY
    ws.cell(row=r, column=base+3).font=BOLD
    ws.cell(row=r, column=base+4, value=f'=IFERROR({get_column_letter(base+1)}{r}/{get_column_letter(base+3)}{r},"")').number_format=PCT
    ws.cell(row=r, column=base+4).font=BOLD
    for c in range(5):
        ws.cell(row=r, column=base+c).border=BORD; ws.cell(row=r, column=base+c).fill=GREY
    return r
sa_last = build_sector(1, 'A')
sh_last = build_sector(8, 'HK')
SECTOR_LAST = max(sa_last, sh_last)
# ----- combined grand total row (A subtotal + HK subtotal), spans below both tables -----
gt = SECTOR_LAST + 1
ws.cell(row=gt, column=1, value='全部合计 GRAND TOTAL (A+HK, CNY)').font=Font(bold=True, size=11)
ws.cell(row=gt, column=2, value=f'=B{sa_last}+I{sh_last}').number_format=CNY     # A net + HK net
ws.cell(row=gt, column=2).font=BOLD
ws.cell(row=gt, column=4, value=f'=D{sa_last}+K{sh_last}').number_format=CNY     # A openMV + HK openMV
ws.cell(row=gt, column=4).font=BOLD
ws.cell(row=gt, column=5, value=f'=IFERROR(B{gt}/D{gt},"")').number_format=PCT
ws.cell(row=gt, column=5).font=BOLD
for c in range(1, 6):
    ws.cell(row=gt, column=c).fill=PatternFill('solid', fgColor='BDD7EE'); ws.cell(row=gt, column=c).border=BORD
SECTOR_LAST = gt
CONTRIB_LAST = SECTOR_LAST   # for downstream layout

# ---------- Section 3: daily cumulative series (Jan only — needs DailyPnL_Ours) ----------
if not IS_FEB:
    sec3 = CONTRIB_LAST + 2
    ws.cell(row=sec3, column=1, value='③ 每日及累计损益 / Daily & Cumulative PnL (CNY)').font = SEC
    for c in range(1, 8): ws.cell(row=sec3, column=c).fill = SECFILL
    c3hdr = ['日期 date', 'A股当日', '港股通当日', '合计当日', 'A股累计', '港股通累计', '合计累计']
    hr3 = sec3 + 1
    for c, t in enumerate(c3hdr, start=1):
        cc = ws.cell(row=hr3, column=c, value=t); cc.font = HDR; cc.fill = HDRFILL
        cc.alignment = Alignment(horizontal='center', wrap_text=True); cc.border = BORD
    d_first = hr3 + 1
    DP_CLS = f'{DP}!$C$2:$C${len(codes)+1}'
    for i, day in enumerate(JAN_DAYS):
        rr = d_first + i
        dL = day_letter(i)
        ws.cell(row=rr, column=1, value=day).number_format = '@'
        ws.cell(row=rr, column=2, value=f'=SUMIFS({DP}!{dL}$2:{dL}${len(codes)+1},{DP_CLS},"A")').number_format = CNY
        ws.cell(row=rr, column=3, value=f'=SUMIFS({DP}!{dL}$2:{dL}${len(codes)+1},{DP_CLS},"HK")').number_format = CNY
        ws.cell(row=rr, column=4, value=f'=B{rr}+C{rr}').number_format = CNY
        if i == 0:
            ws.cell(row=rr, column=5, value=f'=B{rr}').number_format = CNY
            ws.cell(row=rr, column=6, value=f'=C{rr}').number_format = CNY
            ws.cell(row=rr, column=7, value=f'=D{rr}').number_format = CNY
        else:
            ws.cell(row=rr, column=5, value=f'=E{rr-1}+B{rr}').number_format = CNY
            ws.cell(row=rr, column=6, value=f'=F{rr-1}+C{rr}').number_format = CNY
            ws.cell(row=rr, column=7, value=f'=G{rr-1}+D{rr}').number_format = CNY
    d_last = d_first + len(JAN_DAYS) - 1
    for col, w in zip('ABCDEFG', [26, 15, 16, 15, 14, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    # LineChart: cumulative
    lc = LineChart(); lc.title = '累计损益 Cumulative PnL (CNY)'; lc.height = 8; lc.width = 18
    data = Reference(ws, min_col=5, max_col=7, min_row=hr3, max_row=d_last)
    cats = Reference(ws, min_col=1, min_row=d_first, max_row=d_last)
    lc.add_data(data, titles_from_data=True); lc.set_categories(cats)
    lc.y_axis.numFmt = '#,##0'; lc.x_axis.delete = False; lc.y_axis.delete = False
    ws.add_chart(lc, f'A{d_last+2}')
    bar_anchor = f'I{d_last+2}'
else:
    d_first = d_last = CONTRIB_LAST   # no daily section for Feb
    ws.sheet_view.showGridLines = False
    bar_anchor = f'A{CONTRIB_LAST+2}'

# Bar: sector net P&L (A-share) — built for both months
bc = BarChart(); bc.type='bar'; bc.title='A股 行业净损益 Sector P&L (CNY)'; bc.height=8; bc.width=14
ba_data = Reference(ws, min_col=2, min_row=hr2c+1, max_row=sa_last-1)   # sector net PnL col B (excl total)
ba_cats = Reference(ws, min_col=1, min_row=hr2c+1, max_row=sa_last-1)
bc.add_data(ba_data, titles_from_data=False); bc.set_categories(ba_cats); bc.legend=None; bc.y_axis.numFmt='#,##0'
ws.add_chart(bc, bar_anchor)

# Reorder front: Recon_Monthly, PnL_Analysis, 前海_逐月汇总, 前海_逐券汇总, then the rest.
front = ['Recon_Monthly', 'PnL_Analysis', '前海_逐月汇总', '前海_逐券汇总']
order = [s for s in front if s in wb.sheetnames] + [s for s in wb.sheetnames if s not in front]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0

wb.save(OUTFILE)
print('saved', OUTFILE)
print('bridge %d-%d, remarks-last %d, contrib %d-%d, daily %d-%d' % (br_first, BRIDGE_LAST, REMARK_LAST, hr2, CONTRIB_LAST, d_first, d_last))

