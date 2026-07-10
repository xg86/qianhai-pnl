
# -*- coding: utf-8 -*-
# python _extract.py            (January, default)
# PNL_MONTH=feb python _extract.py   (February)
# Step 1 of 2: read the raw source files and write the intermediate json
# (consumed by _build_workbook.py). Run this first.
import openpyxl, csv, sys, io, json, os
from collections import defaultdict
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PNL_MONTH = os.environ.get('PNL_MONTH', 'may').lower()   # 'jan' or 'feb'
OUT_JSON = '_intermediate.json' if PNL_MONTH == 'jan' else f'_intermediate_{PNL_MONTH}.json'

def classify(code): return 'HK' if len(str(code)) == 5 else 'A'
def px_code(code):
    code = str(code)
    if classify(code) == 'HK': return str(int(code)).zfill(4) + '.HK'
    return code + ('.SH' if code.startswith('6') else '.SZ')

# ---- start positions (期初持仓 = Dec-31 book) ----
wb = openpyxl.load_workbook('data/start position.xlsx', data_only=True); ws = wb.active
start_dec31 = {}
for r in list(ws.iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    code = str(r[0])
    start_dec31[code] = dict(code=code, name=r[1], qty=r[2], cost=r[3], px_cny=r[4], mv=r[5], cls=classify(code))

# ---- all 2026 trades (split into Jan / Feb below) ----
wb = openpyxl.load_workbook('data/transaction hist.xlsx', data_only=True); ws = wb.active
rows = list(ws.iter_rows(values_only=True)); h = rows[0]
def c(n): return h.index(n)
all_trades = []
for r in rows[1:]:
    if r[c('证券代码')] is None: continue
    d = r[c('发生日期')]
    dt = datetime.strptime(d, '%Y-%m-%d').date() if isinstance(d, str) else d.date()
    if dt.year != 2026: continue
    code = str(r[c('证券代码')]); sign = 1 if r[c('委托方向')] == '买入' else -1
    qty = r[c('成交数量')]
    all_trades.append(dict(date=dt.isoformat(), month=dt.month, code=code, name=r[c('证券名称')],
                           dir=r[c('委托方向')], sign=sign, qty=qty, signed_qty=sign*qty,
                           cny_amt=r[c('本币成交金额')], trade_amt=r[c('成交金额')],
                           cls=classify(code), mkt=r[c('交易市场')]))

JAN_DAYS = ['2026-01-02','2026-01-05','2026-01-06','2026-01-07','2026-01-08','2026-01-09',
            '2026-01-12','2026-01-13','2026-01-14','2026-01-15','2026-01-16','2026-01-19',
            '2026-01-20','2026-01-21','2026-01-22','2026-01-23','2026-01-26','2026-01-27',
            '2026-01-28','2026-01-29','2026-01-30']

# ---- equity prices: load ALL 2026 + Dec-31 2025 baseline (per instrument, per date) ----
px_all = defaultdict(dict); dec31_hkd = {}
with open('data/EQUITY_SPOT_PRICES_HIST.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        d = row['valuation_date']
        try: dt = datetime.strptime(d,'%m/%d/%Y').date()
        except: dt = datetime.strptime(d,'%Y-%m-%d').date()
        if dt.year==2026: px_all[row['instrument_code']][dt.isoformat()] = float(row['price'])
        if dt.isoformat()=='2025-12-31': dec31_hkd[row['instrument_code']] = float(row['price'])

fx = {}  # legacy retail-FX dict, kept empty so json schema is unchanged

# ---- CSDC settlement mid FX (data/结算汇兑比率.xlsx): used for Feb + as a generic mid lookup ----
_fxwb = openpyxl.load_workbook('data/结算汇兑比率.xlsx', data_only=True)
csdc_mid = {}   # iso date -> HKD->CNY mid
for _r in list(_fxwb['结算汇兑比率'].iter_rows(values_only=True))[1:]:
    if _r[0] is None or _r[3] != 'HKD': continue
    _d = _r[0].date().isoformat() if hasattr(_r[0],'date') else str(_r[0])[:10]
    csdc_mid[_d] = round((float(_r[1]) + float(_r[2]))/2.0, 5)

if PNL_MONTH == 'jan':
    # ============ JANUARY (unchanged) ============
    start = start_dec31
    trades = [t for t in all_trades if t['month'] == 1]
    PERIOD_DAYS = JAN_DAYS
    codes = sorted(start.keys(), key=lambda x:(classify(x), x))
    px = defaultdict(dict)
    for inst, dd in px_all.items():
        for d, v in dd.items():
            if d[:7] == '2026-01': px[inst][d] = v
    # official implied FX backed out from the attribution file (本年累计)
    wb_off = openpyxl.load_workbook('data/2026年1月(1).xlsx', data_only=True)
    ws = wb_off['本年累计']; rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
    dcc = {}
    for i in range(4, 36):
        v = hdr[i]; s = v.date().isoformat() if hasattr(v,'date') else str(v); dcc[s] = i
    cum = defaultdict(dict); cur = None
    for r in rows[1:]:
        if r[2] is not None: cur = r[2]
        if r[3] is not None: cum[cur][r[3]] = r
    def implied_fx(ref):
        nm=start[ref]['name']; q=start[ref]['qty']
        c0=cum[nm]['持仓成本'][4]*10000.0; fv=cum[nm]['持仓公允']
        return {d:(c0+fv[dcc[d]]*10000.0)/q/px[px_code(ref)][d]
                for d in JAN_DAYS if isinstance(fv[dcc[d]],(int,float)) and px.get(px_code(ref),{}).get(d)}
    fx_ref1 = implied_fx('03988')
    official_fx_dec31 = round(start['00941']['px_cny']/dec31_hkd['0941.HK'], 5)   # = 0.8974
    official_fx = {d: (official_fx_dec31 if d=='2026-01-02' else fx_ref1.get(d, official_fx_dec31)) for d in JAN_DAYS}
    dec31_hkd_out = dec31_hkd

else:
    # ============ ANY MONTH after January (feb/mar/apr/may/...) ============
    # StartPos(M) = prior-month-end position: qty + WAVG cost rolled with ALL trades before
    # month M, marked at the last trading day < month M (= prior month's last close).
    MONTH_NUM = {'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,
                 'sep':9,'oct':10,'nov':11,'dec':12}[PNL_MONTH]
    PREFIX = f'2026-{MONTH_NUM:02d}'
    PERIOD_DAYS = sorted({d for dd in px_all.values() for d in dd if d[:7]==PREFIX})
    if not PERIOD_DAYS:
        raise SystemExit(f'no price days for {PREFIX}')
    FIRST = PERIOD_DAYS[0]
    # BASE = last CSDC/price date strictly before the period start = prior month's last trading day
    prior_px_dates = sorted({d for dd in px_all.values() for d in dd if d < FIRST})
    BASE = prior_px_dates[-1]
    base_fx = csdc_mid.get(BASE) or csdc_mid[max(d for d in csdc_mid if d <= BASE)]
    # roll qty + WAVG cost pool through ALL trades before month M
    pre_tr = [t for t in all_trades if t['month'] < MONTH_NUM]
    start = {}
    for code, s0 in start_dec31.items():
        q = s0['qty'] or 0.0
        cost = s0['cost'] or 0.0           # Dec-31 CNY cost pool
        wac = (cost/q) if q else 0.0
        for t in sorted([x for x in pre_tr if x['code']==code], key=lambda x:x['date']):
            if t['sign']==1:
                q += t['qty']; cost += t['cny_amt']; wac = cost/q if q else 0.0
            else:
                cost -= t['qty']*wac; q -= t['qty']
        inst = px_code(code)
        # BASE mark (CNY): A=close; HK=close_HKD × BASE CSDC mid
        cl = px_all.get(inst,{}).get(BASE)
        if cl is None:
            cand = sorted([d for d in px_all.get(inst,{}) if d<=BASE])
            cl = px_all[inst][cand[-1]] if cand else (s0['px_cny'] or 0.0)
        mark_cny = cl*base_fx if classify(code)=='HK' else cl
        start[code] = dict(code=code, name=s0['name'], qty=q, cost=cost,
                           px_cny=round(mark_cny,4), mv=round(q*mark_cny,2), cls=classify(code))
    trades = [t for t in all_trades if t['month'] == MONTH_NUM]
    codes = sorted(start.keys(), key=lambda x:(classify(x), x))
    px = defaultdict(dict)
    for inst, dd in px_all.items():
        for d, v in dd.items():
            if d[:7] == PREFIX: px[inst][d] = v
    # baseline HKD close per inst at BASE (for HK StartPos 原币 / G col): reuse dec31_hkd slot
    dec31_hkd_out = {}
    for inst, dd in px_all.items():
        if dd.get(BASE) is not None: dec31_hkd_out[inst] = dd[BASE]
    official_fx_dec31 = base_fx             # period baseline FX (prior-month-end mid)
    official_fx = {d: csdc_mid.get(d, base_fx) for d in PERIOD_DAYS}

json.dump(dict(start=start, trades=trades, codes=codes, JAN_DAYS=PERIOD_DAYS,
               px={k:px[k] for k in px}, fx=fx, official_fx=official_fx,
               dec31_hkd=dec31_hkd_out, official_fx_dec31=official_fx_dec31,
               px_code={cd:px_code(cd) for cd in codes}, cls={cd:classify(cd) for cd in codes},
               month=PNL_MONTH),
          open(OUT_JSON,'w',encoding='utf-8'), ensure_ascii=False)
print(f'wrote {OUT_JSON}: month={PNL_MONTH}, {len(codes)} codes, {len(trades)} trades,',
      f'{len(PERIOD_DAYS)} days, base_fx={official_fx_dec31}')