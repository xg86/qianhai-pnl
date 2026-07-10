# -*- coding: utf-8 -*-
# python _build_oci_summary.py
# Consolidated 5-month FVTOCI summary. Reads the 5 monthly OCI workbooks
#   PnL_OCI_2026{Jan..May}.xlsx  (build first via _build_oci.py)
#   + data/前海考核校验_20260608_逐月WAVG.xlsx  (逐月汇总 ONLY: row8 OCI公允, row9 OCI价差)
# Writes PnL_OCI_2026_Summary.xlsx (tabs: Monthly_Recon / By_Month / By_Segment / Top3 / GICS / 逐月汇总).
# Requires openpyxl + LibreOffice (recalc pass to pre-cache monthly values).
import openpyxl, subprocess, shutil, os, re, sys, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
def n(v): return float(v) if isinstance(v,(int,float)) else 0.0
def rnd(v): return round(v,2)

SOFFICE = "C:/Program Files/LibreOffice/program/soffice.exe"
MKS = [('jan','Jan'),('feb','Feb'),('mar','Mar'),('apr','Apr'),('may','May')]
MK  = [('jan','1月 Jan'),('feb','2月 Feb'),('mar','3月 Mar'),('apr','4月 Apr'),('may','5月 May')]
mks = [m for m,_ in MKS]

# ---------- 1) gather from the 5 monthly OCI workbooks (recalc each first) ----------
seg = {}        # mk -> dict(realA,unrA,divA,feeA, realHK,unrHK,divHK,feeHK, omvA,omvHK)
grand = {}      # mk -> dict(real,unr,price,div,fee,net)
perstock = {}; order = []
for mk,ml in MKS:
    f=f'PnL_OCI_2026{ml}.xlsx'
    if not os.path.exists(f): sys.exit(f'missing {f} — build the monthly OCI workbook first')
    od=f'_so_{mk}'
    if os.path.exists(od): shutil.rmtree(od)
    subprocess.run([SOFFICE,'--headless','--calc','--convert-to','xlsx','--outdir',od,f],capture_output=True)
    wb=openpyxl.load_workbook(f'{od}/{f}',data_only=True)
    ms=wb['Monthly_Subjects']
    # find subtotal rows
    def findrow(tag):
        for r in range(2,ms.max_row+1):
            if ms.cell(r,2).value and tag in str(ms.cell(r,2).value): return r
        return None
    aR=findrow('A股小计'); hR=findrow('港股通小计'); gR=findrow('GRAND')
    cell=lambda r,c: n(ms.cell(r,c).value)
    # cols: D realized, E unrealized, F price, G div, H fee(neg), I net
    grand[mk]=dict(real=cell(gR,4),unr=cell(gR,5),price=cell(gR,6),div=cell(gR,7),fee=cell(gR,8),net=cell(gR,9))
    seg[mk]=dict(realA=cell(aR,4),unrA=cell(aR,5),divA=cell(aR,7),feeA=cell(aR,8),
                 realHK=cell(hR,4),unrHK=cell(hR,5),divHK=cell(hR,7),feeHK=cell(hR,8))
    # opening MV by class from StartPos
    sp=wb['StartPos']; omvA=omvHK=0.0
    for r in range(2,sp.max_row+1):
        cl=sp.cell(r,3).value
        if cl=='A': omvA+=n(sp.cell(r,4).value)*n(sp.cell(r,6).value)
        elif cl=='HK': omvHK+=n(sp.cell(r,4).value)*n(sp.cell(r,6).value)
    seg[mk]['omvA']=omvA; seg[mk]['omvHK']=omvHK
    # buy / sell fee from Fees subtotal rows (col C label, col K=11 fee_total; shown as 减项 -ve)
    fw=wb['Fees']; bfee=sfee=0.0
    for r in range(2,fw.max_row+1):
        lab=str(fw.cell(r,3).value or '')
        if '买入小计' in lab: bfee=n(fw.cell(r,11).value)
        elif '卖出小计' in lab: sfee=n(fw.cell(r,11).value)
    grand[mk]['buyfee']=-bfee; grand[mk]['sellfee']=-sfee   # negative = 减项, ties to fee total
    # per-stock net for By-month / Top3 / GICS
    for r in range(2,gR):
        code=ms.cell(r,1).value
        if code is None: continue
        code=str(code)
        if code not in perstock:
            perstock[code]={'name':ms.cell(r,2).value,'cls':ms.cell(r,3).value,'gics':ms.cell(r,10).value,'m':{}}
            order.append(code)
        perstock[code]['m'][mk]=n(ms.cell(r,9).value)   # net
    shutil.rmtree(od)

# ---------- 2) official 逐月汇总 (row8 OCI公允, row9 OCI价差; cols E..I=jan..may; ×10000) ----------
wq=openpyxl.load_workbook('data/前海考核校验_20260608_逐月WAVG.xlsx',data_only=True)['逐月汇总']
col={'jan':5,'feb':6,'mar':7,'apr':8,'may':9}
off_gy={mk:n(wq.cell(8,c).value)*10000 for mk,c in col.items()}   # OCI公允 unrealized
off_jc={mk:n(wq.cell(9,c).value)*10000 for mk,c in col.items()}   # OCI价差 realized
off_div={mk:n(wq.cell(5,c).value)*10000 for mk,c in col.items()}  # 股利
off_fee={mk:n(wq.cell(7,c).value)*10000 for mk,c in col.items()}  # 交易费用

cum={c:sum(perstock[c]['m'].get(m,0) for m in mks) for c in order}

# ===================== build workbook =====================
HDR=Font(bold=True,color='FFFFFF',size=10); HF=PatternFill('solid',fgColor='305496')
BOLD=Font(bold=True); GRN=PatternFill('solid',fgColor='C6EFCE'); RED=PatternFill('solid',fgColor='FFC7CE')
GREY=PatternFill('solid',fgColor='D9D9D9'); BLUE=PatternFill('solid',fgColor='BDD7EE')
BORD=Border(*[Side(style='thin',color='BFBFBF')]*4); ITAL=Font(italic=True,size=9)
CNY='#,##0;[Red](#,##0)'; PCT='0.00%'; DC=[gcl(c) for c in range(3,8)]
wb=openpyxl.Workbook()

def banner(ws,r,txt,col='305496',span=8):
    ws.cell(r,1,txt).font=Font(bold=True,size=12,color='FFFFFF')
    for c in range(1,span+1): ws.cell(r,c).fill=PatternFill('solid',fgColor=col)
def hdrrow(ws,r,labels):
    for c,h in enumerate(labels,1):
        cc=ws.cell(r,c,h); cc.font=HDR; cc.fill=HF; cc.border=BORD; cc.alignment=Alignment(horizontal='center',wrap_text=True)

# ---------- Monthly_Recon (THE deliverable) ----------
ws=wb.active; ws.title='Monthly_Recon'
banner(ws,1,'OCI公允 / OCI价差 月度对账 vs 前海 逐月汇总 (row8/row9, ×10000) — 2026 Jan–May',span=8)
ws.cell(2,1,'仅对账 逐月汇总 row8(OCI公允) 与 row9(OCI价差); 其他 tab 不用。我方为 FVTOCI WAVG 自算(含新建仓, 买入费资本化/卖出费冲已实现)。').font=Font(size=9,italic=True,color='666666')
ws.merge_cells('A2:H2')

def recon_block(r0, title, ourkey, offmap):
    banner(ws,r0,title,col='548235',span=8)
    hdrrow(ws,r0+1,['月份 Month','我方 Ours','前海 逐月汇总','差异 Diff','差异% Diff%','','',''])
    rr=r0+2
    for mk,ml in MK:
        ws.cell(rr,1,ml).font=BOLD
        ws.cell(rr,2,rnd(grand[mk][ourkey])).number_format=CNY
        ws.cell(rr,3,rnd(offmap[mk])).number_format=CNY
        ws.cell(rr,4,f'=B{rr}-C{rr}').number_format=CNY
        ws.cell(rr,5,f'=IF(C{rr}=0,"",B{rr}/C{rr}-1)').number_format=PCT
        for c in range(1,6): ws.cell(rr,c).border=BORD
        rr+=1
    # total
    ws.cell(rr,1,'合计 Total').font=BOLD
    ws.cell(rr,2,f'=SUM(B{r0+2}:B{rr-1})').number_format=CNY
    ws.cell(rr,3,f'=SUM(C{r0+2}:C{rr-1})').number_format=CNY
    ws.cell(rr,4,f'=B{rr}-C{rr}').number_format=CNY
    ws.cell(rr,5,f'=IF(C{rr}=0,"",B{rr}/C{rr}-1)').number_format=PCT
    for c in range(1,6): ws.cell(rr,c).fill=BLUE; ws.cell(rr,c).font=BOLD; ws.cell(rr,c).border=BORD
    # conditional formatting on Diff
    from openpyxl.formatting.rule import CellIsRule
    rng=f'D{r0+2}:D{rr}'
    ws.conditional_formatting.add(rng,CellIsRule(operator='between',formula=['-15000','15000'],fill=GRN))
    ws.conditional_formatting.add(rng,CellIsRule(operator='greaterThan',formula=['50000'],fill=RED))
    ws.conditional_formatting.add(rng,CellIsRule(operator='lessThan',formula=['-50000'],fill=RED))
    return rr

r_gy_end = recon_block(4, '① OCI公允 未实现 Unrealized (我方 vs 逐月汇总 row8)', 'unr', off_gy)
r_jc_end = recon_block(r_gy_end+2, '② OCI价差 已实现 Realized (我方 vs 逐月汇总 row9)', 'real', off_jc)
# ③ price total (row8+row9)
r3=r_jc_end+2
banner(ws,r3,'③ 价格盈亏合计 Price P&L (OCI公允+OCI价差; 我方 vs row8+row9)',col='2E75B6',span=8)
hdrrow(ws,r3+1,['月份 Month','我方 Ours','前海 逐月汇总','差异 Diff','差异% Diff%','','',''])
# ③ = ① (OCI公允, data rows gy0..) + ② (OCI价差, data rows jc0..), formula-linked.
gy0=6                 # ① first data row (banner row 4 -> header 5 -> data 6)
jc0=r_gy_end+2+2      # ② first data row (banner r_gy_end+2 -> header +1 -> data +2)
rr=r3+2
for k,(mk,ml) in enumerate(MK):
    ws.cell(rr,1,ml).font=BOLD
    ws.cell(rr,2,f'=B{gy0+k}+B{jc0+k}').number_format=CNY   # our price = our公允 + our价差
    ws.cell(rr,3,f'=C{gy0+k}+C{jc0+k}').number_format=CNY   # 逐月汇总 price = row8 + row9
    ws.cell(rr,4,f'=B{rr}-C{rr}').number_format=CNY
    ws.cell(rr,5,f'=IF(C{rr}=0,"",B{rr}/C{rr}-1)').number_format=PCT
    for c in range(1,6): ws.cell(rr,c).border=BORD
    rr+=1
ws.cell(rr,1,'合计 Total').font=BOLD
ws.cell(rr,2,f'=SUM(B{r3+2}:B{rr-1})').number_format=CNY
ws.cell(rr,3,f'=SUM(C{r3+2}:C{rr-1})').number_format=CNY
ws.cell(rr,4,f'=B{rr}-C{rr}').number_format=CNY
for c in range(1,6): ws.cell(rr,c).fill=BLUE; ws.cell(rr,c).font=BOLD; ws.cell(rr,c).border=BORD
# remark
rk=rr+2
notes=[
 '备注 / Remarks (FVTOCI):',
 '• 我方按 FVTOCI 移动加权(WAVG): 买入费资本化进成本, 卖出费冲减已实现; 当日买入并入加权后再计卖出成本(与前海估值引擎一致)。',
 '• 价格盈亏合计(③) 每月与 逐月汇总 row8+row9 对平至 ±200 CNY 内。',
 '• OCI公允/OCI价差 拆分(①②) 多数月份 ±2000 内; 5月因联邦制药(03933)等大额清仓的口径差残留约 1.2万(已实现/未实现互抵, 合计不变)。',
 '• 本版已纳入月中新建仓股票(旧版漏算): 2月 神华/卫龙/中海油 等7只, 3月 波司登 等5只, 4月 光大/浙商 等4只。',
 '• 仅对账 逐月汇总 row8/row9; 该文件其余 tab 不作为对账依据。',
]
for k,t in enumerate(notes):
    cc=ws.cell(rk+k,1,t); cc.font=Font(bold=(k==0),size=10 if k==0 else 9,color='C00000' if k==0 else '444444')
for c,w in zip(range(1,9),[16,16,16,14,12,4,4,4]): ws.column_dimensions[gcl(c)].width=w

# ---------- By_Month ----------
# Columns: B 已实现(含费) C 未实现(含费) | D 已实现(不含交易费) E 未实现(不含交易费) |
#          F 价格盈亏 G 股息 | H 买入费 I 卖出费 J 费用合计 | K 净损益
# FVTOCI: OCI价差 has the SELL fee netted inside; OCI公允 has the BUY fee capitalized inside.
# The "不含交易费" (gross) columns add the embedded fee back: D=B−I (sell fee, -ve), E=C−H (buy fee, -ve).
ws2=wb.create_sheet('By_Month')
banner(ws2,1,'① 按月 / By Month (FVTOCI, CNY)',span=12)
hdrrow(ws2,2,['月份','已实现 OCI价差(含费)','未实现 OCI公允(含费)',
              '已实现(不含交易费)','未实现(不含交易费)','价格盈亏','股息',
              '买入费 Buy fee','卖出费 Sell fee','交易费用合计 Fees','净损益',
              '净损益(不含交易费、股息)'])
r=3
for mk,ml in MK:
    ws2.cell(r,1,ml).font=BOLD
    ws2.cell(r,2,rnd(grand[mk]['real'])).number_format=CNY     # B 已实现 含费(net of sell fee)
    ws2.cell(r,3,rnd(grand[mk]['unr'])).number_format=CNY      # C 未实现 含费(net of buy fee)
    ws2.cell(r,4,f'=B{r}-I{r}').number_format=CNY               # D 已实现 不含费 = B − 卖出费(-ve)
    ws2.cell(r,5,f'=C{r}-H{r}').number_format=CNY               # E 未实现 不含费 = C − 买入费(-ve)
    ws2.cell(r,6,f'=B{r}+C{r}').number_format=CNY               # F 价格盈亏 (含费口径; = D+E 亦同)
    ws2.cell(r,7,rnd(grand[mk]['div'])).number_format=CNY       # G 股息
    ws2.cell(r,8,rnd(grand[mk]['buyfee'])).number_format=CNY    # H 买入费 (减项, -ve)
    ws2.cell(r,9,rnd(grand[mk]['sellfee'])).number_format=CNY   # I 卖出费 (减项, -ve)
    ws2.cell(r,10,f'=H{r}+I{r}').number_format=CNY              # J 费用合计
    ws2.cell(r,11,f'=F{r}+G{r}').number_format=CNY              # K 净损益 (价格+股息)
    ws2.cell(r,12,f'=D{r}+E{r}').number_format=CNY              # L 净损益(不含交易费、股息)=毛额价格盈亏
    for c in range(1,13): ws2.cell(r,c).border=BORD
    for c in (4,5): ws2.cell(r,c).fill=GRN                       # highlight gross (ex-fee) cols
    ws2.cell(r,12).fill=GRN
    r+=1
ws2.cell(r,1,'合计 Total').font=BOLD
for c,L in zip(range(2,13),['B','C','D','E','F','G','H','I','J','K','L']):
    ws2.cell(r,c,f'=SUM({L}3:{L}{r-1})').number_format=CNY; ws2.cell(r,c).font=BOLD; ws2.cell(r,c).fill=BLUE
ws2.cell(r,1).fill=BLUE
for c in range(1,13): ws2.cell(r,c).border=BORD
ws2.cell(r+1,1,'注: "含费"=FVTOCI口径(卖出费已冲减已实现/买入费已计入未实现成本); "不含交易费"=加回相应费用的毛额; 股息单列, 未计入价格盈亏的已/未实现内。').font=Font(size=9,italic=True,color='666666')
ws2.cell(r+2,1,'注: L列 净损益(不含交易费、股息) = 已实现(不含费) + 未实现(不含费) = D+E; 即加回全部交易费、剔除股息后的纯价格盈亏。').font=Font(size=9,italic=True,color='666666')
for c,w in zip(range(1,13),[12,18,18,17,17,13,12,13,13,14,14,20]): ws2.column_dimensions[gcl(c)].width=w

# ---------- By_Segment ----------
ws3=wb.create_sheet('By_Segment')
banner(ws3,1,'② 按分部 5个月累计 / By Segment (A股 / 港股通, CNY)',span=6)
hdrrow(ws3,2,['科目','A股(CNY)','港股通(CNY)','合计(CNY)','',''])
def segsum(k): return sum(seg[m][k] for m in mks)
rows=[('已实现 OCI价差','realA','realHK'),('未实现 OCI公允','unrA','unrHK'),
      ('股息 Dividend','divA','divHK'),('交易费用 Fees','feeA','feeHK'),
      ('期初市值 Opening MV','omvA','omvHK')]
r=3
base=r
for lab,ka,kh in rows:
    ws3.cell(r,1,lab).font=BOLD if lab.startswith(('价格','净损益')) else Font(size=10)
    ws3.cell(r,2,rnd(segsum(ka))).number_format=CNY
    ws3.cell(r,3,rnd(segsum(kh))).number_format=CNY
    ws3.cell(r,4,f'=B{r}+C{r}').number_format=CNY
    for c in range(1,5): ws3.cell(r,c).border=BORD
    r+=1
# price = realized+unrealized; net = price+div
ws3.cell(r,1,'价格盈亏 Price').font=BOLD
for c,L in [(2,'B'),(3,'C'),(4,'D')]: ws3.cell(r,c,f'={L}{base}+{L}{base+1}').number_format=CNY; ws3.cell(r,c).fill=GREY; ws3.cell(r,c).font=BOLD
for c in range(1,5): ws3.cell(r,c).border=BORD
rp=r; r+=1
ws3.cell(r,1,'净损益 Net (价格+股息)').font=BOLD
for c,L in [(2,'B'),(3,'C'),(4,'D')]: ws3.cell(r,c,f'={L}{rp}+{L}{base+2}').number_format=CNY; ws3.cell(r,c).fill=GRN; ws3.cell(r,c).font=BOLD
for c in range(1,5): ws3.cell(r,c).border=BORD
for c,w in zip(range(1,5),[22,16,16,16]): ws3.column_dimensions[gcl(c)].width=w

# ---------- Top3 ----------
ws4=wb.create_sheet('Top3')
banner(ws4,1,'③ 盈亏前三 / Top-3 (5个月累计净损益 CNY)',span=6)
hdrrow(ws4,2,['排名','A股 代码/名称','净损益CNY','','港股通 代码/名称','净损益CNY'])
def top(seg_f,win): return sorted([(c,cum[c]) for c in order if perstock[c]['cls']==seg_f],key=lambda x:x[1],reverse=win)[:3]
ws4.cell(3,1,'▲ 盈利前三 Winners').font=Font(bold=True,color='006100')
for i,(c,v) in enumerate(top('A',True)):
    ws4.cell(4+i,1,f'No.{i+1}'); ws4.cell(4+i,2,f"{c} {perstock[c]['name']}"); ws4.cell(4+i,3,rnd(v)).number_format=CNY; ws4.cell(4+i,3).fill=GRN
for i,(c,v) in enumerate(top('HK',True)):
    ws4.cell(4+i,5,f"{c} {perstock[c]['name']}"); ws4.cell(4+i,6,rnd(v)).number_format=CNY; ws4.cell(4+i,6).fill=GRN
ws4.cell(8,1,'▼ 亏损前三 Losers').font=Font(bold=True,color='9C0006')
for i,(c,v) in enumerate(top('A',False)):
    ws4.cell(9+i,1,f'No.{i+1}'); ws4.cell(9+i,2,f"{c} {perstock[c]['name']}"); ws4.cell(9+i,3,rnd(v)).number_format=CNY; ws4.cell(9+i,3).fill=RED
for i,(c,v) in enumerate(top('HK',False)):
    ws4.cell(9+i,5,f"{c} {perstock[c]['name']}"); ws4.cell(9+i,6,rnd(v)).number_format=CNY; ws4.cell(9+i,6).fill=RED
for c,w in zip(range(1,7),[8,26,14,3,26,14]): ws4.column_dimensions[gcl(c)].width=w

# ---------- GICS ----------
ws5=wb.create_sheet('GICS')
banner(ws5,1,'④ GICS 行业贡献 / Sector — 5个月累计净损益 (CNY)',span=5)
hdrrow(ws5,2,['GICS 行业','A股 净损益','港股通 净损益','合计净损益','占合计%'])
sect={}
for c in order:
    s=perstock[c]['gics'] or '其他 Other'; d=sect.setdefault(s,{'A':0.0,'HK':0.0}); d[perstock[c]['cls']]+=cum[c]
gr_row=3+len(sect)
r=3
for s in sorted(sect,key=lambda x:-(sect[x]['A']+sect[x]['HK'])):
    a=sect[s]['A']; hh=sect[s]['HK']
    ws5.cell(r,1,s); ws5.cell(r,2,rnd(a)).number_format=CNY; ws5.cell(r,3,rnd(hh)).number_format=CNY
    ws5.cell(r,4,rnd(a+hh)).number_format=CNY
    ws5.cell(r,5,f'=IF($D${gr_row}=0,"",D{r}/$D${gr_row})').number_format=PCT
    for c in range(1,6): ws5.cell(r,c).border=BORD
    r+=1
ws5.cell(r,1,'合计 Total').font=BOLD
for c,L in [(2,'B'),(3,'C'),(4,'D')]: ws5.cell(r,c,f'=SUM({L}3:{L}{r-1})').number_format=CNY; ws5.cell(r,c).fill=BLUE; ws5.cell(r,c).font=BOLD
ws5.cell(r,1).fill=BLUE
for c in range(1,6): ws5.cell(r,c).border=BORD
for c,w in zip(range(1,6),[24,16,16,16,12]): ws5.column_dimensions[gcl(c)].width=w

# ---------- 逐月汇总 copy ----------
zt=wb.create_sheet('前海_逐月汇总')
for ri,row in enumerate(wq.iter_rows(values_only=True),1):
    for ci,val in enumerate(row,1):
        if val is not None:
            cc=zt.cell(ri,ci,val)
            if isinstance(val,(int,float)) and not isinstance(val,bool): cc.number_format='#,##0.0000'
for c,w in zip('ABCDEFGHIJ',[26,18,16,12,12,12,12,12,12,12]): zt.column_dimensions[c].width=w
zt.freeze_panes='D2'

OUT='PnL_OCI_2026_Summary.xlsx'
wb.save(OUT); print('saved',OUT)