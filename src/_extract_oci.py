# -*- coding: utf-8 -*-
# python _extract_oci.py               (January, default)
# PNL_MONTH=feb python _extract_oci.py (February) ... may
# OCI variant of _extract.py. TWO differences from the original:
#   1. Universe = opening holdings UNION any code traded in-month (new mid-month buys get a
#      synthetic qty=0/cost=0 opening row). Post-Jan ALSO rolls codes first bought in a prior
#      month (held into this month) — the original froze the universe to the Dec-31 book.
#   2. January additionally extracts the OFFICIAL daily OCI公允 / OCI价差 per stock per day
#      (from data/2026年1月(1).xlsx, 本年累计, 万元×10000) for the daily-OCI reconciliation.
# Writes _oci_<month>.json (consumed by _build_oci.py). The original _extract.py is untouched.
import openpyxl, csv, sys, io, json, os
from collections import defaultdict
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PNL_MONTH = os.environ.get('PNL_MONTH', 'may').lower()
OUT_JSON = f'_oci_{PNL_MONTH}.json'

# fee rate model (must match _build_oci.py Fees sheet, so the rolled WAVG basis is consistent)
A_STAMP_BPS, A_COMM_BPS, HK_STAMP_BPS, HK_OTHER_BPS = 5.0, 0.641, 10.0, 1.362

def classify(code): return 'HK' if len(str(code)) == 5 else 'A'
def px_code(code):
    code = str(code)
    if classify(code) == 'HK': return str(int(code)).zfill(4) + '.HK'
    return code + ('.SH' if code.startswith('6') else '.SZ')

# ---- start positions (期初持仓 = Dec-31 book) ----
wb = openpyxl.load_workbook('data/start position.xlsx', data_only=True); ws = wb.active
start_dec31 = {}
for r in list(ws.iter_rows(values_only=True))[1:]:
    if r[0] is None: continue
    code = str(r[0])
    start_dec31[code] = dict(code=code, name=r[1], qty=r[2], cost=r[3], px_cny=r[4], mv=r[5], cls=classify(code))

# ---- all 2026 trades ----
wb = openpyxl.load_workbook('data/transaction hist.xlsx', data_only=True); ws = wb.active
rows = list(ws.iter_rows(values_only=True)); h = rows[0]
def c(n): return h.index(n)
all_trades = []
for r in rows[1:]:
    if r[c('证券代码')] is None: continue
    d = r[c('发生日期')]
    dt = datetime.strptime(d, '%Y-%m-%d').date() if isinstance(d, str) else d.date()
    if dt.year != 2026: continue
    code = str(r[c('证券代码')]); sign = 1 if r[c('委托方向')] == '买入' else -1
    qty = r[c('成交数量')]
    all_trades.append(dict(date=dt.isoformat(), month=dt.month, code=code, name=r[c('证券名称')],
                           dir=r[c('委托方向')], sign=sign, qty=qty, signed_qty=sign*qty,
                           cny_amt=r[c('本币成交金额')], trade_amt=r[c('成交金额')],
                           cls=classify(code), mkt=r[c('交易市场')]))

# name lookup (for synthetic opening rows of new codes)
name_by_code = {cd: s['name'] for cd, s in start_dec31.items()}
for t in all_trades:
    name_by_code.setdefault(t['code'], t['name'])

JAN_DAYS = ['2026-01-02','2026-01-05','2026-01-06','2026-01-07','2026-01-08','2026-01-09',
            '2026-01-12','2026-01-13','2026-01-14','2026-01-15','2026-01-16','2026-01-19',
            '2026-01-20','2026-01-21','2026-01-22','2026-01-23','2026-01-26','2026-01-27',
            '2026-01-28','2026-01-29','2026-01-30']

# ---- equity prices: load ALL 2026 + Dec-31 2025 baseline (per instrument, per date) ----
px_all = defaultdict(dict); dec31_hkd = {}
with open('data/EQUITY_SPOT_PRICES_HIST.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        d = row['valuation_date']
        try: dt = datetime.strptime(d,'%m/%d/%Y').date()
        except: dt = datetime.strptime(d,'%Y-%m-%d').date()
        if dt.year==2026: px_all[row['instrument_code']][dt.isoformat()] = float(row['price'])
        if dt.isoformat()=='2025-12-31': dec31_hkd[row['instrument_code']] = float(row['price'])

fx = {}  # legacy retail-FX dict, kept empty so json schema is unchanged

# ---- CSDC settlement mid FX (data/结算汇兑比率.xlsx) ----
_fxwb = openpyxl.load_workbook('data/结算汇兑比率.xlsx', data_only=True)
csdc_mid = {}
for _r in list(_fxwb['结算汇兑比率'].iter_rows(values_only=True))[1:]:
    if _r[0] is None or _r[3] != 'HKD': continue
    _d = _r[0].date().isoformat() if hasattr(_r[0],'date') else str(_r[0])[:10]
    csdc_mid[_d] = round((float(_r[1]) + float(_r[2]))/2.0, 5)

def base_mark_cny(code, base_iso, base_fx):
    """Best-effort baseline CNY mark for a code (used for StartPos col F). Only matters for
    rows with qty_start>0; new (qty=0) codes get a sensible non-zero level for price continuity."""
    inst = px_code(code)
    cl = px_all.get(inst, {}).get(base_iso)
    if cl is None:
        cand = sorted([d for d in px_all.get(inst, {}) if d <= base_iso])
        if cand: cl = px_all[inst][cand[-1]]
        else:
            cand2 = sorted(px_all.get(inst, {}))
            cl = px_all[inst][cand2[0]] if cand2 else 0.0
    return cl*base_fx if classify(code)=='HK' else cl

# ================================================================= official daily OCI (Jan only)
off_oci_daily = {}   # code -> {iso_date: OCI公允 CNY}
off_jc_daily  = {}   # code -> {iso_date: OCI价差 CNY}
off_oci_month = {}   # code -> Σ all daily OCI公允 cols (CNY)  [complete monthly official]
off_jc_month  = {}   # code -> Σ all daily OCI价差 cols (CNY)
if PNL_MONTH == 'jan':
    wb_off = openpyxl.load_workbook('data/2026年1月(1).xlsx', data_only=True)
    ws_o = wb_off['本年累计']
    orows = list(ws_o.iter_rows(values_only=True)); ohdr = orows[0]
    odate = {}
    for i in range(5, 36):
        v = ohdr[i]
        s = v.date().isoformat() if hasattr(v, 'date') else (v if isinstance(v, str) else str(v))
        odate[i] = s
    code_by_name = {s['name']: cd for cd, s in start_dec31.items()}
    for t in all_trades:
        code_by_name.setdefault(t['name'], t['code'])
    cur = None
    for r in orows[1:]:
        if r[2] is not None: cur = r[2]
        if r[3] is None or cur is None: continue
        subj = r[3]
        if subj not in ('OCI公允', 'OCI价差'): continue
        code = code_by_name.get(cur)
        if code is None: continue
        dd = off_oci_daily if subj == 'OCI公允' else off_jc_daily
        mtot = 0.0; series = dd.setdefault(code, {})
        for i in range(5, 36):
            v = r[i]
            if isinstance(v, (int, float)):
                series[odate[i]] = series.get(odate[i], 0.0) + v*10000.0
                mtot += v*10000.0
        if subj == 'OCI公允': off_oci_month[code] = off_oci_month.get(code, 0.0) + mtot
        else:                 off_jc_month[code]  = off_jc_month.get(code, 0.0) + mtot

# ================================================================= build universe + start rows
if PNL_MONTH == 'jan':
    PERIOD_DAYS = JAN_DAYS
    trades = [t for t in all_trades if t['month'] == 1]
    base_iso = '2025-12-31'
    # official implied FX (Jan), backed out from the attribution file (本年累计) — unchanged logic
    wb_off = openpyxl.load_workbook('data/2026年1月(1).xlsx', data_only=True)
    wsf = wb_off['本年累计']; frows = list(wsf.iter_rows(values_only=True)); fhdr = frows[0]
    dcc = {}
    for i in range(4, 36):
        v = fhdr[i]; s = v.date().isoformat() if hasattr(v,'date') else str(v); dcc[s] = i
    cum = defaultdict(dict); cur = None
    for r in frows[1:]:
        if r[2] is not None: cur = r[2]
        if r[3] is not None: cum[cur][r[3]] = r
    pxJ = defaultdict(dict)
    for inst, dd in px_all.items():
        for d, v in dd.items():
            if d[:7] == '2026-01': pxJ[inst][d] = v
    def implied_fx(ref):
        nm=start_dec31[ref]['name']; q=start_dec31[ref]['qty']
        c0=cum[nm]['持仓成本'][4]*10000.0; fv=cum[nm]['持仓公允']
        return {d:(c0+fv[dcc[d]]*10000.0)/q/pxJ[px_code(ref)][d]
                for d in JAN_DAYS if isinstance(fv[dcc[d]],(int,float)) and pxJ.get(px_code(ref),{}).get(d)}
    fx_ref1 = implied_fx('03988')
    official_fx_dec31 = round(start_dec31['00941']['px_cny']/dec31_hkd['0941.HK'], 5)   # 0.8974
    official_fx = {d: (official_fx_dec31 if d=='2026-01-02' else fx_ref1.get(d, official_fx_dec31)) for d in JAN_DAYS}
    base_fx = official_fx_dec31
    # universe = Dec-31 holdings ∪ codes traded in Jan
    universe = set(start_dec31) | {t['code'] for t in trades}
    start = {}
    for code in universe:
        if code in start_dec31:
            start[code] = dict(start_dec31[code])
        else:   # new buy in Jan, no opening holding
            start[code] = dict(code=code, name=name_by_code.get(code, code), qty=0.0, cost=0.0,
                               px_cny=round(base_mark_cny(code, base_iso, base_fx), 4), mv=0.0,
                               cls=classify(code))
    codes = sorted(start.keys(), key=lambda x:(classify(x), x))
    px = pxJ
    dec31_hkd_out = dec31_hkd

else:
    MONTH_NUM = {'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,
                 'sep':9,'oct':10,'nov':11,'dec':12}[PNL_MONTH]
    PREFIX = f'2026-{MONTH_NUM:02d}'
    PERIOD_DAYS = sorted({d for dd in px_all.values() for d in dd if d[:7]==PREFIX})
    if not PERIOD_DAYS: raise SystemExit(f'no price days for {PREFIX}')
    FIRST = PERIOD_DAYS[0]
    prior_px_dates = sorted({d for dd in px_all.values() for d in dd if d < FIRST})
    BASE = prior_px_dates[-1]
    base_fx = csdc_mid.get(BASE) or csdc_mid[max(d for d in csdc_mid if d <= BASE)]
    pre_tr = [t for t in all_trades if t['month'] < MONTH_NUM]
    trades = [t for t in all_trades if t['month'] == MONTH_NUM]
    # universe = Dec-31 holdings ∪ all codes traded BEFORE month M (rolled in) ∪ codes traded IN M
    universe = set(start_dec31) | {t['code'] for t in pre_tr} | {t['code'] for t in trades}
    start = {}
    for code in universe:
        s0 = start_dec31.get(code)
        q = (s0['qty'] if s0 else 0.0) or 0.0
        cost = (s0['cost'] if s0 else 0.0) or 0.0
        wac = (cost/q) if q else 0.0
        # Roll prior-month trades day-ordered. BUY fee is CAPITALIZED into the WAVG cost pool
        # (FVTOCI), exactly as the in-month CostPool sheet does — else rolled-in names carry a
        # cost basis short by accumulated buy fees, shifting P&L from unrealized into realized.
        # Buy fee bps: A = comm only (0.641, no stamp on buys); HK = stamp 10 + other 1.362.
        _bf_bps = A_COMM_BPS if classify(code)=='A' else (HK_STAMP_BPS+HK_OTHER_BPS)
        for t in sorted([x for x in pre_tr if x['code']==code], key=lambda x:x['date']):
            if t['sign']==1:
                buy_fee = t['cny_amt']*_bf_bps/10000.0
                q += t['qty']; cost += t['cny_amt'] + buy_fee; wac = cost/q if q else 0.0
            else:
                cost -= t['qty']*wac; q -= t['qty']
        mark_cny = base_mark_cny(code, BASE, base_fx)
        start[code] = dict(code=code, name=name_by_code.get(code, code), qty=q, cost=cost,
                           px_cny=round(mark_cny,4), mv=round(q*mark_cny,2), cls=classify(code))
    # keep only codes that are held at month start OR traded this month (drop dormant zeros)
    active_in_month = {t['code'] for t in trades}
    start = {cd: s for cd, s in start.items()
             if abs(s['qty'] or 0.0) > 1e-9 or cd in active_in_month}
    codes = sorted(start.keys(), key=lambda x:(classify(x), x))
    px = defaultdict(dict)
    for inst, dd in px_all.items():
        for d, v in dd.items():
            if d[:7] == PREFIX: px[inst][d] = v
    dec31_hkd_out = {}
    for inst, dd in px_all.items():
        if dd.get(BASE) is not None: dec31_hkd_out[inst] = dd[BASE]
    official_fx_dec31 = base_fx
    official_fx = {d: csdc_mid.get(d, base_fx) for d in PERIOD_DAYS}

json.dump(dict(start=start, trades=trades, codes=codes, JAN_DAYS=PERIOD_DAYS,
               px={k:px[k] for k in px}, fx=fx, official_fx=official_fx,
               dec31_hkd=dec31_hkd_out, official_fx_dec31=official_fx_dec31,
               px_code={cd:px_code(cd) for cd in codes}, cls={cd:classify(cd) for cd in codes},
               month=PNL_MONTH,
               off_oci_daily=off_oci_daily, off_jc_daily=off_jc_daily,
               off_oci_month=off_oci_month, off_jc_month=off_jc_month),
          open(OUT_JSON,'w',encoding='utf-8'), ensure_ascii=False)
_new = [cd for cd in codes if (start[cd]['qty'] or 0.0)==0.0 and any(t['code']==cd for t in trades)]
print(f'wrote {OUT_JSON}: month={PNL_MONTH}, {len(codes)} codes, {len(trades)} trades, '
      f'{len(PERIOD_DAYS)} days, base_fx={official_fx_dec31}')
print(f'  new mid-month positions ({len(_new)}): {_new}')
if PNL_MONTH=='jan':
    print(f'  official OCI公允 codes={len(off_oci_month)}, OCI价差 codes={len(off_jc_month)}')
